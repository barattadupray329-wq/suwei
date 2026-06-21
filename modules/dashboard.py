#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""财务统计首页，聚焦日/月/年维度的收入与成本分析。"""

import tkinter as tk
from tkinter import ttk
from datetime import date, datetime

from theme.colors import DarkTheme


class DashboardFrame(ttk.Frame):
    def __init__(self, parent, data_manager):
        super().__init__(parent)
        self.data_manager = data_manager
        self.clock_label = None
        self.start_var = tk.StringVar()
        self.end_var = tk.StringVar()
        self.configure(style="Main.TFrame")
        self._build()
        self._update_clock()

    def _action_button(self, parent, text, command, color):
        btn = tk.Button(parent, text=text, font=DarkTheme.FONT_SMALL, fg="white", bg=color,
                        relief=tk.FLAT, cursor="hand2", command=command, padx=12, pady=5)
        btn.pack(side=tk.LEFT, padx=3)
        DarkTheme.bind_hover(btn, color)
        return btn

    def _build(self):
        self.records = self.data_manager.get_records()
        self.main = tk.Frame(self, bg=DarkTheme.BG_PRIMARY)
        self.main.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        self.main.grid_columnconfigure(0, weight=1)

        head = tk.Frame(self.main, bg=DarkTheme.BG_PRIMARY)
        head.pack(fill=tk.X, pady=(0, 18))
        left = tk.Frame(head, bg=DarkTheme.BG_PRIMARY)
        left.pack(side=tk.LEFT, fill=tk.X, expand=True)
        tk.Label(left, text="💰 财务统计", font=DarkTheme.FONT_TITLE, fg=DarkTheme.TEXT_PRIMARY, bg=DarkTheme.BG_PRIMARY).pack(anchor=tk.W)
        tk.Label(left, text="按日 / 月 / 年查看收入、成本、净额与资金流水", font=DarkTheme.FONT_SMALL,
                 fg=DarkTheme.TEXT_MUTED, bg=DarkTheme.BG_PRIMARY).pack(anchor=tk.W, pady=(4, 0))

        right = tk.Frame(head, bg=DarkTheme.BG_PRIMARY)
        right.pack(side=tk.RIGHT)
        self.clock_label = tk.Label(right, text="", font=DarkTheme.FONT_LABEL, fg=DarkTheme.ACCENT_CYAN, bg=DarkTheme.BG_PRIMARY)
        self.clock_label.pack(anchor=tk.E, pady=(0, 6))
        btn_row = tk.Frame(right, bg=DarkTheme.BG_PRIMARY)
        btn_row.pack(anchor=tk.E)
        self._action_button(btn_row, "🔄 刷新", self._refresh, DarkTheme.ACCENT_BLUE)
        self._action_button(btn_row, "📊 导出报表", self._export_report, DarkTheme.ACCENT_GREEN)

        self._build_filters()
        self._build_period_cards()

    def _build_period_cards(self):
        top = tk.Frame(self.main, bg=DarkTheme.BG_PRIMARY)
        top.pack(fill=tk.X, pady=(0, 12))
        today = date.today()
        month_start = today.replace(day=1)
        year_start = date(today.year, 1, 1)
        ranges = [
            ("今日", (today, today), DarkTheme.ACCENT_CYAN),
            ("本月", (month_start, today), DarkTheme.ACCENT_GREEN),
            ("本年", (year_start, today), DarkTheme.ACCENT_PURPLE),
        ]
        for idx, (title, dr, color) in enumerate(ranges):
            income = self._stats_sum(date_range=dr)
            cost = self._stats_sum(date_range=dr, cost=True)
            net = income - cost
            box = tk.Frame(top, bg=DarkTheme.BG_CARD, highlightbackground=DarkTheme.BORDER_COLOR, highlightthickness=1)
            box.grid(row=0, column=idx, sticky="nsew", padx=6, pady=6)
            top.grid_columnconfigure(idx, weight=1)
            tk.Label(box, text=f"{title}财务", font=DarkTheme.FONT_SUBTITLE, fg=color, bg=DarkTheme.BG_CARD).pack(anchor=tk.W, padx=14, pady=(12, 6))
            for label, value, vcolor in [("收入", income, DarkTheme.ACCENT_GREEN), ("成本", cost, DarkTheme.ACCENT_RED), ("净额", net, DarkTheme.ACCENT_BLUE)]:
                row = tk.Frame(box, bg=DarkTheme.BG_CARD)
                row.pack(fill=tk.X, padx=14, pady=4)
                tk.Label(row, text=label, font=DarkTheme.FONT_LABEL, fg=DarkTheme.TEXT_SECONDARY, bg=DarkTheme.BG_CARD).pack(side=tk.LEFT)
                tk.Label(row, text=f"¥{value:,.2f}", font=("微软雅黑", 14, "bold"), fg=vcolor, bg=DarkTheme.BG_CARD).pack(side=tk.RIGHT)
            tk.Label(box, text=f"统计区间：{dr[0]} 至 {dr[1]}", font=DarkTheme.FONT_SMALL, fg=DarkTheme.TEXT_MUTED, bg=DarkTheme.BG_CARD).pack(anchor=tk.W, padx=14, pady=(4, 12))

    def _build_filters(self):
        box = tk.Frame(self.main, bg=DarkTheme.BG_CARD, highlightbackground=DarkTheme.BORDER_COLOR, highlightthickness=1)
        box.pack(fill=tk.X, pady=(0, 12))
        tk.Label(box, text="日期筛选", font=DarkTheme.FONT_SUBTITLE, fg=DarkTheme.ACCENT_CYAN, bg=DarkTheme.BG_CARD).pack(anchor=tk.W, padx=16, pady=(12, 8))
        row = tk.Frame(box, bg=DarkTheme.BG_CARD)
        row.pack(fill=tk.X, padx=16, pady=(0, 12))
        today = date.today()
        self.start_var.set(today.replace(day=1).strftime("%Y-%m-%d"))
        self.end_var.set(today.strftime("%Y-%m-%d"))
        tk.Label(row, text="开始日期", font=DarkTheme.FONT_LABEL, fg=DarkTheme.TEXT_SECONDARY, bg=DarkTheme.BG_CARD).pack(side=tk.LEFT)
        ttk.Entry(row, textvariable=self.start_var, width=14).pack(side=tk.LEFT, padx=(6, 12))
        tk.Label(row, text="结束日期", font=DarkTheme.FONT_LABEL, fg=DarkTheme.TEXT_SECONDARY, bg=DarkTheme.BG_CARD).pack(side=tk.LEFT)
        ttk.Entry(row, textvariable=self.end_var, width=14).pack(side=tk.LEFT, padx=(6, 12))
        self._action_button(row, "查询", self._refresh, DarkTheme.ACCENT_BLUE)
        self._action_button(row, "本月", self._set_month_range, DarkTheme.ACCENT_GREEN)
        self._action_button(row, "本年", self._set_year_range, DarkTheme.ACCENT_PURPLE)
        self._action_button(row, "历史", self._set_history_range, DarkTheme.BG_HOVER)

    def _parse_date(self, s):
        try:
            return datetime.strptime(s.strip(), "%Y-%m-%d").date()
        except Exception:
            return None

    def _range_from_vars(self):
        s = self._parse_date(self.start_var.get()) or date.min
        e = self._parse_date(self.end_var.get()) or date.max
        if s > e:
            s, e = e, s
        return s, e

    def _period_match(self, dt, start, end):
        return start <= dt <= end

    def _record_date(self, r):
        for k in ("register_date", "created_at", "updated_at"):
            v = r.get(k)
            if v:
                try:
                    return datetime.strptime(v[:10], "%Y-%m-%d").date()
                except Exception:
                    pass
        lease = r.get("lease_info", {}) or {}
        for k in ("start_date", "end_date"):
            v = lease.get(k)
            if v:
                try:
                    return datetime.strptime(v[:10], "%Y-%m-%d").date()
                except Exception:
                    pass
        return None

    def _rent_of(self, r):
        return float((r.get("lease_info", {}) or {}).get("total_rent", 0) or 0)

    def _cost_of(self, r):
        qty = float(r.get("quantity", 1) or 1)
        lease = r.get("lease_info", {}) or {}
        hw = r.get("hardware", {}) or {}
        items = hw.get("items") if isinstance(hw, dict) else []
        if not items and isinstance(hw, dict):
            items = [hw]
        if not items:
            # 兼容老结构：按记录里可见字段做一个粗略成本
            return float(r.get("estimated_cost", 0) or 0)
        total = 0.0
        for it in items:
            total += float(it.get("quantity", 1) or 1) * float(it.get("unit_cost", 0) or 0)
        return total

    def _payment_source_text(self, record, item):
        lease = record.get("lease_info", {}) or {}
        hardware = record.get("hardware", {}) or {}
        renter = record.get("renter", {}) or {}
        parts = []
        source_type = str(item.get("source_type", "")).strip()
        source_name = str(item.get("source_name", "")).strip()
        source = str(item.get("source", "")).strip()
        if source_type:
            parts.append(source_type)
        if source_name:
            parts.append(source_name)
        if source and source not in parts:
            parts.append(source)
        if not parts:
            if item.get("method"):
                parts.append(str(item.get("method")))
            if lease.get("lease_type"):
                parts.append(str(lease.get("lease_type")))
            if hardware.get("category"):
                parts.append(str(hardware.get("category")))
            if hardware.get("brand"):
                parts.append(str(hardware.get("brand")))
            if hardware.get("model_name"):
                parts.append(str(hardware.get("model_name")))
            if renter.get("name"):
                parts.append(f"租户:{renter.get('name')}")
        return " / ".join([p for p in parts if p]) or "未记录"

    def _payment_events(self, records=None):
        records = records or self.records
        events = []
        for r in records:
            lease = r.get("lease_info", {}) or {}
            history = r.get("payment_history", []) or []
            renter_name = (r.get("renter", {}) or {}).get("name", "")
            rental_no = r.get("id", "")
            total_rent = float(lease.get("total_rent", 0) or 0)
            paid_so_far = 0.0

            if not history:
                paid_amount = float(r.get("paid_amount", 0) or 0)
                if paid_amount > 0:
                    events.append({
                        "record_id": rental_no,
                        "customer_name": renter_name,
                        "date": r.get("updated_at") or r.get("register_date", ""),
                        "amount": paid_amount,
                        "unpaid_amount": max(total_rent - paid_amount, 0),
                        "total_rent": total_rent,
                        "end_date": lease.get("end_date", ""),
                        "method": r.get("payment_method", ""),
                        "source_type": "历史补录",
                        "source_name": self.data_manager._derive_payment_source(r, {
                            "method": r.get("payment_method", ""),
                            "note": "历史数据补录",
                        }) if hasattr(self, "data_manager") else r.get("payment_method", ""),
                        "source": self._payment_source_text(r, {
                            "method": r.get("payment_method", ""),
                            "note": "历史数据补录",
                        }),
                        "note": "历史数据补录",
                        "operator": r.get("operator", "系统"),
                    })

            for item in history:
                amt = float(item.get("amount", 0) or 0)
                if amt <= 0:
                    continue
                paid_so_far += amt
                events.append({
                    "record_id": rental_no,
                    "customer_name": renter_name,
                    "date": item.get("payment_date", ""),
                    "amount": amt,
                    "unpaid_amount": max(total_rent - paid_so_far, 0),
                    "total_rent": total_rent,
                    "end_date": lease.get("end_date", ""),
                    "method": item.get("method", ""),
                    "source_type": item.get("source_type", ""),
                    "source_name": item.get("source_name", ""),
                    "source": self._payment_source_text(r, item),
                    "note": item.get("note", ""),
                    "operator": item.get("operator", ""),
                })
        events.sort(key=lambda x: x.get("date", ""), reverse=True)
        return events

    def _stats_sum(self, predicate=None, date_range=None, cost=False):
        start, end = date_range if date_range else (date.min, date.max)
        total = 0.0
        for r in self.records:
            dt = self._record_date(r)
            if dt is None or not (start <= dt <= end):
                continue
            if predicate and not predicate(r):
                continue
            total += self._cost_of(r) if cost else self._rent_of(r)
        return total

    def _today_range(self):
        today = date.today()
        return today, today

    def _build_summary_cards(self):
        top = tk.Frame(self.main, bg=DarkTheme.BG_PRIMARY)
        top.pack(fill=tk.X, pady=(0, 12))
        start, end = self._range_from_vars()
        month_start = date.today().replace(day=1)
        q = (date.today().month - 1) // 3 + 1
        q_start_month = (q - 1) * 3 + 1
        quarter_start = date(date.today().year, q_start_month, 1)
        year_start = date(date.today().year, 1, 1)

        today = date.today()
        today_range = self._today_range()
        groups = [
            ("收入统计", DarkTheme.ACCENT_GREEN, [
                ("今日总租金", self._stats_sum(date_range=today_range), DarkTheme.ACCENT_CYAN),
                ("本月总租金", self._stats_sum(date_range=(month_start, today)), DarkTheme.ACCENT_BLUE),
                ("本季总租金", self._stats_sum(date_range=(quarter_start, today)), DarkTheme.ACCENT_GREEN),
                ("本年总租金", self._stats_sum(date_range=(year_start, today)), DarkTheme.ACCENT_YELLOW),
                ("历史合计总租金", self._stats_sum(), DarkTheme.ACCENT_CYAN),
            ]),
            ("成本统计", DarkTheme.ACCENT_RED, [
                ("今日成本统计", self._stats_sum(date_range=today_range, cost=True), DarkTheme.ACCENT_CYAN),
                ("本月成本统计", self._stats_sum(date_range=(month_start, today), cost=True), DarkTheme.STATUS_ACTIVE),
                ("本季度成本统计", self._stats_sum(date_range=(quarter_start, today), cost=True), DarkTheme.STATUS_EXPIRED),
                ("本年度成本统计", self._stats_sum(date_range=(year_start, today), cost=True), DarkTheme.STATUS_RETURNED),
                ("在租成本", self._stats_sum(lambda r: r.get("status") == "在租", cost=True), DarkTheme.ACCENT_GREEN),
                ("逾期成本", self._stats_sum(lambda r: r.get("status") == "已逾期", cost=True), DarkTheme.ACCENT_RED),
                ("丢失成本", self._stats_sum(lambda r: r.get("status") == "已丢失", cost=True), DarkTheme.ACCENT_PURPLE),
            ]),
            ("当前筛选", DarkTheme.ACCENT_PURPLE, [
                ("选择日期总租金", self._stats_sum(date_range=(start, end)), DarkTheme.ACCENT_PURPLE),
                ("选择日期总成本", self._stats_sum(date_range=(start, end), cost=True), DarkTheme.ACCENT_RED),
                ("筛选净额", self._stats_sum(date_range=(start, end)) - self._stats_sum(date_range=(start, end), cost=True), DarkTheme.ACCENT_GREEN),
                ("筛选条件", f"{start} 至 {end}", DarkTheme.ACCENT_CYAN),
            ]),
        ]

        for idx, (gtitle, gcolor, items) in enumerate(groups):
            box = tk.Frame(top, bg=DarkTheme.BG_CARD, highlightbackground=DarkTheme.BORDER_COLOR, highlightthickness=1)
            box.grid(row=0, column=idx, sticky="nsew", padx=6, pady=6)
            top.grid_columnconfigure(idx, weight=1)
            tk.Label(box, text=gtitle, font=DarkTheme.FONT_SUBTITLE, fg=gcolor, bg=DarkTheme.BG_CARD).pack(anchor=tk.W, padx=14, pady=(12, 8))
            inner = tk.Frame(box, bg=DarkTheme.BG_CARD)
            inner.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 12))
            for r_i, (title, value, color) in enumerate(items):
                row = tk.Frame(inner, bg=DarkTheme.BG_CARD)
                row.pack(fill=tk.X, pady=4)
                tk.Label(row, text=title, font=DarkTheme.FONT_LABEL, fg=DarkTheme.TEXT_SECONDARY, bg=DarkTheme.BG_CARD, wraplength=160, justify=tk.LEFT).pack(side=tk.LEFT)
                disp = value if isinstance(value, str) else f"¥{value:,.2f}"
                tk.Label(row, text=disp, font=("微软雅黑", 14, "bold"), fg=color, bg=DarkTheme.BG_CARD).pack(side=tk.RIGHT)

    def _set_month_range(self):
        d = date.today().replace(day=1)
        self.start_var.set(d.strftime("%Y-%m-%d"))
        self.end_var.set(date.today().strftime("%Y-%m-%d"))
        self._refresh()

    def _set_quarter_range(self):
        today = date.today()
        q = (today.month - 1) // 3 + 1
        m = (q - 1) * 3 + 1
        self.start_var.set(date(today.year, m, 1).strftime("%Y-%m-%d"))
        self.end_var.set(today.strftime("%Y-%m-%d"))
        self._refresh()

    def _set_year_range(self):
        today = date.today()
        self.start_var.set(date(today.year, 1, 1).strftime("%Y-%m-%d"))
        self.end_var.set(today.strftime("%Y-%m-%d"))
        self._refresh()

    def _set_history_range(self):
        self.start_var.set("1970-01-01")
        self.end_var.set(date.today().strftime("%Y-%m-%d"))
        self._refresh()

    def _update_clock(self):
        if self.clock_label:
            now = datetime.now()
            self.clock_label.config(text=f"🕐 {now.strftime('%Y-%m-%d %H:%M:%S')} 星期{['一','二','三','四','五','六','日'][now.weekday()]}")
            self.after(1000, self._update_clock)

    def _refresh(self):
        for w in self.winfo_children():
            w.destroy()
        self._build()
        self._update_clock()

    def _export_report(self):
        from tkinter import messagebox, filedialog
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            messagebox.showerror("错误", "缺少 openpyxl 库，请先安装: pip install openpyxl")
            return

        fp = filedialog.asksaveasfilename(title="导出统计报表", defaultextension=".xlsx", filetypes=[("Excel 文件", "*.xlsx")])
        if not fp:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "财务统计"
        ws["A1"] = "财务统计报表"
        ws["A1"].font = Font(name="微软雅黑", size=14, bold=True)
        ws["A2"] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A4"] = "项目"
        ws["B4"] = "金额"
        for c in ("A4", "B4"):
            ws[c].font = Font(name="微软雅黑", bold=True, color="FFFFFF")
            ws[c].fill = PatternFill(fill_type="solid", fgColor="4472C4")
        row = 5
        start, end = self._range_from_vars()
        today = date.today()
        month_start = today.replace(day=1)
        q = (today.month - 1) // 3 + 1
        quarter_start = date(today.year, (q - 1) * 3 + 1, 1)
        year_start = date(today.year, 1, 1)
        today_range = self._today_range()
        items = [
            ("今日总租金", self._stats_sum(date_range=today_range)),
            ("本月总租金", self._stats_sum(date_range=(month_start, today))),
            ("本季总租金", self._stats_sum(date_range=(quarter_start, today))),
            ("本年总租金", self._stats_sum(date_range=(year_start, today))),
            ("历史合计总租金", self._stats_sum()),
            ("选择日期总租金", self._stats_sum(date_range=(start, end))),
            ("今日成本统计", self._stats_sum(date_range=today_range, cost=True)),
            ("本月成本统计", self._stats_sum(date_range=(month_start, today), cost=True)),
            ("本季度成本统计", self._stats_sum(date_range=(quarter_start, today), cost=True)),
            ("本年度成本统计", self._stats_sum(date_range=(year_start, today), cost=True)),
            ("在租成本", self._stats_sum(lambda r: r.get("status") == "在租", cost=True)),
            ("逾期成本", self._stats_sum(lambda r: r.get("status") == "已逾期", cost=True)),
            ("丢失成本", self._stats_sum(lambda r: r.get("status") == "已丢失", cost=True)),
        ]
        for name, value in items:
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=value)
            row += 1
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 18
        wb.save(fp)
        messagebox.showinfo("成功", f"报表已导出:\n{fp}")
