#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
报表和筛选模块
"""

import csv
import tkinter as tk
from datetime import datetime
from tkinter import filedialog, messagebox, ttk

from theme.colors import DarkTheme


class RenewHistoryDialog:
    """续租历史对话框"""

    def __init__(self, parent, record):
        self.record = record
        self.win = tk.Toplevel(parent)
        self.win.title(f"续租历史 - {record.get('id', '')}")
        self.win.geometry("860x520")
        self.win.transient(parent)
        self.win.grab_set()
        self.win.configure(bg=DarkTheme.BG_PRIMARY)
        self.win.protocol("WM_DELETE_WINDOW", self._on_close)
        self._center()
        self._build()

    def _on_close(self):
        """安全关闭窗口，释放 grab"""
        try:
            self.win.grab_release()
        except Exception:
            pass
        self.win.destroy()

    def _center(self):
        self.win.update_idletasks()
        w, h = 860, 520
        x = (self.win.winfo_screenwidth() // 2) - (w // 2)
        y = (self.win.winfo_screenheight() // 2) - (h // 2)
        self.win.geometry(f"{w}x{h}+{x}+{y}")

    def _build(self):
        main = tk.Frame(self.win, bg=DarkTheme.BG_PRIMARY)
        main.pack(fill=tk.BOTH, expand=True, padx=16, pady=14)

        tk.Label(
            main,
            text="🔄 续租历史",
            font=DarkTheme.FONT_SUBTITLE,
            fg=DarkTheme.ACCENT_CYAN,
            bg=DarkTheme.BG_PRIMARY,
        ).pack(anchor=tk.W, pady=(0, 10))

        self.renew_history = self.record.get("renew_history", []) or []
        if not self.renew_history:
            tk.Label(
                main,
                text="暂无续租记录",
                font=DarkTheme.FONT_NORMAL,
                fg=DarkTheme.TEXT_SECONDARY,
                bg=DarkTheme.BG_PRIMARY,
            ).pack(pady=20)
        else:
            frame = tk.Frame(main, bg=DarkTheme.BG_PRIMARY)
            frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

            cols = ("续租时间", "时长", "单位", "金额", "原到期", "新到期", "操作人")
            tree = ttk.Treeview(frame, columns=cols, show="headings", height=14)
            widths = {
                "续租时间": 150,
                "时长": 70,
                "单位": 60,
                "金额": 90,
                "原到期": 110,
                "新到期": 110,
                "操作人": 100,
            }
            for c in cols:
                tree.heading(c, text=c)
                tree.column(c, width=widths.get(c, 90), anchor="center")

            vbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
            tree.configure(yscrollcommand=vbar.set)
            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            vbar.pack(side=tk.RIGHT, fill=tk.Y)

            for item in self.renew_history:
                tree.insert(
                    "",
                    tk.END,
                    values=(
                        item.get("renew_date", ""),
                        item.get("renew_time", ""),
                        item.get("renew_unit", ""),
                        f"¥{float(item.get('renew_amount', 0) or 0):.2f}",
                        item.get("old_end_date", ""),
                        item.get("new_end_date", ""),
                        item.get("operator", ""),
                    ),
                )

        btns = tk.Frame(main, bg=DarkTheme.BG_PRIMARY)
        btns.pack(fill=tk.X)
        exp_csv_btn = tk.Button(btns, text="📤 导出 CSV", font=DarkTheme.FONT_BUTTON, fg="white",
            bg=DarkTheme.ACCENT_GREEN, relief=tk.FLAT, cursor="hand2",
            command=self._export_csv, padx=14, pady=8)
        exp_csv_btn.pack(side=tk.LEFT, padx=(0, 4))
        DarkTheme.bind_hover(exp_csv_btn, DarkTheme.ACCENT_GREEN)
        exp_xlsx_btn = tk.Button(btns, text="📊 导出 Excel", font=DarkTheme.FONT_BUTTON, fg="white",
            bg=DarkTheme.ACCENT_BLUE, relief=tk.FLAT, cursor="hand2",
            command=self._export_xlsx, padx=14, pady=8)
        exp_xlsx_btn.pack(side=tk.LEFT, padx=(0, 4))
        DarkTheme.bind_hover(exp_xlsx_btn, DarkTheme.ACCENT_BLUE)
        cls_btn = tk.Button(btns, text="关闭", font=DarkTheme.FONT_BUTTON, fg="white",
            bg=DarkTheme.BG_HOVER, relief=tk.FLAT, cursor="hand2",
            command=self.win.destroy, padx=14, pady=8)
        cls_btn.pack(side=tk.LEFT)
        DarkTheme.bind_hover(cls_btn, DarkTheme.BG_HOVER)

    def _export_csv(self):
        if not self.renew_history:
            messagebox.showwarning("提示", "没有可导出的续租记录")
            return
        fp = filedialog.asksaveasfilename(
            title="导出续租历史",
            defaultextension=".csv",
            filetypes=[("CSV 文件", "*.csv"), ("所有文件", "*.*")],
            initialfile=f"续租历史_{self.record.get('id', '')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        )
        if not fp:
            return
        try:
            with open(fp, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(["续租时间", "时长", "单位", "金额", "原到期", "新到期", "操作人"])
                for h in self.renew_history:
                    writer.writerow(
                        [
                            h.get("renew_date", ""),
                            h.get("renew_time", ""),
                            h.get("renew_unit", ""),
                            h.get("renew_amount", ""),
                            h.get("old_end_date", ""),
                            h.get("new_end_date", ""),
                            h.get("operator", ""),
                        ]
                    )
            messagebox.showinfo("成功", f"已导出续租历史\n{fp}")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败：{e}")

    def _export_xlsx(self):
        """导出续租历史为 Excel 格式"""
        if not self.renew_history:
            messagebox.showwarning("提示", "没有可导出的续租记录")
            return
        fp = filedialog.asksaveasfilename(
            title="导出续租历史 Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
            initialfile=f"续租历史_{self.record.get('id', '')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )
        if not fp:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            wb = Workbook()
            ws = wb.active
            ws.title = "续租历史"

            headers = ["续租时间", "时长", "单位", "金额", "原到期", "新到期", "操作人"]
            header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for row_idx, h in enumerate(self.renew_history, 2):
                ws.cell(row=row_idx, column=1, value=h.get("renew_date", ""))
                ws.cell(row=row_idx, column=2, value=h.get("renew_time", ""))
                ws.cell(row=row_idx, column=3, value=h.get("renew_unit", ""))
                ws.cell(row=row_idx, column=4, value=float(h.get("renew_amount", 0) or 0))
                ws.cell(row=row_idx, column=5, value=h.get("old_end_date", ""))
                ws.cell(row=row_idx, column=6, value=h.get("new_end_date", ""))
                ws.cell(row=row_idx, column=7, value=h.get("operator", ""))

            for col in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max_length + 4

            wb.save(fp)
            messagebox.showinfo("成功", f"已导出 {len(self.renew_history)} 条续租记录\n{fp}")
        except ImportError:
            messagebox.showerror("错误", "缺少 openpyxl 库，请先安装: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败：{e}")


class AdvancedFilterDialog:
    """高级筛选对话框"""

    def __init__(self, parent, current_filters):
        self.result = None
        self.current = current_filters or {}
        self.entries = {}

        self.win = tk.Toplevel(parent)
        self.win.title("高级筛选")
        self.win.geometry("560x520")
        self.win.transient(parent)
        self.win.grab_set()
        self.win.configure(bg=DarkTheme.BG_PRIMARY)
        self.win.protocol("WM_DELETE_WINDOW", self._on_close)
        self._center()
        self._build()

    def _on_close(self):
        """安全关闭窗口，释放 grab"""
        try:
            self.win.grab_release()
        except Exception:
            pass
        self.win.destroy()

    def _center(self):
        self.win.update_idletasks()
        w, h = 560, 520
        x = (self.win.winfo_screenwidth() // 2) - (w // 2)
        y = (self.win.winfo_screenheight() // 2) - (h // 2)
        self.win.geometry(f"{w}x{h}+{x}+{y}")

    def _row(self, parent, label, key):
        row = tk.Frame(parent, bg=DarkTheme.BG_PRIMARY)
        row.pack(fill=tk.X, pady=4)
        tk.Label(
            row,
            text=label,
            font=DarkTheme.FONT_LABEL,
            fg=DarkTheme.TEXT_SECONDARY,
            bg=DarkTheme.BG_PRIMARY,
            width=16,
            anchor=tk.W,
        ).pack(side=tk.LEFT)
        ent = ttk.Entry(row, width=30, font=DarkTheme.FONT_NORMAL)
        ent.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ent.insert(0, str(self.current.get(key, "")))
        self.entries[key] = ent

    def _build(self):
        main = tk.Frame(self.win, bg=DarkTheme.BG_PRIMARY)
        main.pack(fill=tk.BOTH, expand=True, padx=16, pady=14)

        tk.Label(
            main,
            text="🧪 高级筛选条件",
            font=DarkTheme.FONT_SUBTITLE,
            fg=DarkTheme.ACCENT_CYAN,
            bg=DarkTheme.BG_PRIMARY,
        ).pack(anchor=tk.W, pady=(0, 10))

        self._row(main, "关键词(全字段)", "keyword")
        self._row(main, "起租日期从", "start_date_from")
        self._row(main, "起租日期到", "start_date_to")
        self._row(main, "到期日期从", "end_date_from")
        self._row(main, "到期日期到", "end_date_to")
        self._row(main, "总租金最小值", "total_rent_min")
        self._row(main, "总租金最大值", "total_rent_max")
        self._row(main, "已付金额最小值", "paid_min")
        self._row(main, "已付金额最大值", "paid_max")

        tk.Label(
            main,
            text="日期格式：YYYY-MM-DD；金额可留空",
            font=DarkTheme.FONT_NORMAL,
            fg=DarkTheme.TEXT_MUTED,
            bg=DarkTheme.BG_PRIMARY,
        ).pack(anchor=tk.W, pady=(8, 10))

        btns = tk.Frame(main, bg=DarkTheme.BG_PRIMARY)
        btns.pack(fill=tk.X)
        for txt, cmd, clr in [
            ("🔍 应用筛选", self._apply, DarkTheme.ACCENT_BLUE),
            ("🗑 清空条件", self._clear, DarkTheme.ACCENT_RED),
            ("取消", self.win.destroy, DarkTheme.BG_HOVER),
        ]:
            b = tk.Button(btns, text=txt, font=DarkTheme.FONT_BUTTON, fg="white", bg=clr,
                          relief=tk.FLAT, cursor="hand2", command=cmd, padx=14, pady=8)
            b.pack(side=tk.LEFT, padx=3)
            DarkTheme.bind_hover(b, clr)

    def _apply(self):
        self.result = {k: v.get().strip() for k, v in self.entries.items()}
        self.win.destroy()

    def _clear(self):
        self.result = {
            "keyword": "",
            "start_date_from": "",
            "start_date_to": "",
            "end_date_from": "",
            "end_date_to": "",
            "total_rent_min": "",
            "total_rent_max": "",
            "paid_min": "",
            "paid_max": "",
        }
        self.win.destroy()

    def show(self):
        self.win.wait_window()
        return self.result


class ReportDialog:
    """报表窗口"""

    def __init__(self, parent, data_manager):
        self.dm = data_manager
        self.win = tk.Toplevel(parent)
        self.win.title("租赁统计报表")
        self.win.geometry("920x620")
        self.win.transient(parent)
        self.win.grab_set()
        self.win.configure(bg=DarkTheme.BG_PRIMARY)
        self.win.protocol("WM_DELETE_WINDOW", self._on_close)
        self._center()
        self._build()

    def _on_close(self):
        """安全关闭窗口，释放 grab"""
        try:
            self.win.grab_release()
        except Exception:
            pass
        self.win.destroy()

    def _center(self):
        self.win.update_idletasks()
        w, h = 920, 620
        x = (self.win.winfo_screenwidth() // 2) - (w // 2)
        y = (self.win.winfo_screenheight() // 2) - (h // 2)
        self.win.geometry(f"{w}x{h}+{x}+{y}")

    def _build(self):
        main = tk.Frame(self.win, bg=DarkTheme.BG_PRIMARY)
        main.pack(fill=tk.BOTH, expand=True, padx=16, pady=14)

        tk.Label(
            main,
            text="📊 租赁统计报表",
            font=DarkTheme.FONT_SUBTITLE,
            fg=DarkTheme.ACCENT_CYAN,
            bg=DarkTheme.BG_PRIMARY,
        ).pack(anchor=tk.W, pady=(0, 10))

        stats = self.dm.get_stats()
        records = self.dm.get_records()
        total_rent = sum(float(r.get("lease_info", {}).get("total_rent", 0) or 0) for r in records)
        paid_amount = sum(float(r.get("paid_amount", 0) or 0) for r in records)
        unpaid_amount = total_rent - paid_amount

        cards = tk.Frame(main, bg=DarkTheme.BG_PRIMARY)
        cards.pack(fill=tk.X, pady=(0, 12))
        card_items = [
            ("总记录数", stats["total"], DarkTheme.ACCENT_CYAN),
            ("在租中", stats["active"], DarkTheme.ACCENT_BLUE),
            ("已逾期", stats["expired"], DarkTheme.ACCENT_RED),
            ("已退租", stats["returned"], DarkTheme.ACCENT_GREEN),
            ("已丢失", stats["lost"], DarkTheme.ACCENT_YELLOW),
            ("已买断", stats["bought"], DarkTheme.ACCENT_PURPLE),
        ]
        for label, value, color in card_items:
            card = tk.Frame(cards, bg=DarkTheme.BG_CARD, highlightbackground=DarkTheme.BORDER_COLOR, highlightthickness=1)
            card.pack(side=tk.LEFT, padx=4, fill=tk.BOTH, expand=True)
            tk.Label(card, text=label, font=DarkTheme.FONT_LABEL, fg=DarkTheme.TEXT_SECONDARY, bg=DarkTheme.BG_CARD).pack(pady=(8, 2))
            tk.Label(card, text=str(value), font=("微软雅黑", 18, "bold"), fg=color, bg=DarkTheme.BG_CARD).pack(pady=(2, 8))

        amount_box = tk.Frame(main, bg=DarkTheme.BG_CARD, highlightbackground=DarkTheme.BORDER_COLOR, highlightthickness=1)
        amount_box.pack(fill=tk.X, pady=(0, 12))
        tk.Label(amount_box, text="金额概览", font=("微软雅黑", 12, "bold"), fg=DarkTheme.ACCENT_BLUE, bg=DarkTheme.BG_CARD).pack(anchor=tk.W, padx=12, pady=(10, 6))
        for lbl, val, color in [
            ("总租金", total_rent, DarkTheme.ACCENT_BLUE),
            ("已收金额", paid_amount, DarkTheme.ACCENT_GREEN),
            ("未收金额", unpaid_amount, DarkTheme.ACCENT_RED),
        ]:
            row = tk.Frame(amount_box, bg=DarkTheme.BG_CARD)
            row.pack(fill=tk.X, padx=12, pady=2)
            tk.Label(row, text=f"{lbl}：", font=DarkTheme.FONT_LABEL, fg=DarkTheme.TEXT_SECONDARY, bg=DarkTheme.BG_CARD, width=10, anchor=tk.W).pack(side=tk.LEFT)
            tk.Label(row, text=f"¥{val:.2f}", font=("微软雅黑", 14, "bold"), fg=color, bg=DarkTheme.BG_CARD).pack(side=tk.LEFT)

        status_box = tk.Frame(main, bg=DarkTheme.BG_CARD, highlightbackground=DarkTheme.BORDER_COLOR, highlightthickness=1)
        status_box.pack(fill=tk.BOTH, expand=True)
        tk.Label(status_box, text="状态分布", font=("微软雅黑", 12, "bold"), fg=DarkTheme.ACCENT_BLUE, bg=DarkTheme.BG_CARD).pack(anchor=tk.W, padx=12, pady=(10, 6))
        total = max(1, stats["total"])
        for label, value, color in [
            ("在租", stats["active"], DarkTheme.STATUS_ACTIVE),
            ("逾期", stats["expired"], DarkTheme.STATUS_EXPIRED),
            ("退租", stats["returned"], DarkTheme.STATUS_RETURNED),
            ("丢失", stats["lost"], DarkTheme.STATUS_LOST),
            ("买断", stats["bought"], DarkTheme.STATUS_BOUGHT),
        ]:
            row = tk.Frame(status_box, bg=DarkTheme.BG_CARD)
            row.pack(fill=tk.X, padx=12, pady=3)
            tk.Label(row, text=f"{label}：{value}", font=DarkTheme.FONT_LABEL, fg=DarkTheme.TEXT_SECONDARY, bg=DarkTheme.BG_CARD, width=12, anchor=tk.W).pack(side=tk.LEFT)
            bar = tk.Frame(row, bg=DarkTheme.BG_INPUT, height=14)
            bar.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(4, 0))
            fill = tk.Frame(bar, bg=color, height=14, width=max(2, int((value / total) * 360)))
            fill.pack(side=tk.LEFT, fill=tk.Y)

        btns = tk.Frame(main, bg=DarkTheme.BG_PRIMARY)
        btns.pack(fill=tk.X, pady=(12, 0))
        chart_btn = tk.Button(btns, text="📊 查看图表", font=DarkTheme.FONT_BUTTON, fg="white",
            bg=DarkTheme.ACCENT_PURPLE, relief=tk.FLAT, cursor="hand2",
            command=lambda: self._show_charts(stats, total_rent, paid_amount, unpaid_amount), padx=14, pady=8)
        chart_btn.pack(side=tk.LEFT, padx=(0, 8))
        DarkTheme.bind_hover(chart_btn, DarkTheme.ACCENT_PURPLE)
        exp_excel_btn = tk.Button(btns, text="📈 导出 Excel", font=DarkTheme.FONT_BUTTON, fg="white",
            bg=DarkTheme.ACCENT_BLUE, relief=tk.FLAT, cursor="hand2",
            command=lambda: self._export_excel(stats, total_rent, paid_amount, unpaid_amount), padx=14, pady=8)
        exp_excel_btn.pack(side=tk.LEFT, padx=(0, 8))
        DarkTheme.bind_hover(exp_excel_btn, DarkTheme.ACCENT_BLUE)
        exp_btn = tk.Button(btns, text="📤 导出报表 CSV", font=DarkTheme.FONT_BUTTON, fg="white",
            bg=DarkTheme.ACCENT_GREEN, relief=tk.FLAT, cursor="hand2",
            command=lambda: self._export(stats, total_rent, paid_amount, unpaid_amount), padx=14, pady=8)
        exp_btn.pack(side=tk.LEFT, padx=(0, 8))
        DarkTheme.bind_hover(exp_btn, DarkTheme.ACCENT_GREEN)
        cls_btn = tk.Button(btns, text="关闭", font=DarkTheme.FONT_BUTTON, fg="white",
            bg=DarkTheme.BG_HOVER, relief=tk.FLAT, cursor="hand2",
            command=self._on_close, padx=14, pady=8)
        cls_btn.pack(side=tk.LEFT)
        DarkTheme.bind_hover(cls_btn, DarkTheme.BG_HOVER)

    def _show_charts(self, stats, total_rent, paid_amount, unpaid_amount):
        """显示 matplotlib 统计图表"""
        try:
            import matplotlib
            matplotlib.use("TkAgg")
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "Arial Unicode MS"]
            plt.rcParams["axes.unicode_minus"] = False
        except Exception as e:
            messagebox.showerror("错误", f"图表加载失败：{e}")
            return

        win = tk.Toplevel(self.win)
        win.title("📊 统计图表")
        win.geometry("900x600")
        win.transient(self.win)
        win.grab_set()
        win.configure(bg=DarkTheme.BG_PRIMARY)

        def _safe_close(w=win):
            try:
                w.grab_release()
            except Exception:
                pass
            w.destroy()

        win.protocol("WM_DELETE_WINDOW", _safe_close)

        main = tk.Frame(win, bg=DarkTheme.BG_PRIMARY)
        main.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)

        tk.Label(main, text="📊 数据可视化", font=DarkTheme.FONT_SUBTITLE,
                 fg=DarkTheme.ACCENT_CYAN, bg=DarkTheme.BG_PRIMARY).pack(anchor=tk.W, pady=(0, 8))

        # 创建 2x2 图表网格
        chart_frame = tk.Frame(main, bg=DarkTheme.BG_PRIMARY)
        chart_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 8))

        fig, axes = plt.subplots(2, 2, figsize=(12, 9))
        fig.patch.set_facecolor("#1e1e2e")

        # 1. 饼图 - 租赁状态分布
        ax1 = axes[0, 0]
        status_labels = ["在租", "逾期", "退租", "丢失", "买断"]
        status_values = [stats["active"], stats["expired"], stats["returned"], stats["lost"], stats["bought"]]
        status_colors = ["#22c55e", "#ef4444", "#3b82f6", "#f59e0b", "#a855f7"]
        if sum(status_values) > 0:
            ax1.pie(status_values, labels=status_labels, colors=status_colors, autopct='%1.1f%%',
                    textprops={'color': 'white', 'fontsize': 9})
        ax1.set_title("租赁状态分布", color='white', fontsize=12, fontweight='bold')
        ax1.set_facecolor('#2d2d3f')

        # 2. 柱状图 - 状态数量对比
        ax2 = axes[0, 1]
        ax2.bar(status_labels, status_values, color=status_colors, edgecolor='white', linewidth=0.5)
        ax2.set_title("状态数量对比", color='white', fontsize=12, fontweight='bold')
        ax2.set_facecolor('#2d2d3f')
        ax2.tick_params(colors='white')
        for i, v in enumerate(status_values):
            ax2.text(i, v + 0.1, str(v), ha='center', color='white', fontsize=9)

        # 3. 环形图 - 财务分布
        ax3 = axes[1, 0]
        fin_labels = ["已收金额", "未收金额"]
        fin_values = [paid_amount, max(0, unpaid_amount)]
        fin_colors = ["#22c55e", "#ef4444"]
        if sum(fin_values) > 0:
            wedges, texts, autotexts = ax3.pie(fin_values, labels=fin_labels, colors=fin_colors,
                                               autopct='%1.1f%%', textprops={'color': 'white', 'fontsize': 9})
        ax3.set_title("财务分布", color='white', fontsize=12, fontweight='bold')
        ax3.set_facecolor('#2d2d3f')

        # 4. 金额柱状图
        ax4 = axes[1, 1]
        fin_labels2 = ["总租金", "已收", "未收"]
        fin_values2 = [total_rent, paid_amount, max(0, unpaid_amount)]
        fin_colors2 = ["#3b82f6", "#22c55e", "#ef4444"]
        ax4.bar(fin_labels2, fin_values2, color=fin_colors2, edgecolor='white', linewidth=0.5)
        ax4.set_title("金额对比", color='white', fontsize=12, fontweight='bold')
        ax4.set_facecolor('#2d2d3f')
        ax4.tick_params(colors='white')
        for i, v in enumerate(fin_values2):
            ax4.text(i, v + total_rent * 0.01, f"¥{v:.0f}", ha='center', color='white', fontsize=8)

        plt.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        # 保存按钮
        btn_row = tk.Frame(main, bg=DarkTheme.BG_PRIMARY)
        btn_row.pack(fill=tk.X, pady=(0, 8))
        tk.Button(btn_row, text="💾 保存图表", font=DarkTheme.FONT_BUTTON, fg="white",
                  bg=DarkTheme.ACCENT_BLUE, relief=tk.FLAT, cursor="hand2",
                  command=lambda: self._save_chart(fig), padx=14, pady=6).pack(side=tk.LEFT, padx=(0, 8))
        tk.Button(btn_row, text="关闭", font=DarkTheme.FONT_BUTTON, fg="white",
                  bg=DarkTheme.BG_HOVER, relief=tk.FLAT, cursor="hand2",
                  command=_safe_close, padx=14, pady=6).pack(side=tk.LEFT)

    def _save_chart(self, fig):
        """保存图表为图片"""
        fp = filedialog.asksaveasfilename(
            title="保存统计图表",
            defaultextension=".png",
            filetypes=[("PNG 图片", "*.png"), ("所有文件", "*.*")],
            initialfile=f"租赁统计图表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png",
        )
        if not fp:
            return
        try:
            fig.savefig(fp, dpi=150, bbox_inches='tight', facecolor='#1e1e2e')
            messagebox.showinfo("成功", f"图表已保存\n{fp}")
        except Exception as e:
            messagebox.showerror("错误", f"保存失败：{e}")

    def _export_excel(self, stats, total_rent, paid_amount, unpaid_amount):
        """导出统计报表为 Excel 格式"""
        fp = filedialog.asksaveasfilename(
            title="导出租赁报表 Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
            initialfile=f"租赁报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )
        if not fp:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            wb = Workbook()
            ws = wb.active
            ws.title = "租赁统计报表"

            # 样式
            title_font = Font(name="微软雅黑", size=14, bold=True, color="4472C4")
            header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            normal_font = Font(name="微软雅黑", size=11)
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )

            # 标题
            ws.merge_cells("A1:B1")
            ws["A1"] = "速维电脑租赁管理系统 — 统计报表"
            ws["A1"].font = title_font
            ws["A2"] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws["A2"].font = normal_font

            # 基本统计
            ws["A4"] = "基本统计"
            ws["A4"].font = Font(name="微软雅黑", size=12, bold=True, color="4472C4")
            row = 5
            ws.cell(row=row, column=1, value="项目").font = header_font
            ws.cell(row=row, column=1).fill = header_fill
            ws.cell(row=row, column=2, value="数量").font = header_font
            ws.cell(row=row, column=2).fill = header_fill
            for label, key in [("总记录", "total"), ("在租", "active"), ("逾期", "expired"),
                               ("退租", "returned"), ("丢失", "lost"), ("买断", "bought")]:
                row += 1
                c1 = ws.cell(row=row, column=1, value=label)
                c1.font = normal_font
                c1.border = thin_border
                c2 = ws.cell(row=row, column=2, value=stats[key])
                c2.font = normal_font
                c2.border = thin_border
                row += 1

            # 财务汇总
            row += 1
            ws.cell(row=row, column=1, value="财务汇总").font = Font(name="微软雅黑", size=12, bold=True, color="4472C4")
            row += 1
            for label, val in [("总租金", total_rent), ("已收金额", paid_amount), ("未收金额", unpaid_amount)]:
                c1 = ws.cell(row=row, column=1, value=label)
                c1.font = normal_font
                c1.border = thin_border
                c2 = ws.cell(row=row, column=2, value=val)
                c2.font = normal_font
                c2.number_format = '"¥"#,##0.00'
                c2.border = thin_border
                row += 1

            # 列宽
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 18

            wb.save(fp)
            messagebox.showinfo("成功", f"报表已导出\n{fp}")
        except ImportError:
            messagebox.showerror("错误", "缺少 openpyxl 库，请先安装: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败：{e}")

    def _export(self, stats, total_rent, paid_amount, unpaid_amount):
        """导出报表（CSV）"""
        fp = filedialog.asksaveasfilename(
            title="导出租赁报表",
            defaultextension=".csv",
            filetypes=[("CSV 文件", "*.csv"), ("所有文件", "*.*")],
            initialfile=f"租赁报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        )
        if not fp:
            return
        try:
            with open(fp, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(["项目", "值"])
                writer.writerow(["总记录数", stats["total"]])
                writer.writerow(["在租", stats["active"]])
                writer.writerow(["逾期", stats["expired"]])
                writer.writerow(["退租", stats["returned"]])
                writer.writerow(["丢失", stats["lost"]])
                writer.writerow(["买断", stats["bought"]])
                writer.writerow([])
                writer.writerow(["总租金", total_rent])
                writer.writerow(["已收金额", paid_amount])
                writer.writerow(["未收金额", unpaid_amount])
                writer.writerow(["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            messagebox.showinfo("成功", f"报表已导出\n{fp}")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败：{e}")
