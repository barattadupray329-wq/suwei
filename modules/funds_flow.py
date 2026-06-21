#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""资金流水页面，面向财务统计与对账。"""

import csv
import tkinter as tk
from datetime import date, datetime
from tkinter import filedialog, messagebox, ttk

from theme.colors import DarkTheme
from modules.rental_mgmt import RentalDetailDialog


class FundsFlowFrame(ttk.Frame):
    """页签式财务中心：收款流水、买断流水、汇总统计。"""

    def __init__(self, parent, data_manager):
        super().__init__(parent)
        self.data_manager = data_manager
        self.search_var = tk.StringVar()
        self.type_var = tk.StringVar(value="全部")
        self.start_var = tk.StringVar()
        self.end_var = tk.StringVar()
        self.summary_var = tk.StringVar(value="")
        self._selected_event = None
        self._selected_receipt = None
        self._build()
        self._set_month_range()

    def _action_button(self, parent, text, command, color):
        btn = tk.Button(parent, text=text, font=DarkTheme.FONT_SMALL, fg="white", bg=color,
                        relief=tk.FLAT, cursor="hand2", command=command, padx=12, pady=5)
        btn.pack(side=tk.LEFT, padx=3)
        DarkTheme.bind_hover(btn, color)
        return btn

    def _build(self):
        main = tk.Frame(self, bg=DarkTheme.BG_PRIMARY)
        main.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        head = tk.Frame(main, bg=DarkTheme.BG_PRIMARY)
        head.pack(fill=tk.X, pady=(0, 16))
        tk.Label(head, text="📒 收支流水", font=DarkTheme.FONT_TITLE, fg=DarkTheme.TEXT_PRIMARY, bg=DarkTheme.BG_PRIMARY).pack(anchor=tk.W)
        tk.Label(head, text="页签式财务中心：收款流水、买断流水与汇总统计", font=DarkTheme.FONT_SMALL,
                 fg=DarkTheme.TEXT_MUTED, bg=DarkTheme.BG_PRIMARY).pack(anchor=tk.W, pady=(4, 0))

        top = tk.Frame(main, bg=DarkTheme.BG_CARD, highlightbackground=DarkTheme.BORDER_COLOR, highlightthickness=1)
        top.pack(fill=tk.X, pady=(0, 12))
        tk.Label(top, textvariable=self.summary_var, font=DarkTheme.FONT_LABEL,
                 fg=DarkTheme.ACCENT_CYAN, bg=DarkTheme.BG_CARD).pack(anchor=tk.W, padx=16, pady=(12, 4))
        tk.Label(top, text="提示：可通过页签查看不同类型流水，点击记录可查看详情。", font=DarkTheme.FONT_SMALL,
                 fg=DarkTheme.TEXT_MUTED, bg=DarkTheme.BG_CARD).pack(anchor=tk.W, padx=16, pady=(0, 12))

        ctrl = tk.Frame(main, bg=DarkTheme.BG_CARD, highlightbackground=DarkTheme.BORDER_COLOR, highlightthickness=1)
        ctrl.pack(fill=tk.X, pady=(0, 12))
        row = tk.Frame(ctrl, bg=DarkTheme.BG_CARD)
        row.pack(fill=tk.X, padx=16, pady=(12, 6))
        tk.Label(row, text="搜索", font=DarkTheme.FONT_LABEL, fg=DarkTheme.TEXT_SECONDARY, bg=DarkTheme.BG_CARD).pack(side=tk.LEFT)
        ttk.Entry(row, textvariable=self.search_var, width=24).pack(side=tk.LEFT, padx=(8, 10))
        tk.Label(row, text="类型", font=DarkTheme.FONT_LABEL, fg=DarkTheme.TEXT_SECONDARY, bg=DarkTheme.BG_CARD).pack(side=tk.LEFT)
        ttk.Combobox(row, textvariable=self.type_var, values=["全部", "租金", "买断"], width=10, state="readonly").pack(side=tk.LEFT, padx=(8, 10))
        self._action_button(row, "查询", self._refresh, DarkTheme.ACCENT_BLUE)
        self._action_button(row, "刷新", self._set_default_range_and_refresh, DarkTheme.ACCENT_GREEN)
        self._action_button(row, "导出 CSV", self._export_csv, DarkTheme.ACCENT_PURPLE)
        self._action_button(row, "导出 Excel", self._export_excel, DarkTheme.ACCENT_YELLOW)

        range_row = tk.Frame(ctrl, bg=DarkTheme.BG_CARD)
        range_row.pack(fill=tk.X, padx=16, pady=(0, 12))
        tk.Label(range_row, text="起始日期", font=DarkTheme.FONT_LABEL, fg=DarkTheme.TEXT_SECONDARY, bg=DarkTheme.BG_CARD).pack(side=tk.LEFT)
        ttk.Entry(range_row, textvariable=self.start_var, width=14).pack(side=tk.LEFT, padx=(8, 12))
        tk.Label(range_row, text="结束日期", font=DarkTheme.FONT_LABEL, fg=DarkTheme.TEXT_SECONDARY, bg=DarkTheme.BG_CARD).pack(side=tk.LEFT)
        ttk.Entry(range_row, textvariable=self.end_var, width=14).pack(side=tk.LEFT, padx=(8, 12))
        self._action_button(range_row, "本月", self._set_month_range, DarkTheme.ACCENT_CYAN)
        self._action_button(range_row, "本季", self._set_quarter_range, DarkTheme.ACCENT_PURPLE)
        self._action_button(range_row, "本年", self._set_year_range, DarkTheme.ACCENT_YELLOW)
        self._action_button(range_row, "历史", self._set_history_range, DarkTheme.BG_HOVER)

        quick = tk.Frame(ctrl, bg=DarkTheme.BG_CARD)
        quick.pack(fill=tk.X, padx=16, pady=(0, 12))
        self._action_button(quick, "全部", lambda: self._set_type_and_refresh("全部"), DarkTheme.ACCENT_BLUE)
        self._action_button(quick, "租金", lambda: self._set_type_and_refresh("租金"), DarkTheme.ACCENT_GREEN)
        self._action_button(quick, "买断", lambda: self._set_type_and_refresh("买断"), DarkTheme.ACCENT_PURPLE)
        self._action_button(quick, "清空筛选", self._reset_filters, DarkTheme.BG_HOVER)

        self.tabs = ttk.Notebook(main)
        self.tabs.pack(fill=tk.BOTH, expand=True)

        self.receipt_tab = self._make_tab("收款流水")
        self.buyout_tab = self._make_tab("买断流水")
        self.summary_tab = self._make_tab("汇总统计")
        self.tabs.add(self.receipt_tab, text="收款流水")
        self.tabs.add(self.buyout_tab, text="买断流水")
        self.tabs.add(self.summary_tab, text="汇总统计")

        self._build_receipt_tab(self.receipt_tab)
        self._build_buyout_tab(self.buyout_tab)
        self._build_summary_tab(self.summary_tab)

    def _make_tab(self, name):
        tab = tk.Frame(self.tabs, bg=DarkTheme.BG_PRIMARY)
        tab._tab_name = name
        return tab

    def _build_tree_panel(self, parent, title):
        panel = tk.Frame(parent, bg=DarkTheme.BG_CARD, highlightbackground=DarkTheme.BORDER_COLOR, highlightthickness=1)
        panel.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)
        tk.Label(panel, text=title, font=DarkTheme.FONT_SUBTITLE,
                 fg=DarkTheme.ACCENT_GREEN, bg=DarkTheme.BG_CARD).pack(anchor=tk.W, padx=16, pady=(12, 8))
        cols = ("类型", "时间", "名称", "单号", "金额")
        tree = ttk.Treeview(panel, columns=cols, show="headings", height=16)
        widths = {"类型": 80, "时间": 145, "名称": 160, "单号": 120, "金额": 100}
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=widths[c], anchor="center")
        vbar = ttk.Scrollbar(panel, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vbar.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(16, 0), pady=(0, 12))
        vbar.pack(side=tk.RIGHT, fill=tk.Y, pady=(0, 12))
        return panel, tree

    def _build_detail_block(self, parent):
        block = tk.Frame(parent, bg=DarkTheme.BG_CARD, highlightbackground=DarkTheme.BORDER_COLOR, highlightthickness=1)
        block.pack(fill=tk.X, pady=(10, 0))
        tk.Label(block, text="详情", font=DarkTheme.FONT_SUBTITLE, fg=DarkTheme.ACCENT_CYAN, bg=DarkTheme.BG_CARD).pack(anchor=tk.W, padx=16, pady=(12, 8))
        text = tk.Text(block, height=8, wrap=tk.WORD, font=DarkTheme.FONT_NORMAL,
                       bg=DarkTheme.BG_INPUT, fg=DarkTheme.TEXT_PRIMARY, insertbackground=DarkTheme.TEXT_PRIMARY)
        text.pack(fill=tk.X, padx=16, pady=(0, 12))
        text.configure(state="disabled")
        return block, text

    def _build_receipt_tab(self, parent):
        container = tk.Frame(parent, bg=DarkTheme.BG_PRIMARY)
        container.pack(fill=tk.BOTH, expand=True)
        self.receipt_panel, self.receipt_tree = self._build_tree_panel(container, "收款流水（基于 payment_history）")
        self.receipt_tree.bind("<<TreeviewSelect>>", lambda e: self._on_select(self.receipt_tree, self.receipt_detail_text))
        self.receipt_tree.bind("<Double-1>", lambda e: self._open_selected_record(self.receipt_tree))
        _, self.receipt_detail_text = self._build_detail_block(container)
        self.receipt_panel.pack(fill=tk.BOTH, expand=True)

    def _build_buyout_tab(self, parent):
        container = tk.Frame(parent, bg=DarkTheme.BG_PRIMARY)
        container.pack(fill=tk.BOTH, expand=True)
        self.buyout_panel, self.buyout_tree = self._build_tree_panel(container, "买断流水")
        self.buyout_tree.bind("<<TreeviewSelect>>", lambda e: self._on_select(self.buyout_tree, self.buyout_detail_text))
        self.buyout_tree.bind("<Double-1>", lambda e: self._open_selected_record(self.buyout_tree))
        _, self.buyout_detail_text = self._build_detail_block(container)
        self.buyout_panel.pack(fill=tk.BOTH, expand=True)

    def _build_summary_tab(self, parent):
        container = tk.Frame(parent, bg=DarkTheme.BG_PRIMARY)
        container.pack(fill=tk.BOTH, expand=True)
        top = tk.Frame(container, bg=DarkTheme.BG_CARD, highlightbackground=DarkTheme.BORDER_COLOR, highlightthickness=1)
        top.pack(fill=tk.X, pady=(0, 12))
        tk.Label(top, text="汇总统计", font=DarkTheme.FONT_SUBTITLE, fg=DarkTheme.ACCENT_PURPLE, bg=DarkTheme.BG_CARD).pack(anchor=tk.W, padx=16, pady=(12, 8))
        self.summary_detail_text = tk.Text(top, height=12, wrap=tk.WORD, font=DarkTheme.FONT_NORMAL,
                                           bg=DarkTheme.BG_INPUT, fg=DarkTheme.TEXT_PRIMARY, insertbackground=DarkTheme.TEXT_PRIMARY)
        self.summary_detail_text.pack(fill=tk.X, padx=16, pady=(0, 12))
        self.summary_detail_text.configure(state="disabled")

    def _parse_date(self, text):
        try:
            return datetime.strptime((text or "").strip()[:10], "%Y-%m-%d").date()
        except Exception:
            return None

    def _get_date_range(self):
        start = self._parse_date(self.start_var.get()) or date(1970, 1, 1)
        end = self._parse_date(self.end_var.get()) or date.today()
        if start > end:
            start, end = end, start
        return start, end

    def _record_matches_range(self, dt):
        start, end = self._get_date_range()
        return dt is not None and start <= dt <= end

    def _payment_events(self):
        records = self.data_manager.get_records()
        events = []
        for rec in records:
            renter = rec.get("renter", {}) or {}
            lease = rec.get("lease_info", {}) or {}
            rid = rec.get("id", "")
            name = renter.get("name", "")
            total_rent = float(lease.get("total_rent", 0) or 0)
            end_date = lease.get("end_date", "")
            payment_history = rec.get("payment_history", []) or []
            if self.type_var.get() in ("全部", "租金"):
                if payment_history:
                    paid_running = 0.0
                    for item in payment_history:
                        amount = float(item.get("amount", 0) or 0)
                        if amount <= 0:
                            continue
                        pay_dt = self._parse_date(item.get("payment_date", "")) or self._parse_date(rec.get("register_date", ""))
                        if not self._record_matches_range(pay_dt):
                            continue
                        paid_running += amount
                        events.append({
                            "type": "租金",
                            "time": item.get("payment_date", ""),
                            "name": name,
                            "no": rid,
                            "amount": amount,
                            "status": rec.get("status", ""),
                            "note": f"已收累计：¥{paid_running:.2f}；到期：{end_date}" if end_date else f"已收累计：¥{paid_running:.2f}",
                            "source": item.get("source", ""),
                            "operator": item.get("operator", ""),
                        })
                else:
                    rent_dt = self._parse_date(rec.get("register_date", "") or lease.get("start_date", ""))
                    if self._record_matches_range(rent_dt):
                        events.append({
                            "type": "租金",
                            "time": rec.get("register_date", "") or lease.get("start_date", ""),
                            "name": name,
                            "no": rid,
                            "amount": total_rent,
                            "status": rec.get("status", ""),
                            "note": f"到期：{end_date}" if end_date else "",
                            "source": lease.get("lease_type", "租金"),
                            "operator": rec.get("operator", ""),
                        })
            if self.type_var.get() in ("全部", "买断"):
                if rec.get("status") == "已买断" or rec.get("settlement_amount") not in (None, "", 0, 0.0):
                    buyout_amount = float(rec.get("settlement_amount", 0) or 0)
                    buyout_dt = self._parse_date(rec.get("settlement_date", "") or rec.get("updated_at", "") or rec.get("register_date", ""))
                    if self._record_matches_range(buyout_dt):
                        events.append({
                            "type": "买断",
                            "time": rec.get("settlement_date", "") or rec.get("updated_at", ""),
                            "name": name,
                            "no": rid,
                            "amount": buyout_amount,
                            "status": rec.get("status", ""),
                            "note": "买断结算",
                            "source": "买断费",
                            "operator": rec.get("operator", ""),
                        })
        q = self.search_var.get().strip().lower()
        if q:
            events = [e for e in events if q in f"{e['type']}{e['time']}{e['name']}{e['no']}{e['amount']}{e.get('note','')}{e.get('source','')}{e.get('operator','')}".lower()]
        events.sort(key=lambda x: x.get("time", ""), reverse=True)
        return events

    def _receipt_events(self):
        return [e for e in self._payment_events() if e.get("type") == "租金"]

    def _buyout_events(self):
        return [e for e in self._payment_events() if e.get("type") == "买断"]

    def _set_default_range_and_refresh(self):
        self._set_month_range()

    def _set_type_and_refresh(self, value):
        self.type_var.set(value)
        self._refresh()

    def _reset_filters(self):
        self.search_var.set("")
        self.type_var.set("全部")
        self._set_month_range()

    def _set_month_range(self):
        today = date.today()
        self.start_var.set(today.replace(day=1).strftime("%Y-%m-%d"))
        self.end_var.set(today.strftime("%Y-%m-%d"))
        self._refresh()

    def _set_quarter_range(self):
        today = date.today()
        q = (today.month - 1) // 3 + 1
        month = (q - 1) * 3 + 1
        self.start_var.set(date(today.year, month, 1).strftime("%Y-%m-%d"))
        self.end_var.set(today.strftime("%Y-%m-%d"))
        self._refresh()

    def _set_year_range(self):
        today = date.today()
        self.start_var.set(date(today.year, 1, 1).strftime("%Y-%m-%d"))
        self.end_var.set(today.strftime("%Y-%m-%d"))
        self._refresh()

    def _set_history_range(self):
        self.start_var.set("1970-01-01")
        self.end_var.set(date.today().strftime("%Y-%m-%d"))
        self._refresh()

    def _fill_tree(self, tree, events):
        for item in tree.get_children():
            tree.delete(item)
        for e in events:
            tree.insert("", tk.END, values=(
                e.get("type", ""),
                e.get("time", ""),
                e.get("name", ""),
                e.get("no", ""),
                f"¥{float(e.get('amount', 0) or 0):.2f}",
            ))

    def _set_text(self, text_widget, lines):
        text_widget.configure(state="normal")
        text_widget.delete("1.0", tk.END)
        text_widget.insert("1.0", "\n".join(lines))
        text_widget.configure(state="disabled")

    def _on_select(self, tree, text_widget):
        sel = tree.selection()
        if not sel:
            return
        idx = tree.index(sel[0])
        if tree is self.receipt_tree:
            events = self._receipt_events()
        else:
            events = self._buyout_events()
        if 0 <= idx < len(events):
            e = events[idx]
            self._selected_event = e
            lines = [
                f"类型：{e.get('type','')}",
                f"时间：{e.get('time','')}",
                f"名称：{e.get('name','')}",
                f"单号：{e.get('no','')}",
                f"金额：¥{float(e.get('amount',0) or 0):.2f}",
                f"来源：{e.get('source','') or '-'}",
                f"备注：{e.get('note','') or '-'}",
                f"操作人：{e.get('operator','') or '-'}",
                f"状态：{e.get('status','') or '-'}",
            ]
            self._set_text(text_widget, lines)

    def _open_selected_record(self, tree):
        sel = tree.selection()
        if not sel:
            return
        idx = tree.index(sel[0])
        events = self._receipt_events() if tree is self.receipt_tree else self._buyout_events()
        if not (0 <= idx < len(events)):
            return
        event = events[idx]
        rid = event.get("no", "")
        if not rid:
            return
        rec = None
        getter = getattr(self.data_manager, "get_record_by_id", None)
        if callable(getter):
            try:
                rec = getter(rid)
            except Exception:
                rec = None
        if rec is None:
            for item in self.data_manager.get_records():
                if item.get("id") == rid:
                    rec = item
                    break
        if rec:
            try:
                RentalDetailDialog(self.master, rec)
            except Exception:
                messagebox.showinfo("提示", f"已定位到单号 {rid}，但详情窗口打开失败。")

    def _refresh(self):
        receipt_events = self._receipt_events()
        buyout_events = self._buyout_events()
        self._fill_tree(self.receipt_tree, receipt_events)
        self._fill_tree(self.buyout_tree, buyout_events)

        rent_total = sum(float(e.get("amount", 0) or 0) for e in receipt_events)
        buyout_total = sum(float(e.get("amount", 0) or 0) for e in buyout_events)
        range_text = f"{self.start_var.get().strip() or '不限'} 至 {self.end_var.get().strip() or '不限'}"
        self.summary_var.set(f"区间 {range_text} ｜ 收款 {len(receipt_events)} 笔 / ¥{rent_total:.2f} ｜ 买断 {len(buyout_events)} 笔 / ¥{buyout_total:.2f} ｜ 合计 {len(receipt_events) + len(buyout_events)} 笔")
        self._set_text(self.receipt_detail_text, ["当前显示收款流水。"])
        self._set_text(self.buyout_detail_text, ["当前显示买断流水。"])
        self._build_summary_text(receipt_events, buyout_events)
        if receipt_events:
            self.receipt_tree.selection_set(self.receipt_tree.get_children()[0])
            self._on_select(self.receipt_tree, self.receipt_detail_text)
        if buyout_events:
            self.buyout_tree.selection_set(self.buyout_tree.get_children()[0])
            self._on_select(self.buyout_tree, self.buyout_detail_text)

    def _build_summary_text(self, receipt_events, buyout_events):
        lines = [
            f"收款流水：{len(receipt_events)} 笔",
            f"收款合计：¥{sum(float(e.get('amount', 0) or 0) for e in receipt_events):.2f}",
            f"买断流水：{len(buyout_events)} 笔",
            f"买断合计：¥{sum(float(e.get('amount', 0) or 0) for e in buyout_events):.2f}",
            "",
            "说明：收款流水优先取 payment_history，买断流水取已买断记录和结算金额。",
        ]
        self._set_text(self.summary_detail_text, lines)

    def _iter_current_events(self):
        tab = self.tabs.index(self.tabs.select())
        if tab == 0:
            return self._receipt_events()
        if tab == 1:
            return self._buyout_events()
        return self._payment_events()

    def _export_csv(self):
        events = self._iter_current_events()
        if not events:
            messagebox.showinfo("提示", "没有可导出的数据")
            return
        current = self.tabs.tab(self.tabs.select(), "text")
        fp = filedialog.asksaveasfilename(title="导出 CSV", defaultextension=".csv", filetypes=[("CSV 文件", "*.csv")], initialfile=f"{current}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
        if not fp:
            return
        try:
            with open(fp, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(["类型", "时间", "名称", "单号", "金额", "来源/备注", "操作人"])
                for e in events:
                    writer.writerow([e.get("type", ""), e.get("time", ""), e.get("name", ""), e.get("no", ""), e.get("amount", 0), e.get("source", "") or e.get("note", ""), e.get("operator", "")])
            messagebox.showinfo("成功", f"已导出\n{fp}")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败：{e}")

    def _export_excel(self):
        events = self._iter_current_events()
        if not events:
            messagebox.showinfo("提示", "没有可导出的数据")
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("错误", "缺少 openpyxl 库，请先安装: pip install openpyxl")
            return
        current = self.tabs.tab(self.tabs.select(), "text")
        fp = filedialog.asksaveasfilename(title="导出 Excel", defaultextension=".xlsx", filetypes=[("Excel 文件", "*.xlsx")], initialfile=f"{current}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        if not fp:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = current
            headers = ["类型", "时间", "名称", "单号", "金额", "来源/备注", "操作人"]
            header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
            header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = Alignment(horizontal="center")
            for row, e in enumerate(events, 2):
                ws.cell(row=row, column=1, value=e.get("type", ""))
                ws.cell(row=row, column=2, value=e.get("time", ""))
                ws.cell(row=row, column=3, value=e.get("name", ""))
                ws.cell(row=row, column=4, value=e.get("no", ""))
                ws.cell(row=row, column=5, value=float(e.get("amount", 0) or 0))
                ws.cell(row=row, column=6, value=e.get("source", "") or e.get("note", ""))
                ws.cell(row=row, column=7, value=e.get("operator", ""))
            for row in range(2, len(events) + 2):
                ws.cell(row=row, column=5).number_format = '"¥"#,##0.00'
            widths = {"A": 10, "B": 20, "C": 18, "D": 16, "E": 12, "F": 26, "G": 12}
            for col, width in widths.items():
                ws.column_dimensions[col].width = width
            wb.save(fp)
            messagebox.showinfo("成功", f"已导出\n{fp}")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败：{e}")
