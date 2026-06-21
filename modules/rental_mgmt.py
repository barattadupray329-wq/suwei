#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""租赁管理模块 - 核心CRUD功能"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
import csv
import json
from theme.colors import DarkTheme
from modules.hardware_mgmt import HardwareDialog, HardwareItemDialog


class RentalDetailDialog:
    """租赁记录详情对话框 - 替代右侧详情面板"""

    def __init__(self, parent_frame, record):
        self.frame = parent_frame
        self.rec = record
        self.dm = parent_frame.dm
        self.app = parent_frame.app
        owner = parent_frame.winfo_toplevel()
        self.win = tk.Toplevel(owner)
        self.win.withdraw()
        self.win.title(f"租赁详情 - {record.get('id', 'N/A')}")
        self.win.geometry("760x560")
        self.win.minsize(700, 520)
        self.win.transient(owner)
        self.win.configure(bg=DarkTheme.BG_PRIMARY)
        self._build()
        # 子窗口按主程序窗口中心对齐
        owner.update_idletasks()
        self._center_on_main(self.win)
        self.win.deiconify()
        self.win.lift()
        self.win.focus_force()
        self.win.grab_set()

    def _center_on_main(self, win, w=None, h=None):
        self.frame._center_on_main(win, w, h)

    def _section_label(self, parent, text, color):
        return tk.Label(parent, text=text, font=("微软雅黑", 12, "bold"),
                        fg=color, bg=DarkTheme.BG_PRIMARY)

    def _action_button(self, parent, text, command, color=DarkTheme.ACCENT_BLUE):
        btn = tk.Button(parent, text=text, font=DarkTheme.FONT_BUTTON, fg="white",
                        bg=color, relief=tk.FLAT, cursor="hand2",
                        command=command, padx=20, pady=10)
        btn.pack(side=tk.LEFT, padx=8)
        DarkTheme.bind_hover(btn, DarkTheme.darken(color, 15))
        return btn

    def _build(self):
        main = tk.Frame(self.win, bg=DarkTheme.BG_PRIMARY)
        main.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        head = tk.Frame(main, bg=DarkTheme.BG_PRIMARY)
        head.pack(fill=tk.X, pady=(0, 8))
        tk.Label(head, text="📋 租赁详情", font=DarkTheme.FONT_SUBTITLE,
                 fg=DarkTheme.ACCENT_CYAN, bg=DarkTheme.BG_PRIMARY).pack(side=tk.LEFT)
        tk.Label(head, text="双栏卡片 · 自适应布局", font=DarkTheme.FONT_SMALL,
                 fg=DarkTheme.TEXT_MUTED, bg=DarkTheme.BG_PRIMARY).pack(side=tk.LEFT, padx=(8, 0), pady=(8, 0))

        notebook = ttk.Notebook(main)
        notebook.pack(fill=tk.BOTH, expand=True)

        page1 = tk.Frame(notebook, bg=DarkTheme.BG_PRIMARY)
        notebook.add(page1, text="基础信息")

        self._build_info_tab(page1)

        btn_row = tk.Frame(main, bg=DarkTheme.BG_PRIMARY)
        btn_row.pack(fill=tk.X, pady=(8, 0))
        self._action_button(btn_row, "关闭", self.win.destroy, DarkTheme.BG_HOVER)

    def _build_info_tab(self, parent):
        canvas = tk.Canvas(parent, bg=DarkTheme.BG_PRIMARY, highlightthickness=0)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        content = tk.Frame(canvas, bg=DarkTheme.BG_PRIMARY)
        content.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        window_id = canvas.create_window((0, 0), window=content, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        def _sync_width(event):
            canvas.itemconfigure(window_id, width=event.width)
        canvas.bind("<Configure>", _sync_width)

        def _wheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _wheel)

        self.dm.refresh_record_business_fields(self.rec)
        renter = self.rec.get("renter", {})
        lease = self.rec.get("lease_info", {})
        st = self.rec.get("status", "")
        hw_summary = self.dm.summarize_hardware(self.rec)
        unpaid = float(self.rec.get('unpaid_amount', 0) or 0)

        top = tk.Frame(content, bg=DarkTheme.BG_PRIMARY)
        top.pack(fill=tk.X, padx=10, pady=(6, 8))

        left = tk.Frame(top, bg=DarkTheme.BG_INPUT, highlightbackground=DarkTheme.BORDER_COLOR, highlightthickness=1)
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 6))
        right = tk.Frame(top, bg=DarkTheme.BG_INPUT, highlightbackground=DarkTheme.BORDER_COLOR, highlightthickness=1)
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(6, 0))

        tk.Label(left, text="👤 租赁人信息", font=DarkTheme.FONT_LABEL, fg=DarkTheme.ACCENT_CYAN, bg=DarkTheme.BG_INPUT).pack(anchor=tk.W, padx=12, pady=(10, 6))
        for lbl, val in [
            ("数量", f"{int(self.rec.get('quantity', 1))} 套"),
            ("租赁人", renter.get("name", "")),
            ("电话", renter.get("phone", "")),
            ("身份证", renter.get("id_card", "未填")),
            ("地址", renter.get("address", "未填")),
            ("状态", st),
        ]:
            self._add_row(left, lbl, val)

        tk.Label(right, text="📅 租赁信息", font=DarkTheme.FONT_LABEL, fg=DarkTheme.ACCENT_BLUE, bg=DarkTheme.BG_INPUT).pack(anchor=tk.W, padx=12, pady=(10, 6))
        for lbl, val in [
            ("起租日期", lease.get("start_date", "")),
            ("到期日期", lease.get("end_date", "")),
            ("月租", f"¥{float(lease.get('monthly_rent', 0)):.2f}"),
            ("总租金", f"¥{float(lease.get('total_rent', 0)):.2f}"),
            ("已付", f"¥{float(self.rec.get('paid_amount', 0)):.2f}"),
            ("未付", f"¥{unpaid:.2f}"),
        ]:
            self._add_row(right, lbl, val)

        hw_card = tk.Frame(content, bg=DarkTheme.BG_INPUT, highlightbackground=DarkTheme.BORDER_COLOR, highlightthickness=1)
        hw_card.pack(fill=tk.X, padx=10, pady=(0, 8))
        tk.Label(hw_card, text="🖥️ 硬件摘要", font=DarkTheme.FONT_LABEL, fg=DarkTheme.ACCENT_PURPLE, bg=DarkTheme.BG_INPUT).pack(anchor=tk.W, padx=12, pady=(10, 6))
        tk.Label(hw_card, text=hw_summary or "未填写硬件配置", font=DarkTheme.FONT_NORMAL,
                 fg=DarkTheme.TEXT_PRIMARY, bg=DarkTheme.BG_INPUT, justify=tk.LEFT, wraplength=660).pack(anchor=tk.W, padx=12, pady=(0, 10), fill=tk.X)

        hint = tk.Label(content, text="提示：如需查看更完整的变更记录和操作历史，可结合列表页双击进入。",
                        font=DarkTheme.FONT_SMALL, fg=DarkTheme.TEXT_MUTED, bg=DarkTheme.BG_PRIMARY)
        hint.pack(anchor=tk.W, padx=12, pady=(2, 8))

    def _build_config_tab(self, parent):
        return

    def _build_history_tab(self, parent):
        return

    def _build_action_tab(self, parent):
        return

    def _build_actions_tab(self, parent):
        return

    def _add_row(self, parent, lbl, val):
        row = tk.Frame(parent, bg=parent.cget("bg"))
        row.pack(fill=tk.X, pady=2, padx=12)
        tk.Label(row, text=f"{lbl}:", font=DarkTheme.FONT_LABEL, fg=DarkTheme.TEXT_SECONDARY, bg=parent.cget("bg"), width=10, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 6))
        tk.Label(row, text=str(val), font=DarkTheme.FONT_NORMAL, fg=DarkTheme.TEXT_PRIMARY, bg=parent.cget("bg"), justify=tk.LEFT, wraplength=300).pack(side=tk.LEFT, fill=tk.X, expand=True)

    def _render_snapshot_summary(self, parent, original, current):
        summary_lines = self._snapshot_diff_lines(current, original)
        text = self._format_snapshot(current)
        tk.Label(parent, text=f"当前：{text}", font=DarkTheme.FONT_NORMAL, fg=DarkTheme.TEXT_PRIMARY, bg=DarkTheme.BG_INPUT, justify=tk.LEFT, wraplength=780).pack(anchor=tk.W, padx=10, pady=(8, 2))
        if summary_lines:
            tk.Label(parent, text=f"变更项 {len(summary_lines)} 条：" + "；".join(summary_lines[:3]), font=DarkTheme.FONT_SMALL, fg=DarkTheme.ACCENT_YELLOW, bg=DarkTheme.BG_INPUT, justify=tk.LEFT, wraplength=780).pack(anchor=tk.W, padx=10, pady=(0, 8))
        else:
            tk.Label(parent, text="配置未检测到差异", font=DarkTheme.FONT_SMALL, fg=DarkTheme.TEXT_MUTED, bg=DarkTheme.BG_INPUT).pack(anchor=tk.W, padx=10, pady=(0, 8))

    def _render_snapshot_view(self, parent, snapshot, compare_snapshot=None, mode="current"):
        text = self._format_snapshot(snapshot)
        diff_lines = self._snapshot_diff_lines(snapshot, compare_snapshot)
        box = tk.Frame(parent, bg=DarkTheme.BG_INPUT)
        box.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        tk.Label(box, text=text, font=DarkTheme.FONT_NORMAL, fg=DarkTheme.TEXT_PRIMARY, bg=DarkTheme.BG_INPUT, justify=tk.LEFT, wraplength=360).pack(anchor=tk.W)
        if diff_lines:
            tk.Label(box, text="差异：", font=DarkTheme.FONT_LABEL, fg=DarkTheme.ACCENT_YELLOW, bg=DarkTheme.BG_INPUT).pack(anchor=tk.W, pady=(8, 2))
            for line in diff_lines[:8]:
                tk.Label(box, text=line, font=DarkTheme.FONT_SMALL, fg=DarkTheme.ACCENT_YELLOW, bg=DarkTheme.BG_INPUT, justify=tk.LEFT, wraplength=360).pack(anchor=tk.W)

    def _format_snapshot(self, snapshot):
        if not snapshot:
            return "未记录"
        if isinstance(snapshot, dict):
            return self.dm.summarize_hardware({"hardware": snapshot})
        return str(snapshot)

    def _snapshot_diff_lines(self, a, b):
        if not isinstance(a, dict) or not isinstance(b, dict):
            return []
        lines = []
        keys = sorted(set(a.keys()) | set(b.keys()))
        for k in keys:
            va = a.get(k)
            vb = b.get(k)
            if va != vb:
                lines.append(f"{k}: {vb or '未填'} → {va or '未填'}")
        return lines

    def _show_change_detail_dialog(self, tree, history):
        return

    def _format_change_item(self, item):
        old_name = item.get("old_part_name", "") or "旧件"
        new_name = item.get("new_part_name", "") or "新件"
        return f"{old_name} → {new_name}"

    def _open_change_dialog(self):
        return

    def _format_hw_change(self, old_hw, new_hw):
        return ""

    def _add_history_tree(self, parent, data, cols, keys):
        return

    def _do_action(self, kind):
        return


class RentalManagementFrame(ttk.Frame):
    """租赁管理"""

    STATUS_COLORS = {
        "在租": DarkTheme.STATUS_ACTIVE, "已逾期": DarkTheme.STATUS_EXPIRED,
        "已退租": DarkTheme.STATUS_RETURNED, "已丢失": DarkTheme.STATUS_LOST,
        "已买断": DarkTheme.STATUS_BOUGHT,
    }

    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.dm = app.data_manager
        self._all, self._shown = [], []
        self._record_index = {}  # ID→记录索引映射，加速查找
        self._right_frame = None  # 右侧面板引用
        self._current_form_refs = {}  # 当前表单控件引用
        # 大数据量优化
        self._page_size = 100  # 每页显示记录数
        self._current_page = 0  # 当前页码
        self._total_pages = 0  # 总页数
        self._build()
        self._refresh()

    def _build(self):
        main = tk.Frame(self, bg=DarkTheme.BG_PRIMARY)
        main.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)

        # 标题栏：美化的标题和统计信息
        title_bar = tk.Frame(main, bg=DarkTheme.BG_PRIMARY)
        title_bar.pack(anchor=tk.W, pady=(0, 16), fill=tk.X)
        tk.Label(title_bar, text="📋 租赁管理", font=(DarkTheme.FONT_TITLE[0], DarkTheme.FONT_TITLE[1] + 2, "bold"),
                 fg=DarkTheme.ACCENT_CYAN, bg=DarkTheme.BG_PRIMARY).pack(side=tk.LEFT, anchor=tk.W)
        self._stat_label = tk.Label(title_bar, text="", font=DarkTheme.FONT_LABEL,
                 fg=DarkTheme.TEXT_SECONDARY, bg=DarkTheme.BG_PRIMARY)
        self._stat_label.pack(side=tk.RIGHT, anchor=tk.E)

        # 单栏布局：只有左侧列表
        left = tk.Frame(main, bg=DarkTheme.BG_PRIMARY)
        left.pack(fill=tk.BOTH, expand=True)
        
        # 不再使用右侧面板，不流出空白区域
        self._right_frame = None

        # 搜索和筛选栏：紧凑版本，节省空间
        ctrl = tk.Frame(left, bg=DarkTheme.BG_CARD, relief=tk.FLAT, highlightthickness=0)
        ctrl.pack(fill=tk.X, pady=(0, 8), padx=0, ipady=4)

        # 搜索 + 筛选在同一行，搜索框较窄
        search_frame = tk.Frame(ctrl, bg=DarkTheme.BG_CARD)
        search_frame.pack(fill=tk.X, padx=8, pady=2)
        tk.Label(search_frame, text="🔍", font=(DarkTheme.FONT_LABEL[0], DarkTheme.FONT_LABEL[1] - 1),
                 fg=DarkTheme.TEXT_SECONDARY, bg=DarkTheme.BG_CARD).pack(side=tk.LEFT, padx=(0, 3))
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *_: self._apply_filter())
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var, width=12, font=DarkTheme.FONT_NORMAL)
        search_entry.pack(side=tk.LEFT, padx=(0, 12), ipady=2)
        
        # 状态筛选
        tk.Label(search_frame, text="⚙️", font=(DarkTheme.FONT_LABEL[0], DarkTheme.FONT_LABEL[1] - 1),
                 fg=DarkTheme.TEXT_SECONDARY, bg=DarkTheme.BG_CARD).pack(side=tk.LEFT, padx=(0, 3))
        self.status_var = tk.StringVar(value="全部")
        self.status_combo = ttk.Combobox(search_frame, textvariable=self.status_var, width=10,
                                         state="readonly", values=["全部", "在租", "已退租", "已丢失", "已买断", "已逾期"])
        self.status_combo.pack(side=tk.LEFT, ipady=2)
        self.status_combo.bind("<<ComboboxSelected>>", lambda *_: self._apply_filter())

        # ── 操作按钮区 ──（紧凑版本，节省空间）
        btn_frame = tk.Frame(left, bg=DarkTheme.BG_CARD)
        btn_frame.pack(fill=tk.X, pady=(0, 8), padx=0, ipady=3)

        action_btns = [
            ("➕ 新增", self.add_new_record, DarkTheme.ACCENT_CYAN),
            ("🗑️ 删除", self.delete_record, DarkTheme.ACCENT_RED),
            ("🔄 续租/退租", self.renew_lease, DarkTheme.ACCENT_BLUE),
            ("📤 导出", self.export_rentals, DarkTheme.ACCENT_GREEN),
        ]
        
        def _pack_action_row(parent, buttons):
            row = tk.Frame(parent, bg=DarkTheme.BG_CARD)
            row.pack(fill=tk.X, pady=2, padx=4)
            for txt, cmd, clr in buttons:
                b = tk.Button(row, text=txt, font=(DarkTheme.FONT_BUTTON[0], DarkTheme.FONT_BUTTON[1] - 2, "bold"),
                              fg="white", bg=clr, relief=tk.RAISED, cursor="hand2", command=cmd,
                              padx=4, pady=3, width=9, activebackground=DarkTheme.darken(clr, 15),
                              bd=1, highlightthickness=0)
                b.pack(side=tk.LEFT, padx=1, fill=tk.BOTH, expand=True)
                DarkTheme.bind_hover(b, DarkTheme.darken(clr, 15))

        # 紧凑一行显示所有按钮
        _pack_action_row(btn_frame, action_btns)

        # 分页容器
        self._page_frame = tk.Frame(left, bg=DarkTheme.BG_PRIMARY)
        self._page_frame.pack(fill=tk.X, pady=(4, 0))

        table_frame = tk.Frame(left, bg=DarkTheme.BG_PRIMARY)
        table_frame.pack(fill=tk.BOTH, expand=True)

        cols = ("ID", "数量", "租赁人", "联系电话", "到期时间", "总租金", "已付", "未付", "逾期(天)", "状态")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=14)
        widths = {"ID": 60, "数量": 40, "租赁人": 80, "联系电话": 100, "到期时间": 100, "总租金": 70, "已付": 70, "未付": 70, "逾期(天)": 70, "状态": 60}
        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=widths.get(c, 80), anchor="center")

        vbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vbar.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.bind("<Double-1>", self.show_detail)
        self.tree.bind("<<TreeviewSelect>>", lambda *_: self._on_tree_select())



    def _show_right_placeholder(self):
        """右侧占位 - 不再使用"""
        return

    def _on_tree_select(self):
        """单击选择只选中记录，不在右侧显示详情；详情通过双击弹窗查看"""
        return

    def _show_detail_panel(self, rec):
        """右侧详情面板已移除；详情通过双击弹窗查看。"""
        return

    def _refresh(self):
        self.dm.check_overdue()
        self._all = self.dm.get_records()
        # 构建索引
        self._record_index = {r.get("id"): r for r in self._all}
        self._current_page = 0  # 重置到第一页
        self._apply_filter()

    def _find_record(self, rid):
        """使用索引加速查找，O(1)复杂度"""
        return self._record_index.get(rid)

    def _apply_filter(self):
        q = self.search_var.get().strip().lower()
        s = self.status_var.get()
        self._shown = []
        for r in self._all:
            if s != "全部" and r.get("status") != s:
                continue
            if q:
                renter = r.get("renter", {})
                haystack = (str(r.get("id", "")) + str(renter.get("name", "")) + 
                            str(renter.get("phone", ""))).lower()
                if q not in haystack:
                    continue
            self._shown.append(r)
        self._current_page = 0  # 重置到第一页
        self._render_tree()

    def _render_tree(self):
        """分页渲染 Treeview，支持大数据量"""
        total = len(self._shown)
        if hasattr(self, "_stat_label"):
            active = sum(1 for r in self._all if r.get("status") == "在租")
            overdue = sum(1 for r in self._all if r.get("status") == "已逾期")
            self._stat_label.configure(text=f"共 {len(self._all)} 单｜在租 {active}｜逾期 {overdue}｜当前 {total}")
        if total == 0:
            for i in self.tree.get_children():
                self.tree.delete(i)
            return

        # 计算分页
        self._total_pages = max(1, (total + self._page_size - 1) // self._page_size)
        self._current_page = min(self._current_page, self._total_pages - 1)

        start = self._current_page * self._page_size
        end = min(start + self._page_size, total)
        page_data = self._shown[start:end]

        # 批量删除并插入
        for i in self.tree.get_children():
            self.tree.delete(i)

        # 批量插入（比逐条插入快）
        items = []
        for r in page_data:
            # 只刷新当前页记录的业务字段，减少计算量
            self.dm.refresh_record_business_fields(r)
            renter = r.get("renter", {})
            lease = r.get("lease_info", {})
            st = r.get("status", "")
            total_rent = lease.get("total_rent", 0) or 0
            paid = r.get("paid_amount", 0) or 0
            unpaid = r.get("unpaid_amount", 0) or 0
            overdue_days = r.get("overdue_days", 0) or 0
            items.append((
                r.get("id", ""),
                int(r.get("quantity", 1)),
                renter.get("name", ""),
                renter.get("phone", ""),
                lease.get("end_date", ""),
                f"¥{total_rent:.2f}",
                f"¥{paid:.2f}",
                f"¥{unpaid:.2f}",
                overdue_days,
                st
            ))

        for vals in items:
            self.tree.insert("", tk.END, values=vals, tags=(vals[9],))
        for st, clr in self.STATUS_COLORS.items():
            self.tree.tag_configure(st, foreground=clr)

        # 更新分页信息
        self._update_pagination(total, start, end)

    def _update_pagination(self, total, start, end):
        """更新分页控件和状态"""
        # 清空旧的分页控件内容
        for w in self._page_frame.winfo_children():
            w.destroy()

        # 上一页按钮
        prev_btn = tk.Button(self._page_frame, text="◀ 上一页", font=DarkTheme.FONT_SMALL,
                            fg="white", bg=DarkTheme.ACCENT_BLUE if self._current_page > 0 else DarkTheme.BG_HOVER,
                            relief=tk.FLAT, cursor="hand2" if self._current_page > 0 else "arrow",
                            command=self._prev_page, padx=8, pady=3)
        prev_btn.pack(side=tk.LEFT, padx=2)
        if self._current_page > 0:
            DarkTheme.bind_hover(prev_btn, DarkTheme.darken(DarkTheme.ACCENT_BLUE, 15))

        # 页码信息
        page_info = f"第 {self._current_page + 1}/{self._total_pages} 页 (共 {total} 条，显示 {start+1}-{end})"
        tk.Label(self._page_frame, text=page_info, font=DarkTheme.FONT_LABEL,
                 fg=DarkTheme.TEXT_SECONDARY, bg=DarkTheme.BG_PRIMARY).pack(side=tk.LEFT, padx=10)

        # 下一页按钮
        next_btn = tk.Button(self._page_frame, text="下一页 ▶", font=DarkTheme.FONT_SMALL,
                            fg="white", bg=DarkTheme.ACCENT_BLUE if self._current_page < self._total_pages - 1 else DarkTheme.BG_HOVER,
                            relief=tk.FLAT, cursor="hand2" if self._current_page < self._total_pages - 1 else "arrow",
                            command=self._next_page, padx=8, pady=3)
        next_btn.pack(side=tk.LEFT, padx=2)
        if self._current_page < self._total_pages - 1:
            DarkTheme.bind_hover(next_btn, DarkTheme.darken(DarkTheme.ACCENT_BLUE, 15))

    def _prev_page(self):
        if self._current_page > 0:
            self._current_page -= 1
            self._render_tree()

    def _next_page(self):
        if self._current_page < self._total_pages - 1:
            self._current_page += 1
            self._render_tree()

    def _selected_id(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("提示", "请先选择一条记录")
            return None
        return self.tree.item(sel[0])["values"][0]


    def _advanced_filter(self):
        """当前版本不使用高级筛选面板。"""
        messagebox.showinfo("提示", "高级筛选功能已移除，请使用顶部搜索和状态筛选。")

    def add_new_record(self):
        """弹出新增租赁记录窗口"""
        self._show_add_form_dialog()

    def import_rentals(self):
        """导入租赁数据。当前版本未提供独立导入向导，避免按钮点击直接报错。"""
        messagebox.showinfo("提示", "当前版本暂未开放租赁导入功能。")

    def export_rentals(self):
        """导出租赁记录。默认导出当前筛选结果。"""
        records = list(self._shown) if self._shown else list(self._all)
        if not records:
            messagebox.showinfo("提示", "当前没有可导出的租赁记录")
            return
        self._batch_export(records, self.winfo_toplevel())

    def open_ai(self):
        """打开 AI 助手入口。"""
        messagebox.showinfo("提示", "AI 助手功能当前未在租赁模块中启用。")

    def renew_lease(self):
        """基于当前选中记录打开续租/退租弹窗"""
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("提示", "请先选择一条记录")
            return
        rid = self.tree.item(sel[0])["values"][0]
        rec = self._find_record(rid)
        if rec:
            self._show_renew_form(rec)
        else:
            messagebox.showwarning("提示", "未找到该记录")

    def edit_record(self):
        """基于当前选中记录打开编辑弹窗"""
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("提示", "请先选择一条记录")
            return
        rid = self.tree.item(sel[0])["values"][0]
        rec = self._find_record(rid)
        if rec:
            self._show_edit_form(rec)
        else:
            messagebox.showwarning("提示", "未找到该记录")

    def delete_record(self):
        rid = self._selected_id()
        if rid is None:
            return
        if not messagebox.askyesno("确认", f"确定删除记录 {rid}？"):
            return
        rec = self._find_record(rid)
        if rec and rec.get("status") in ("在租", "已逾期"):
            self._sync_inventory_for_record(rec, direction=-1, trans_type="租赁删除", note="删除租赁记录释放库存")
        self.dm.delete_record(rid)
        messagebox.showinfo("成功", "已删除")
        self._refresh()

    def show_detail(self, event=None):
        """双击记录时，弹出详情对话框"""
        sel = self.tree.selection()
        if not sel:
            return
        rid = self.tree.item(sel[0])["values"][0]
        rec = self._find_record(rid)
        if rec:
            RentalDetailDialog(self, rec)

    def _show_renew_history(self, rec):
        return

    def _edit_hardware_in_record(self, rec):
        """在记录中编辑硬件，并自动写入更换历史。"""
        hardware_data = rec.get("hardware", {})
        old_hardware = json.loads(json.dumps(hardware_data, ensure_ascii=False)) if hardware_data else {}
        dlg = HardwareDialog(self, hardware_data, data_manager=self.dm)
        result = dlg.show()
        if result is not None:
            rec["hardware"] = result
            self.dm.append_hardware_history(
                rec,
                old_hardware,
                result,
                operator=getattr(self.app, "username", "系统"),
                note="租赁记录内硬件配置变更",
            )
            self._sync_inventory_for_hardware_change(rec, old_hardware, result)
            rec["current_config_snapshot"] = json.loads(json.dumps(result, ensure_ascii=False)) if result else {}
            if not rec.get("original_config_snapshot"):
                rec["original_config_snapshot"] = old_hardware
            rec["config_updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.dm.refresh_record_business_fields(rec)
            self.dm.save()
            self._refresh()

    def _edit_hardware_inline(self, hardware_dict, on_updated=None):
        """从表单面板编辑硬件，必要时回调同步租赁字段。"""
        dlg = HardwareDialog(self, hardware_dict, data_manager=self.dm)
        result = dlg.show()
        if result is not None:
            hardware_dict.clear()
            hardware_dict.update(result)
            if callable(on_updated):
                try:
                    on_updated(hardware_dict)
                except Exception:
                    pass

    # ═══ 高级筛选面板 ═══


    def _apply_filter_inline(self, refs):
        """应用筛选条件到左侧列表"""
        kw = refs.get("keyword", tk.StringVar()).get().strip().lower() if hasattr(refs.get("keyword"), 'get') else refs.get("keyword", "")
        start_from = refs.get("start_date_from", tk.StringVar()).get().strip() if hasattr(refs.get("start_date_from"), 'get') else refs.get("start_date_from", "")
        start_to = refs.get("start_date_to", tk.StringVar()).get().strip() if hasattr(refs.get("start_date_to"), 'get') else refs.get("start_date_to", "")
        end_from = refs.get("end_date_from", tk.StringVar()).get().strip() if hasattr(refs.get("end_date_from"), 'get') else refs.get("end_date_from", "")
        end_to = refs.get("end_date_to", tk.StringVar()).get().strip() if hasattr(refs.get("end_date_to"), 'get') else refs.get("end_date_to", "")
        rent_min_str = refs.get("total_rent_min", tk.StringVar()).get().strip() if hasattr(refs.get("total_rent_min"), 'get') else refs.get("total_rent_min", "")
        rent_max_str = refs.get("total_rent_max", tk.StringVar()).get().strip() if hasattr(refs.get("total_rent_max"), 'get') else refs.get("total_rent_max", "")
        paid_min_str = refs.get("paid_min", tk.StringVar()).get().strip() if hasattr(refs.get("paid_min"), 'get') else refs.get("paid_min", "")
        paid_max_str = refs.get("paid_max", tk.StringVar()).get().strip() if hasattr(refs.get("paid_max"), 'get') else refs.get("paid_max", "")

        rent_min = float(rent_min_str) if rent_min_str else 0
        rent_max = float(rent_max_str) if rent_max_str else float('inf')
        paid_min = float(paid_min_str) if paid_min_str else 0
        paid_max = float(paid_max_str) if paid_max_str else float('inf')

        filtered = []
        for r in self._all:
            if kw:
                renter = r.get("renter", {})
                haystack = (str(r.get("id", "")) + str(renter.get("name", "")) + str(renter.get("phone", ""))).lower()
                if kw not in haystack:
                    continue

            lease = r.get("lease_info", {})
            start_date = lease.get("start_date", "")
            end_date = lease.get("end_date", "")

            if start_from and start_date < start_from:
                continue
            if start_to and start_date > start_to:
                continue
            if end_from and end_date < end_from:
                continue
            if end_to and end_date > end_to:
                continue

            total_rent = float(lease.get("total_rent", 0) or 0)
            paid_amount = float(r.get("paid_amount", 0) or 0)
            if total_rent < rent_min or total_rent > rent_max:
                continue
            if paid_amount < paid_min or paid_amount > paid_max:
                continue

            filtered.append(r)

        self._shown = filtered
        self._render_tree()
        messagebox.showinfo("筛选结果", f"筛选得到 {len(filtered)} 条记录")

    def _clear_filter(self, refs):
        """清空筛选条件"""
        for key, ent in refs.items():
            if hasattr(ent, 'delete'):
                ent.delete(0, tk.END)
        self._shown = list(self._all)
        self._render_tree()
        self.search_var.set("")

    # ═══ 报表面板 ═══

    def _show_report_panel(self, stats, total_rent, paid_amount, unpaid_amount):
        """打开租赁统计报表弹窗。"""
        win = tk.Toplevel(self.winfo_toplevel())
        win.title("租赁统计报表")
        win.geometry("860x620")
        win.transient(self.winfo_toplevel())
        win.grab_set()
        win.configure(bg=DarkTheme.BG_PRIMARY)
        self._center_on_main(win, 860, 620)

        canvas = tk.Canvas(win, bg=DarkTheme.BG_PRIMARY, highlightthickness=0)
        scrollbar = ttk.Scrollbar(win, orient="vertical", command=canvas.yview)
        content = tk.Frame(canvas, bg=DarkTheme.BG_PRIMARY)
        content.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=content, anchor="nw", width=820)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(12, 0), pady=12)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, pady=12)

        hdr = tk.Frame(content, bg=DarkTheme.BG_PRIMARY)
        hdr.pack(fill=tk.X, padx=12, pady=(12, 8))
        tk.Label(hdr, text="📊 租赁统计报表", font=DarkTheme.FONT_SUBTITLE,
                 fg=DarkTheme.ACCENT_CYAN, bg=DarkTheme.BG_PRIMARY).pack(side=tk.LEFT)
        tk.Button(hdr, text="✕", font=DarkTheme.FONT_SMALL, fg=DarkTheme.TEXT_MUTED,
                  bg=DarkTheme.BG_PRIMARY, relief=tk.FLAT, cursor="hand2",
                  command=win.destroy).pack(side=tk.RIGHT)

        cards = tk.Frame(content, bg=DarkTheme.BG_PRIMARY)
        cards.pack(fill=tk.X, padx=12, pady=(0, 12))
        for label, value, color in [("总记录数", stats["total"], DarkTheme.ACCENT_CYAN), ("在租中", stats["active"], DarkTheme.ACCENT_BLUE), ("已逾期", stats["expired"], DarkTheme.ACCENT_RED), ("已退租", stats["returned"], DarkTheme.ACCENT_GREEN), ("已丢失", stats["lost"], DarkTheme.ACCENT_YELLOW), ("已买断", stats["bought"], DarkTheme.ACCENT_PURPLE)]:
            card = tk.Frame(cards, bg=DarkTheme.BG_CARD, highlightbackground=DarkTheme.BORDER_COLOR, highlightthickness=1)
            card.pack(side=tk.LEFT, padx=3, fill=tk.BOTH, expand=True)
            tk.Label(card, text=label, font=DarkTheme.FONT_LABEL, fg=DarkTheme.TEXT_SECONDARY, bg=DarkTheme.BG_CARD).pack(pady=(8, 2))
            tk.Label(card, text=str(value), font=("微软雅黑", 18, "bold"), fg=color, bg=DarkTheme.BG_CARD).pack(pady=(2, 8))

        # 金额概览
        amount_box = tk.Frame(content, bg=DarkTheme.BG_CARD, highlightbackground=DarkTheme.BORDER_COLOR, highlightthickness=1)
        amount_box.pack(fill=tk.X, padx=12, pady=(0, 12))
        tk.Label(amount_box, text="💰 金额概览", font=("微软雅黑", 12, "bold"), fg=DarkTheme.ACCENT_BLUE, bg=DarkTheme.BG_CARD).pack(anchor=tk.W, padx=12, pady=(10, 6))
        for lbl, val, color in [
            ("总租金", total_rent, DarkTheme.ACCENT_BLUE),
            ("已收金额", paid_amount, DarkTheme.ACCENT_GREEN),
            ("未收金额", unpaid_amount, DarkTheme.ACCENT_RED),
        ]:
            row = tk.Frame(amount_box, bg=DarkTheme.BG_CARD)
            row.pack(fill=tk.X, padx=12, pady=2)
            tk.Label(row, text=f"{lbl}：", font=DarkTheme.FONT_LABEL, fg=DarkTheme.TEXT_SECONDARY, bg=DarkTheme.BG_CARD, width=12, anchor=tk.W).pack(side=tk.LEFT)
            tk.Label(row, text=f"¥{val:.2f}", font=("微软雅黑", 14, "bold"), fg=color, bg=DarkTheme.BG_CARD).pack(side=tk.LEFT)

        # 状态分布
        status_box = tk.Frame(content, bg=DarkTheme.BG_CARD, highlightbackground=DarkTheme.BORDER_COLOR, highlightthickness=1)
        status_box.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 12))
        tk.Label(status_box, text="📈 状态分布", font=("微软雅黑", 12, "bold"), fg=DarkTheme.ACCENT_BLUE, bg=DarkTheme.BG_CARD).pack(anchor=tk.W, padx=12, pady=(10, 6))
        total = max(1, stats["total"])
        for label, value, color in [
            ("在租", stats["active"], DarkTheme.STATUS_ACTIVE),
            ("逾期", stats["expired"], DarkTheme.STATUS_EXPIRED),
            ("退租", stats["returned"], DarkTheme.STATUS_RETURNED),
            ("丢失", stats["lost"], DarkTheme.STATUS_LOST),
            ("买断", stats["bought"], DarkTheme.STATUS_BOUGHT),
        ]:
            row = tk.Frame(status_box, bg=DarkTheme.BG_CARD)
            row.pack(fill=tk.X, padx=12, pady=3)
            tk.Label(row, text=f"{label}：{value}", font=DarkTheme.FONT_LABEL, fg=DarkTheme.TEXT_SECONDARY, bg=DarkTheme.BG_CARD, width=12, anchor=tk.W).pack(side=tk.LEFT)
            bar = tk.Frame(row, bg=DarkTheme.BG_INPUT, height=14)
            bar.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(4, 0))
            fill = tk.Frame(bar, bg=color, height=14, width=max(2, int((value / total) * 360)))
            fill.pack(side=tk.LEFT, fill=tk.Y)

    def _export_report(self, stats, total_rent, paid_amount, unpaid_amount):
        fp = filedialog.asksaveasfilename(
            title="导出租赁报表", defaultextension=".csv",
            filetypes=[("CSV 文件", "*.csv"), ("所有文件", "*.*")],
            initialfile=f"租赁报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
        if not fp:
            return
        try:
            with open(fp, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(["项目", "值"])
                for k, v in stats.items():
                    writer.writerow([k, v])
                writer.writerow([])
                writer.writerow(["总租金", f"{total_rent:.2f}"])
                writer.writerow(["已收金额", f"{paid_amount:.2f}"])
                writer.writerow(["未收金额", f"{unpaid_amount:.2f}"])
            messagebox.showinfo("成功", f"报表已导出\n{fp}")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败：{e}")

    def _export_renew_history(self, history, rid):
        return

    def _batch_operations(self):
        return

        def _safe_close(w=win):
            try:
                w.grab_release()
            except Exception:
                pass
            w.destroy()

        win.protocol("WM_DELETE_WINDOW", _safe_close)

        main = tk.Frame(win, bg=DarkTheme.BG_PRIMARY)
        main.pack(fill=tk.BOTH, expand=True, padx=16, pady=14)

        tk.Label(main, text="⚡ 批量操作", font=DarkTheme.FONT_SUBTITLE,
                 fg=DarkTheme.ACCENT_YELLOW, bg=DarkTheme.BG_PRIMARY).pack(anchor=tk.W, pady=(0, 10))

        # 筛选条件
        filter_frame = tk.Frame(main, bg=DarkTheme.BG_PRIMARY)
        filter_frame.pack(fill=tk.X, pady=(0, 10))

        tk.Label(filter_frame, text="操作范围：", font=DarkTheme.FONT_LABEL,
                 fg=DarkTheme.TEXT_SECONDARY, bg=DarkTheme.BG_PRIMARY).pack(anchor=tk.W)
        
        scope_var = tk.StringVar(value="selected")
        for val, txt in [("selected", "已选中项"), ("current_filter", "当前筛选结果"), ("all", "全部记录")]:
            tk.Radiobutton(filter_frame, text=txt, variable=scope_var, value=val,
                          font=DarkTheme.FONT_NORMAL, bg=DarkTheme.BG_PRIMARY,
                          fg=DarkTheme.TEXT_PRIMARY, selectcolor=DarkTheme.ACCENT_BLUE).pack(anchor=tk.W)

        tk.Label(main, text="操作类型：", font=DarkTheme.FONT_LABEL,
                 fg=DarkTheme.TEXT_SECONDARY, bg=DarkTheme.BG_PRIMARY).pack(anchor=tk.W, pady=(10, 4))

        action_var = tk.StringVar(value="status_change")
        action_frame = tk.Frame(main, bg=DarkTheme.BG_PRIMARY)
        action_frame.pack(fill=tk.X, pady=(0, 10))
        
        actions = [
            ("status_change", "📝 批量更改状态"),
            ("export", "📤 批量导出为 Excel"),
            ("renew", "🔄 批量续租（加 1 个月）"),
            ("delete", "🗑️ 批量删除"),
        ]
        for val, txt in actions:
            tk.Radiobutton(action_frame, text=txt, variable=action_var, value=val,
                          font=DarkTheme.FONT_NORMAL, bg=DarkTheme.BG_PRIMARY,
                          fg=DarkTheme.TEXT_PRIMARY, selectcolor=DarkTheme.ACCENT_BLUE).pack(anchor=tk.W)

        # 操作按钮
        btn = tk.Frame(main, bg=DarkTheme.BG_PRIMARY)
        btn.pack(fill=tk.X, pady=(16, 0))
        tk.Button(btn, text="✅ 执行操作", font=DarkTheme.FONT_BUTTON, fg="white",
                  bg=DarkTheme.ACCENT_BLUE, relief=tk.FLAT, cursor="hand2",
                  command=lambda: self._execute_batch(scope_var.get(), action_var.get(), win),
                  padx=14, pady=8).pack(side=tk.LEFT, padx=(0, 8))
        tk.Button(btn, text="取消", font=DarkTheme.FONT_BUTTON, fg="white",
                  bg=DarkTheme.BG_HOVER, relief=tk.FLAT, cursor="hand2",
                  command=_safe_close, padx=14, pady=8).pack(side=tk.LEFT)

    def _execute_batch(self, scope, action, win):
        """执行批量操作"""
        # 获取目标记录
        if scope == "selected":
            selected = self.tree.selection()
            if not selected:
                messagebox.showwarning("提示", "请先在列表中选中记录")
                return
            target_ids = [self.tree.item(iid)["values"][0] for iid in selected]
            target_records = [self._find_record(rid) for rid in target_ids]
        elif scope == "current_filter":
            target_records = list(self._shown)
        else:  # all
            target_records = list(self._all)

        if not target_records:
            messagebox.showinfo("提示", "没有可操作的记录")
            return

        if action == "status_change":
            self._batch_status_change(target_records, win)
        elif action == "export":
            self._batch_export(target_records, win)
        elif action == "renew":
            self._batch_renew(target_records, win)
        elif action == "delete":
            self._batch_delete(target_records, win)

    def _batch_status_change(self, records, win):
        """批量更改状态"""
        status_win = tk.Toplevel(win)
        status_win.title("📝 批量更改状态")
        status_win.geometry("350x180")
        status_win.transient(win)
        status_win.grab_set()
        status_win.configure(bg=DarkTheme.BG_PRIMARY)

        def _safe_close(w=status_win):
            try:
                w.grab_release()
            except Exception:
                pass
            w.destroy()

        status_win.protocol("WM_DELETE_WINDOW", _safe_close)

        main = tk.Frame(status_win, bg=DarkTheme.BG_PRIMARY)
        main.pack(fill=tk.BOTH, expand=True, padx=16, pady=14)

        tk.Label(main, text=f"将 {len(records)} 条记录更改为：", font=DarkTheme.FONT_LABEL,
                 fg=DarkTheme.TEXT_PRIMARY, bg=DarkTheme.BG_PRIMARY).pack(anchor=tk.W, pady=(0, 8))

        status_var = tk.StringVar(value="在租")
        status_combo = ttk.Combobox(main, textvariable=status_var, state="readonly",
                                    values=["在租", "已退租", "已丢失", "已买断", "已逾期"], width=20)
        status_combo.pack(fill=tk.X, pady=(0, 12))

        btn = tk.Frame(main, bg=DarkTheme.BG_PRIMARY)
        btn.pack(fill=tk.X)
        tk.Button(btn, text="✅ 确认", font=DarkTheme.FONT_BUTTON, fg="white",
                  bg=DarkTheme.ACCENT_BLUE, relief=tk.FLAT, cursor="hand2",
                  command=lambda: self._do_batch_status(records, status_var.get(), status_win, win),
                  padx=14, pady=6).pack(side=tk.LEFT, padx=(0, 8))
        tk.Button(btn, text="取消", font=DarkTheme.FONT_BUTTON, fg="white",
                  bg=DarkTheme.BG_HOVER, relief=tk.FLAT, cursor="hand2",
                  command=_safe_close, padx=14, pady=6).pack(side=tk.LEFT)

    def _do_batch_status(self, records, new_status, status_win, batch_win):
        """执行批量状态更改"""
        if not messagebox.askyesno("确认", f"确定将 {len(records)} 条记录更改为「{new_status}」吗？"):
            return
        
        count = 0
        for rec in records:
            old_status = rec.get("status")
            rec["status"] = new_status
            self.dm.refresh_record_business_fields(rec)
            self._sync_inventory_for_status_change(rec, old_status, new_status)
            count += 1
        
        self.dm.save()
        messagebox.showinfo("成功", f"已批量更改 {count} 条记录状态为「{new_status}」")
        status_win.destroy()
        batch_win.destroy()
        self._refresh()

    def _batch_export(self, records, win):
        """按明细模板导出租赁记录。"""
        fp = filedialog.asksaveasfilename(
            title="导出租机明细全表",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
            initialfile=f"{datetime.now().strftime('%Y年%m月%d日')}租机明细全表.xlsx"
        )
        if not fp:
            return

        try:
            from datetime import date
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "23-2024"
            ws.freeze_panes = "A2"

            header_titles = [
                "租赁开始时间", "序号", "原始配置", "数量", "备用", "类别", "租赁时间段", "单价",
                "月租金", "已收款", "租赁结束时间", "滞纳金", "备注", "管理员", "退租信息", "位置",
            ] + ["" for _ in range(18)]

            header_fill = PatternFill("solid", fgColor="4472C4")
            header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
            header_border = Border(
                left=Side(style="thin", color="D9D9D9"),
                right=Side(style="thin", color="D9D9D9"),
                top=Side(style="thin", color="D9D9D9"),
                bottom=Side(style="thin", color="D9D9D9"),
            )
            normal_border = header_border
            overdue_fill = PatternFill("solid", fgColor="FDE9E7")
            due_soon_fill = PatternFill("solid", fgColor="FFF2CC")
            active_fill = PatternFill("solid", fgColor="E2F0D9")
            returned_fill = PatternFill("solid", fgColor="E7E6E6")

            # 模板头部
            for col_idx, title in enumerate(header_titles, 1):
                cell = ws.cell(row=1, column=col_idx, value=title)
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = header_fill
                cell.border = header_border

            ws.auto_filter.ref = f"A1:{get_column_letter(len(header_titles))}{len(records) + 1}"

            def _fmt_date(value):
                if not value:
                    return ""
                text = str(value).strip()
                return text.replace("-", ".") if "-" in text else text

            def _safe_float(value):
                try:
                    return float(value or 0)
                except Exception:
                    return 0.0

            def _parse_date(text):
                if not text:
                    return None
                t = str(text).strip().replace(".", "-").split(" ")[0]
                for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
                    try:
                        return datetime.strptime(t, fmt).date()
                    except Exception:
                        continue
                return None

            def _join_hardware(rec):
                hw = rec.get("hardware", {}) if isinstance(rec, dict) else {}
                items = hw.get("items", []) if isinstance(hw, dict) else []
                parts = []
                for item in items:
                    name = item.get("name") or item.get("device_type") or "设备"
                    qty = item.get("quantity", 1)
                    spec_bits = [item.get(k, "") for k in ("cpu", "motherboard", "ram", "disk", "gpu", "brand_model", "monitor_brand", "monitor_model")]
                    spec = " ".join([str(x) for x in spec_bits if x])
                    parts.append(f"{name}×{qty}{(' ' + spec) if spec else ''}")
                return "；".join(parts)

            today = date.today()
            for row_idx, rec in enumerate(records, 2):
                lease = rec.get("lease_info", {}) or {}
                start = _fmt_date(lease.get("start_date", ""))
                end = _fmt_date(lease.get("end_date", ""))
                end_date = _parse_date(end)
                monthly = _safe_float(lease.get("monthly_rent", 0))
                total = _safe_float(lease.get("total_rent", 0))
                paid = _safe_float(rec.get("paid_amount", 0))
                unpaid = _safe_float(rec.get("unpaid_amount", total - paid))
                notes = rec.get("notes", "") or _join_hardware(rec)
                status = rec.get("status", "")
                row_fill = active_fill
                if status in ("已退租", "已丢失", "已买断"):
                    row_fill = returned_fill
                elif end_date:
                    if end_date < today:
                        row_fill = overdue_fill
                    elif (end_date - today).days <= 7:
                        row_fill = due_soon_fill

                values = [
                    start,
                    rec.get("id", ""),
                    _join_hardware(rec),
                    int(rec.get("quantity", 1) or 1),
                    "",
                    status,
                    f"{start}-{end}" if start or end else "",
                    monthly,
                    total,
                    paid,
                    end,
                    max(0.0, unpaid - 0),
                    notes,
                    getattr(self.app, "username", "系统"),
                    rec.get("settlement_amount", "") or "",
                    rec.get("location", "") or rec.get("renter", {}).get("address", ""),
                ]
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = Font(name="微软雅黑", size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = normal_border
                    cell.fill = row_fill
                    if col_idx in (8, 9, 10, 12):
                        cell.number_format = '"¥"#,##0.00'

            for col_idx in range(1, len(header_titles) + 1):
                col_letter = get_column_letter(col_idx)
                cell_values = [ws.cell(row=r, column=col_idx).value for r in range(1, len(records) + 2)]
                max_len = max(len(str(v)) if v is not None else 0 for v in cell_values)
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 8), 40)

            widths = {
                "A": 14, "B": 14, "C": 42, "D": 10, "E": 8, "F": 12, "G": 18, "H": 12,
                "I": 12, "J": 12, "K": 14, "L": 12, "M": 36, "N": 12, "O": 14, "P": 18,
            }
            for col, width in widths.items():
                ws.column_dimensions[col].width = width

            wb.save(fp)
            messagebox.showinfo("成功", f"已按明细模板导出 {len(records)} 条记录\n{fp}")
            win.destroy()
        except ImportError:
            messagebox.showerror("错误", "缺少 openpyxl 库，请先安装: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败：{e}")

    def _sync_inventory_for_status_change(self, rec, old_status, new_status):
        hardware = rec.get("hardware", {}) or {}
        items = hardware.get("items") if isinstance(hardware, dict) else None
        if not items:
            return
        for item in items:
            model = self._find_inventory_model(item)
            if not model:
                continue
            qty = int(float(item.get("quantity", 1) or 1))
            mid = model.get("id")
            current = self.dm.get_model_by_id(mid)
            if not current:
                continue
            stock_total = int(current.get("stock_total") or 0)
            stock_rented = int(current.get("stock_rented") or 0)
            if old_status in ("在租", "已逾期") and new_status in ("已退租", "已丢失", "已买断"):
                stock_rented = max(0, stock_rented - qty)
            elif old_status in ("已退租", "已丢失", "已买断") and new_status in ("在租", "已逾期"):
                stock_rented += qty
            self.dm.update_model(mid, stock_total=stock_total, stock_rented=stock_rented)
            self.dm.add_inventory_transaction(
                model_id=mid,
                trans_type=f"状态同步:{old_status}->{new_status}",
                quantity=qty,
                related_rental_id=rec.get("id", ""),
                operator=getattr(self.app, "username", "系统"),
                note="批量状态变更同步库存",
            )

    def _sync_inventory_for_hardware_change(self, rec, old_hardware, new_hardware):
        old_items = (old_hardware or {}).get("items") if isinstance(old_hardware, dict) else None
        new_items = (new_hardware or {}).get("items") if isinstance(new_hardware, dict) else None
        if not old_items and not new_items:
            return
        old_map = {}
        for item in old_items or []:
            model = self._find_inventory_model(item)
            if model:
                old_map[model.get("id")] = int(float(item.get("quantity", 1) or 1))
        new_map = {}
        for item in new_items or []:
            model = self._find_inventory_model(item)
            if model:
                new_map[model.get("id")] = int(float(item.get("quantity", 1) or 1))
        all_ids = set(old_map) | set(new_map)
        for mid in all_ids:
            current = self.dm.get_model_by_id(mid)
            if not current:
                continue
            stock_total = int(current.get("stock_total") or 0)
            stock_rented = int(current.get("stock_rented") or 0)
            delta = new_map.get(mid, 0) - old_map.get(mid, 0)
            if delta != 0:
                stock_rented = max(0, stock_rented + delta)
                self.dm.update_model(mid, stock_total=stock_total, stock_rented=stock_rented)
                self.dm.add_inventory_transaction(
                    model_id=mid,
                    trans_type="硬件变更同步",
                    quantity=delta,
                    related_rental_id=rec.get("id", ""),
                    operator=getattr(self.app, "username", "系统"),
                    note="租赁记录硬件变更同步库存",
                )

    def _find_inventory_model(self, item):
        model_id = item.get("model_id")
        if model_id:
            try:
                model = self.dm.get_model_by_id(int(model_id))
                if model:
                    return model
            except Exception:
                pass
        query = " ".join([str(item.get("cpu", "")), str(item.get("motherboard", "")), str(item.get("ram", "")), str(item.get("disk", "")), str(item.get("gpu", "")), str(item.get("monitor_brand", "")), str(item.get("monitor_model", "")), str(item.get("brand_model", "")), str(item.get("model", ""))]).strip()
        if not query:
            query = str(item.get("device_type", ""))
        matches = self.dm.search_models(query, limit=10)
        if matches:
            return matches[0]
        return None

    def _sync_inventory_for_record(self, rec, direction=1, trans_type="租赁同步", note=""):
        hardware = rec.get("hardware", {}) or {}
        items = hardware.get("items") if isinstance(hardware, dict) else None
        if not items:
            return
        operator = getattr(self.app, "username", "系统")
        for item in items:
            model = self._find_inventory_model(item)
            if not model:
                continue
            mid = model.get("id")
            if not mid:
                continue
            qty = int(float(item.get("quantity", 1) or 1)) * int(direction)
            current = self.dm.get_model_by_id(mid)
            if not current:
                continue
            stock_total = int(current.get("stock_total") or 0)
            stock_rented = max(0, int(current.get("stock_rented") or 0) + qty)
            self.dm.update_model(mid, stock_total=stock_total, stock_rented=stock_rented)
            self.dm.add_inventory_transaction(
                model_id=mid,
                trans_type=trans_type,
                quantity=qty,
                related_rental_id=rec.get("id", ""),
                operator=operator,
                note=note or trans_type,
            )

    def _batch_renew(self, records, win):
        """批量续租（加 1 个月）"""
        if not messagebox.askyesno("确认续租", f"确定将 {len(records)} 条在租记录续租 1 个月吗？"):
            return
        
        count = 0
        for rec in records:
            if rec.get("status") not in ("在租", "已逾期"):
                continue
            lease = rec.get("lease_info", {})
            end_str = lease.get("end_date", "")
            if not end_str:
                continue
            
            try:
                cur_end = datetime.strptime(end_str, "%Y-%m-%d")
                new_end = cur_end + timedelta(days=30)
                monthly_rent = float(lease.get("monthly_rent", 0) or 0)
                
                lease["end_date"] = new_end.strftime("%Y-%m-%d")
                lease["total_rent"] = float(lease.get("total_rent", 0) or 0) + monthly_rent
                
                rec.setdefault("renew_history", []).append({
                    "renew_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "renew_time": 1.0, "renew_unit": "月", "renew_amount": monthly_rent,
                    "old_end_date": cur_end.strftime("%Y-%m-%d"),
                    "new_end_date": new_end.strftime("%Y-%m-%d"),
                    "operator": self.app.username or "系统",
                })
                self.dm.refresh_record_business_fields(rec)
                count += 1
            except Exception:
                continue
        
        if count > 0:
            self.dm.save()
            messagebox.showinfo("成功", f"已批量续租 {count} 条记录")
        else:
            messagebox.showinfo("提示", "没有可续租的记录")
        
        win.destroy()
        self._refresh()

    def _batch_delete(self, records, win):
        """批量删除"""
        if not messagebox.askyesno("确认删除", f"确定删除 {len(records)} 条记录吗？此操作不可恢复！"):
            return
        
        count = 0
        for rec in records:
            rid = rec.get("id", "")
            if rid:
                self.dm.delete_record(rid)
                count += 1
        
        messagebox.showinfo("成功", f"已批量删除 {count} 条记录")
        win.destroy()
        self._refresh()

    def _center_on_main(self, win, w=None, h=None):
        """子窗口按主程序窗口当前可见区域中心对齐。"""
        win.update_idletasks()
        main_root = getattr(self, "app", None).root if getattr(self, "app", None) and getattr(self.app, "root", None) else self.winfo_toplevel()
        main_root.update_idletasks()

        if w is None:
            w = win.winfo_reqwidth() or win.winfo_width()
        if h is None:
            h = win.winfo_reqheight() or win.winfo_height()

        mw = main_root.winfo_width()
        mh = main_root.winfo_height()
        if mw <= 1 or mh <= 1:
            geo = main_root.winfo_geometry()
            try:
                size = geo.split("+")[0]
                mw, mh = map(int, size.split("x", 1))
            except Exception:
                mw, mh = main_root.winfo_screenwidth() // 2, main_root.winfo_screenheight() // 2

        mx = main_root.winfo_rootx()
        my = main_root.winfo_rooty()
        x = mx + (mw - w) // 2
        y = my + (mh - h) // 2
        win.geometry(f"{w}x{h}+{x}+{y}")

    def _clear_right_panel(self):
        """清空右侧面板"""
        self._current_form_refs = {}
        for w in self._right_frame.winfo_children():
            w.destroy()

    def _show_add_form_dialog(self):
        """弹出新增租赁记录窗口（两列卡片布局）"""
        win = tk.Toplevel(self.winfo_toplevel())
        win.title("新增租赁记录")
        win.geometry("1240x760")
        win.minsize(1120, 680)
        win.configure(bg=DarkTheme.BG_PRIMARY)
        self._center_on_main(win, 1240, 760)

        style = ttk.Style(win)
        style.configure("TEntry", padding=(4, 1), relief="flat")
        style.configure("TCombobox", padding=(4, 1), relief="flat")

        hdr = tk.Frame(win, bg=DarkTheme.BG_PRIMARY)
        hdr.pack(fill=tk.X, padx=18, pady=(14, 10))
        tk.Label(hdr, text="➡️ 新增租赁记录", font=DarkTheme.FONT_SUBTITLE,
                 fg=DarkTheme.ACCENT_CYAN, bg=DarkTheme.BG_PRIMARY).pack(side=tk.LEFT)

        main_frame = tk.Frame(win, bg=DarkTheme.BG_PRIMARY)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=18, pady=(0, 14))

        # 两列布局：左侧信息卡片，右侧硬件卡片
        left_col = tk.Frame(main_frame, bg=DarkTheme.BG_PRIMARY, width=300)
        left_col.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
        left_col.pack_propagate(False)

        right_col = tk.Frame(main_frame, bg=DarkTheme.BG_PRIMARY, width=880)
        right_col.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(10, 0))
        right_col.pack_propagate(False)

        refs = {}
        cbg = DarkTheme.BG_CARD

        # ═══ 左侧：信息卡片（包含客户、租赁、付款、状态） ═══
        info_card = tk.Frame(left_col, bg=cbg, highlightbackground=DarkTheme.BORDER_COLOR, highlightthickness=1)
        info_card.pack(fill=tk.BOTH, expand=True)

        # 可滚动的内容区
        canvas = tk.Canvas(info_card, bg=cbg, highlightthickness=0)
        scrollbar = ttk.Scrollbar(info_card, orient="vertical", command=canvas.yview)
        info_content = tk.Frame(canvas, bg=cbg)
        info_content.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        info_window = canvas.create_window((0, 0), window=info_content, anchor="nw", width=280)
        def _sync_info_width(event):
            canvas.itemconfigure(info_window, width=event.width)
        canvas.bind("<Configure>", _sync_info_width)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # 标题
        tk.Label(info_content, text="📋 基本信息", font=DarkTheme.FONT_SUBTITLE,
                 fg=DarkTheme.ACCENT_CYAN, bg=cbg).pack(anchor=tk.W, padx=8, pady=(4, 2))

        # ① 客户信息部分
        tk.Label(info_content, text="👤 租赁人", font=DarkTheme.FONT_LABEL,
                 fg=DarkTheme.ACCENT_CYAN, bg=cbg).pack(anchor=tk.W, padx=8, pady=(2, 1))
        refs['name_e'] = self._make_field_in(info_content, "姓名*", "", bg=cbg, width=12, compact=True)
        refs['phone_e'] = self._make_field_in(info_content, "电话*", "", bg=cbg, width=12, compact=True)
        refs['id_e'] = self._make_field_in(info_content, "身份证", "", bg=cbg, width=12, compact=True)
        refs['name_e'].pack_configure(fill=tk.X, expand=True)
        refs['phone_e'].pack_configure(fill=tk.X, expand=True)
        refs['id_e'].pack_configure(fill=tk.X, expand=True)
        addr_row = tk.Frame(info_content, bg=cbg)
        addr_row.pack(fill=tk.X, pady=1, padx=10)
        tk.Label(addr_row, text="地址", font=DarkTheme.FONT_LABEL, fg=DarkTheme.TEXT_SECONDARY,
                 bg=cbg, width=7, anchor=tk.W).pack(side=tk.LEFT, padx=(0, 4))
        addr_t = tk.Text(addr_row, height=2, font=DarkTheme.FONT_NORMAL, wrap=tk.WORD,
                        bg=DarkTheme.BG_INPUT, fg=DarkTheme.TEXT_PRIMARY, insertbackground=DarkTheme.TEXT_PRIMARY)
        addr_t.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 4))
        refs['addr_t'] = addr_t

        # ② 租赁信息部分
        tk.Label(info_content, text="📅 租赁期限", font=DarkTheme.FONT_LABEL,
                 fg=DarkTheme.ACCENT_BLUE, bg=cbg).pack(anchor=tk.W, padx=8, pady=(1, 0))
        refs['start_e'] = self._make_field_in(info_content, "起租*", datetime.now().strftime("%Y-%m-%d"), bg=cbg, width=12, compact=True)
        refs['months_e'] = self._make_field_in(info_content, "月数*", "1", bg=cbg, width=12, compact=True)
        refs['end_e'] = self._make_field_in(info_content, "到期*", (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d"), bg=cbg, width=12, compact=True)
        refs['start_e'].pack_configure(fill=tk.X, expand=True)
        refs['months_e'].pack_configure(fill=tk.X, expand=True)
        refs['end_e'].pack_configure(fill=tk.X, expand=True)

        # ③ 租金与数量部分（只读，由硬件自动计算）
        tk.Label(info_content, text="💰 租金与数量", font=DarkTheme.FONT_LABEL,
                 fg=DarkTheme.ACCENT_BLUE, bg=cbg).pack(anchor=tk.W, padx=8, pady=(2, 1))
        qty_row = tk.Frame(info_content, bg=cbg)
        qty_row.pack(fill=tk.X, pady=1, padx=8)
        tk.Label(qty_row, text="数量*", font=DarkTheme.FONT_LABEL, fg=DarkTheme.TEXT_SECONDARY,
                 bg=cbg, width=6, anchor=tk.W).pack(side=tk.LEFT, padx=(0, 3))
        refs['quantity_e'] = ttk.Entry(qty_row, width=16, font=DarkTheme.FONT_NORMAL, state="readonly")
        refs['quantity_e'].pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 3))
        refs['quantity_e'].insert(0, "0")

        mr_row = tk.Frame(info_content, bg=cbg)
        mr_row.pack(fill=tk.X, pady=1, padx=8)
        tk.Label(mr_row, text="月租*", font=DarkTheme.FONT_LABEL, fg=DarkTheme.TEXT_SECONDARY,
                 bg=cbg, width=6, anchor=tk.W).pack(side=tk.LEFT, padx=(0, 3))
        refs['monthly_e'] = ttk.Entry(mr_row, width=16, font=DarkTheme.FONT_NORMAL, state="readonly")
        refs['monthly_e'].pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 3))
        refs['monthly_e'].insert(0, "0")

        refs['total_e'] = self._make_field_in(info_content, "总租金*", "0", bg=cbg, width=12, compact=True)
        refs['total_e'].pack_configure(fill=tk.X, expand=True)

        # ④ 付款与状态部分
        tk.Label(info_content, text="💳 付款与状态", font=DarkTheme.FONT_LABEL,
                 fg=DarkTheme.ACCENT_GREEN, bg=cbg).pack(anchor=tk.W, padx=8, pady=(1, 0))
        refs['deposit_e'] = self._make_field_in(info_content, "押金", "0", bg=cbg, width=10, compact=True)
        refs['paid_e'] = self._make_field_in(info_content, "已付金额", "0", bg=cbg, width=10, compact=True)
        refs['deposit_e'].pack_configure(fill=tk.X, expand=True)
        refs['paid_e'].pack_configure(fill=tk.X, expand=True)
        sr = tk.Frame(info_content, bg=cbg)
        sr.pack(fill=tk.X, pady=1, padx=14)
        tk.Label(sr, text="状态", font=DarkTheme.FONT_LABEL, fg=DarkTheme.TEXT_SECONDARY,
                 bg=cbg, width=8, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 4))
        refs['status_var'] = tk.StringVar(value="在租")
        sc = ttk.Combobox(sr, textvariable=refs['status_var'], state="readonly", width=16,
                         values=["在租", "已退租", "已丢失", "已买断", "已逾期"])
        sc.pack(side=tk.LEFT, fill=tk.X, expand=True)
        tk.Label(info_content, text="💡 提示：数量和月租由硬件清单自动计算。",
                 font=DarkTheme.FONT_SMALL, fg=DarkTheme.TEXT_MUTED, bg=cbg).pack(anchor=tk.W, padx=14, pady=(12, 12))

        # ═══ 右侧：硬件卡片 ═══
        hw_card = tk.Frame(right_col, bg=cbg, highlightbackground=DarkTheme.BORDER_COLOR, highlightthickness=1)
        hw_card.pack(fill=tk.BOTH, expand=True)

        tk.Label(hw_card, text="⚙️ 硬件配置", font=DarkTheme.FONT_SUBTITLE,
                 fg=DarkTheme.ACCENT_PURPLE, bg=cbg).pack(anchor=tk.W, padx=14, pady=(14, 10))
        refs['hardware_data'] = {"items": []}
        refs['last_hw_price'] = 120

        hw_body = tk.Frame(hw_card, bg=cbg)
        hw_body.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)
        hw_canvas = tk.Canvas(hw_body, bg=cbg, highlightthickness=0)
        hw_scrollbar = ttk.Scrollbar(hw_body, orient="vertical", command=hw_canvas.yview)
        hw_content = tk.Frame(hw_canvas, bg=cbg)
        hw_content.bind("<Configure>", lambda e: hw_canvas.configure(scrollregion=hw_canvas.bbox("all")))
        hw_window = hw_canvas.create_window((0, 0), window=hw_content, anchor="nw")
        def _sync_hw_width(event):
            hw_canvas.itemconfigure(hw_window, width=event.width)
        hw_canvas.bind("<Configure>", _sync_hw_width)
        hw_canvas.configure(yscrollcommand=hw_scrollbar.set)
        hw_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        hw_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        def _hw_mousewheel(event):
            hw_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        hw_canvas.bind_all("<MouseWheel>", _hw_mousewheel)

        hw_root = hw_content
        refs['hardware_data']['items'] = [{
            "device_type": "台式机",
            "name": "台式机",
            "quantity": 1,
            "unit_rent": 120,
            "cpu": "",
            "motherboard": "",
            "ram": "",
            "disk": "",
            "gpu": "",
            "case": "",
            "psu": "",
            "fan": "",
        }]

        def _update_from_hardware():
            """根据硬件清单自动更新数量和月租"""
            hw = refs.get('hardware_data', {})
            items = hw.get("items", []) if isinstance(hw, dict) else []
            qty = 0
            total_rent = 0.0
            for item in items:
                q = float(item.get("quantity", 0) or 0)
                r = float(item.get("unit_rent", 0) or 0)
                qty += q
                total_rent += r * q
            # 更新数量（只读）
            refs['quantity_e'].config(state="normal")
            refs['quantity_e'].delete(0, tk.END)
            refs['quantity_e'].insert(0, str(int(qty)) if qty == int(qty) else f"{qty:g}")
            refs['quantity_e'].config(state="readonly")
            # 更新月租（只读）
            refs['monthly_e'].config(state="normal")
            refs['monthly_e'].delete(0, tk.END)
            refs['monthly_e'].insert(0, f"{total_rent:.2f}")
            refs['monthly_e'].config(state="readonly")
            _refresh_hw_display()
            _auto_calc()

        def _build_hw_item(hw_type="台式机", qty=1, price=120):
            item = {"device_type": hw_type, "name": hw_type, "quantity": qty, "unit_rent": price}
            if hw_type == "台式机":
                item.update({"cpu": "", "motherboard": "", "ram": "", "disk": "", "gpu": "", "case": "", "psu": "", "fan": ""})
            elif hw_type == "笔记本":
                item.update({"brand_model": "", "screen_size": "", "cpu": "", "ram": "", "disk": ""})
            elif hw_type == "显示器":
                item.update({"monitor_brand": "", "monitor_model": "", "screen_size": "", "resolution": "", "refresh_rate": "", "condition": ""})
            return item

        def _append_default_hw_if_empty():
            hw = refs.get('hardware_data', {})
            items = hw.get('items', []) if isinstance(hw, dict) else []
            if not items:
                items.append(_build_hw_item("台式机", 1, refs.get('last_hw_price', 120)))
                hw['items'] = items
                _update_from_hardware()

        # 硬件清单区域（默认一行，右键下拉详细配置）
        hw_list_frame = tk.Frame(hw_content, bg=DarkTheme.BG_INPUT)
        hw_list_frame.pack(fill=tk.X, padx=14, pady=(0, 4))
        tk.Label(hw_list_frame, text="清单（右键新增配置）", font=DarkTheme.FONT_LABEL, fg=DarkTheme.TEXT_SECONDARY,
                 bg=DarkTheme.BG_INPUT).pack(anchor=tk.W, padx=4, pady=(2, 1))
        hw_tree_frame = tk.Frame(hw_list_frame, bg=DarkTheme.BG_INPUT)
        hw_tree_frame.pack(fill=tk.X, padx=4, pady=(0, 2))
        
        hw_tree = ttk.Treeview(hw_tree_frame, columns=["名称", "数量", "单价/月", "状态"], show="headings", height=5)
        hw_tree.heading("名称", text="配置名称")
        hw_tree.heading("数量", text="数量")
        hw_tree.heading("单价/月", text="单价/月")
        hw_tree.heading("状态", text="状态")
        hw_tree.column("名称", width=260)
        hw_tree.column("数量", width=60, anchor="center")
        hw_tree.column("单价/月", width=90, anchor="center")
        hw_tree.column("状态", width=80, anchor="center")
        hw_tree_scroll = ttk.Scrollbar(hw_tree_frame, orient="vertical", command=hw_tree.yview)
        hw_tree.configure(yscrollcommand=hw_tree_scroll.set)
        hw_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        hw_tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        refs['hw_tree'] = hw_tree

        # 删除硬件项目
        def _delete_hw_item(idx):
            items = refs['hardware_data'].get("items", [])
            if 0 <= idx < len(items):
                items.pop(idx)
                _update_from_hardware()

        def _find_hw_item(idx):
            items = refs['hardware_data'].get("items", [])
            if 0 <= idx < len(items):
                return items[idx]
            return None

        # 编辑硬件项目（右键展开详细编辑）
        edit_panel = tk.Frame(hw_content, bg=DarkTheme.BG_INPUT)
        edit_panel_visible = {'shown': False, 'idx': None}
        edit_panel.pack(fill=tk.X, padx=14, pady=(0, 4))
        edit_panel.pack_forget()
        edit_scroll = None

        def _hide_edit_panel():
            edit_panel_visible['shown'] = False
            edit_panel_visible['idx'] = None
            for w in edit_panel.winfo_children():
                w.destroy()
            edit_panel.pack_forget()

        def _show_edit_panel(idx):
            items = refs['hardware_data'].get("items", [])
            if not (0 <= idx < len(items)):
                return
            item = dict(items[idx])
            dev_type = item.get('device_type', item.get('name', '台式机'))
            edit_panel_visible['shown'] = True
            edit_panel_visible['idx'] = idx
            for w in edit_panel.winfo_children():
                w.destroy()

            tk.Label(edit_panel, text=f"✏️ 编辑：{dev_type}", font=DarkTheme.FONT_LABEL,
                     fg=DarkTheme.ACCENT_CYAN, bg=DarkTheme.BG_INPUT).pack(anchor=tk.W, padx=10, pady=(4, 2))

            type_row = tk.Frame(edit_panel, bg=DarkTheme.BG_INPUT)
            type_row.pack(fill=tk.X, padx=10, pady=(2, 4))
            tk.Label(type_row, text="类型", font=DarkTheme.FONT_SMALL, fg=DarkTheme.TEXT_SECONDARY,
                     bg=DarkTheme.BG_INPUT, width=10, anchor=tk.W).pack(side=tk.LEFT)
            type_var = tk.StringVar(value=dev_type)
            type_combo = ttk.Combobox(type_row, textvariable=type_var, state="readonly",
                                      values=["台式机", "笔记本", "显示器", "一体机"], font=DarkTheme.FONT_SMALL)
            type_combo.pack(side=tk.LEFT, fill=tk.X, expand=True)

            def _field_row(parent, label, value="", width=20):
                row = tk.Frame(parent, bg=DarkTheme.BG_INPUT)
                row.pack(fill=tk.X, padx=10, pady=1)
                tk.Label(row, text=label, font=DarkTheme.FONT_SMALL, fg=DarkTheme.TEXT_SECONDARY,
                         bg=DarkTheme.BG_INPUT, width=8, anchor=tk.W).pack(side=tk.LEFT)
                ent = ttk.Entry(row, font=("微软雅黑", 9))
                ent.pack(side=tk.LEFT, fill=tk.X, expand=True)
                ent.insert(0, str(value))
                return ent

            controls = {}
            field_container = tk.Frame(edit_panel, bg=DarkTheme.BG_INPUT)
            field_container.pack(fill=tk.BOTH, expand=True)

            def _render_fields(kind):
                for w in field_container.winfo_children():
                    w.destroy()
                controls.clear()
                if kind == "台式机":
                    fields = [("cpu", "CPU"), ("motherboard", "主板"), ("ram", "内存"), ("disk", "硬盘"),
                              ("gpu", "显卡"), ("case", "机箱"), ("psu", "电源"), ("fan", "风扇/散热")]
                elif kind == "笔记本":
                    fields = [("brand_model", "品牌型号"), ("screen_size", "屏幕尺寸(寸)"), ("cpu", "CPU"), ("ram", "内存"), ("disk", "硬盘")]
                elif kind == "显示器":
                    fields = [("monitor_brand", "品牌"), ("monitor_model", "型号"), ("screen_size", "尺寸(寸)"),
                              ("resolution", "分辨率"), ("refresh_rate", "刷新率(Hz)"), ("condition", "新旧程度")]
                elif kind == "一体机":
                    fields = [("brand_model", "品牌型号"), ("screen_size", "屏幕尺寸(寸)"), ("cpu", "CPU"), ("ram", "内存"), ("disk", "硬盘")]
                else:
                    fields = [("name", "名称"), ("specs", "规格")]
                for key, label in fields:
                    controls[key] = _field_row(field_container, label, item.get(key, ""))

            current_type = {'value': dev_type}
            _render_fields(dev_type)

            def _on_type_change(event=None):
                kind = type_var.get().strip()
                current_type['value'] = kind
                _render_fields(kind)

            type_combo.bind("<<ComboboxSelected>>", _on_type_change)

            qty_e = _field_row(edit_panel, "数量", item.get('quantity', 1))
            price_e = _field_row(edit_panel, "单价/月", item.get('unit_rent', 0))

            btns = tk.Frame(edit_panel, bg=DarkTheme.BG_INPUT)
            btns.pack(fill=tk.X, padx=10, pady=(6, 8))

            def _save_inline_edit():
                try:
                    updated = dict(items[idx])
                    for key, ent in controls.items():
                        val = ent.get().strip()
                        if val:
                            updated[key] = val
                    dev_type = type_var.get().strip() or dev_type
                    updated['device_type'] = dev_type
                    if dev_type in ("台式机", "笔记本", "显示器", "一体机"):
                        updated['name'] = dev_type
                    else:
                        updated['name'] = updated.get('name') or dev_type
                    updated['quantity'] = float(qty_e.get().strip() or 1)
                    updated['unit_rent'] = float(price_e.get().strip() or 0)
                    items[idx] = updated
                    refs['last_hw_price'] = updated['unit_rent']
                    _hide_edit_panel()
                    _update_from_hardware()
                except ValueError:
                    messagebox.showwarning("提示", "数量和单价必须为数字")

            tk.Button(btns, text="保存", font=DarkTheme.FONT_SMALL, fg="white", bg=DarkTheme.ACCENT_BLUE,
                      relief=tk.FLAT, cursor="hand2", command=_save_inline_edit, padx=10, pady=4).pack(side=tk.LEFT, padx=(0, 6))
            tk.Button(btns, text="取消", font=DarkTheme.FONT_SMALL, fg="white", bg=DarkTheme.BG_HOVER,
                      relief=tk.FLAT, cursor="hand2", command=_hide_edit_panel, padx=10, pady=4).pack(side=tk.LEFT)
            edit_panel.pack(fill=tk.X, padx=14, pady=(0, 8))

        def _edit_hw_item(idx):
            _show_edit_panel(idx)

        def _save_hw_edit(idx, name, qty_str, price_str):
            try:
                items = refs['hardware_data'].get("items", [])
                if 0 <= idx < len(items):
                    name = name.strip()
                    if not name:
                        messagebox.showwarning("提示", "设备名称不能为空")
                        return
                    items[idx]['name'] = name
                    items[idx]['quantity'] = float(qty_str or 1)
                    items[idx]['unit_rent'] = float(price_str or 0)
                    _update_from_hardware()
            except ValueError:
                messagebox.showwarning("提示", "数量和单价必须为数字")

        def _add_after_hw_item(idx):
            items = refs['hardware_data'].get("items", [])
            if not (0 <= idx < len(items)):
                return
            base = items[idx]
            next_type = base.get('device_type', '台式机')
            new_item = _build_hw_item(next_type, 1, refs.get('last_hw_price', 120))
            items.insert(idx + 1, new_item)
            _update_from_hardware()
            hw_card.after(30, lambda: _edit_hw_item(idx + 1))

        def _show_hw_actions(event=None):
            sel = hw_tree.selection()
            if not sel:
                return
            item_idx = int(sel[0]) if str(sel[0]).isdigit() else None
            if item_idx is None:
                return
            popup = tk.Menu(hw_tree, tearoff=0, bg=DarkTheme.BG_SECONDARY, fg=DarkTheme.TEXT_PRIMARY,
                            activebackground=DarkTheme.ACCENT_BLUE, activeforeground="white")
            popup.add_command(label="编辑", command=lambda: _edit_hw_item(item_idx))
            popup.add_command(label="新增", command=lambda: _add_after_hw_item(item_idx))
            popup.add_command(label="删除", command=lambda: _delete_hw_item(item_idx))
            try:
                popup.tk_popup(event.x_root, event.y_root)
            finally:
                popup.grab_release()

        # 刷新硬件树（添加编辑/删除按钮）
        def _summarize_item(item):
            dev_type = item.get('device_type', item.get('name', '台式机'))
            if dev_type == '台式机':
                parts = []
                cpu = item.get('cpu', '')
                ram = item.get('ram', '')
                disk = item.get('disk', '')
                gpu = item.get('gpu', '')
                parts.append(f"CPU:{cpu or '未填'}")
                parts.append(f"内存:{ram or '未填'}")
                parts.append(f"硬盘:{disk or '未填'}")
                if gpu:
                    parts.append(f"显卡:{gpu}")
            elif dev_type == '笔记本':
                parts = [f"型号:{item.get('brand_model','未填')}", f"尺寸:{item.get('screen_size','未填')}", f"CPU:{item.get('cpu','未填')}" ]
            elif dev_type == '显示器':
                parts = [f"品牌:{item.get('monitor_brand','未填')}", f"型号:{item.get('monitor_model','未填')}", f"尺寸:{item.get('screen_size','未填')}", f"分辨率:{item.get('resolution','未填')}" ]
            else:
                parts = [item.get('name', '未命名'), item.get('specs', '')]
            return " | ".join([p for p in parts if p]) if parts else item.get('name', '未命名')

        def _refresh_hw_display():
            hw = refs.get('hardware_data', {})
            items = hw.get("items", []) if isinstance(hw, dict) else []
            for row in hw_tree.get_children():
                hw_tree.delete(row)
            for idx, item in enumerate(items):
                hw_tree.insert("", tk.END, iid=str(idx), values=[
                    _summarize_item(item),
                    int(item.get('quantity', 1)),
                    f"¥{item.get('unit_rent', 0)}",
                    "已展开" if item.get("expand") else "待编辑",
                ])
        hw_tree.bind("<Double-1>", lambda e: _edit_hw_item(int(hw_tree.selection()[0])) if hw_tree.selection() else None)
        hw_tree.bind("<Button-3>", _show_hw_actions)
        _append_default_hw_if_empty()
        _refresh_hw_display()


        def _auto_calc(*_):
            try:
                m_rent = float(refs['monthly_e'].get().strip() or 0)
                months = float(refs['months_e'].get().strip() or 0)
                start_str = refs['start_e'].get().strip()
                total = m_rent * months
                refs['total_e'].delete(0, tk.END)
                refs['total_e'].insert(0, f"{total:.2f}")
                if start_str and months > 0:
                    try:
                        ds = datetime.strptime(start_str, "%Y-%m-%d")
                        de = ds + timedelta(days=int(months * 30))
                        refs['end_e'].delete(0, tk.END)
                        refs['end_e'].insert(0, de.strftime("%Y-%m-%d"))
                    except Exception:
                        pass
            except Exception:
                pass

        refs['months_e'].bind("<KeyRelease>", _auto_calc)
        refs['start_e'].bind("<KeyRelease>", _auto_calc)

        # 按钮行
        btn_frame = tk.Frame(win, bg=DarkTheme.BG_PRIMARY)
        btn_frame.pack(fill=tk.X, padx=16, pady=(0, 12))

        def _save_and_close():
            self._save_new_record_from_dialog(refs, win)

        tk.Frame(btn_frame, bg=DarkTheme.BG_PRIMARY).pack(side=tk.LEFT, expand=True)
        create_btn = tk.Button(btn_frame, text="✔️ 创建", font=DarkTheme.FONT_BUTTON,
                              fg="white", bg=DarkTheme.ACCENT_GREEN, relief=tk.FLAT, cursor="hand2",
                              command=_save_and_close, padx=14, pady=8)
        create_btn.pack(side=tk.RIGHT, padx=(0, 8))
        DarkTheme.bind_hover(create_btn, DarkTheme.ACCENT_GREEN)
        cancel_btn = tk.Button(btn_frame, text="✗ 取消", font=DarkTheme.FONT_BUTTON,
                              fg="white", bg=DarkTheme.BG_HOVER, relief=tk.FLAT, cursor="hand2",
                              command=win.destroy, padx=14, pady=8)
        cancel_btn.pack(side=tk.RIGHT)
        DarkTheme.bind_hover(cancel_btn, DarkTheme.BG_HOVER)
    
    def _save_new_record_from_dialog(self, refs, win):
        """保存新增记录。"""
        try:
            name = refs['name_e'].get().strip()
            phone = refs['phone_e'].get().strip()
            if not name or not phone:
                messagebox.showwarning("提示", "租赁人、联系电话必填")
                return

            start_str = refs['start_e'].get().strip()
            end_str = refs['end_e'].get().strip()
            monthly_widget = refs['monthly_e']
            monthly_str = monthly_widget.get().strip() if hasattr(monthly_widget, "get") else monthly_widget.cget("text")
            total_str = refs['total_e'].get().strip()
            if not start_str or not end_str or not monthly_str or not total_str:
                messagebox.showwarning("提示", "租赁信息中的待填项有空")
                return

            hardware_data = refs.get('hardware_data', {}) or {}
            items = hardware_data.get("items", []) if isinstance(hardware_data, dict) else []
            qty = 0.0
            for item in items:
                qty += float(item.get("quantity", 0) or 0)
            if qty <= 0:
                messagebox.showwarning("提示", "请至少添加一个硬件条目")
                return

            try:
                start_dt = datetime.strptime(start_str, "%Y-%m-%d")
                end_dt = datetime.strptime(end_str, "%Y-%m-%d")
            except ValueError:
                messagebox.showwarning("提示", "起租日期或到期日期格式不正确，应为 YYYY-MM-DD")
                return
            if start_dt > end_dt:
                messagebox.showwarning("提示", "起租日期不能晚于到期日期")
                return

            payload = {
                "id": self.dm.generate_unique_id(),
                "quantity": qty,
                "renter": {
                    "name": name,
                    "phone": phone,
                    "id_card": refs['id_e'].get().strip(),
                    "address": refs['addr_t'].get("1.0", "end").strip(),
                },
                "lease_info": {
                    "start_date": start_str,
                    "end_date": end_str,
                    "monthly_rent": float(monthly_str or 0),
                    "total_rent": float(total_str or 0),
                    "deposit": float(refs['deposit_e'].get().strip() or 0),
                },
                "paid_amount": float(refs['paid_e'].get().strip() or 0),
                "status": refs['status_var'].get(),
                "hardware": hardware_data,
            }

            rec = self.dm.create_rental_record(payload)
            for item in rec.get("hardware", {}).get("items", []) if isinstance(rec.get("hardware", {}), dict) else []:
                if not item.get("model_id"):
                    model = self._find_inventory_model(item)
                    if model and model.get("id"):
                        item["model_id"] = int(model.get("id"))
                        item.setdefault("brand", model.get("brand", ""))
                        item.setdefault("model_name", model.get("model_name", ""))
                        item.setdefault("reference_rent", model.get("reference_rent", 0))
                        item.setdefault("reference_cost", model.get("reference_cost", 0))
            rec["current_config_snapshot"] = json.loads(json.dumps(rec.get("hardware", {}) or {}, ensure_ascii=False))
            rec.setdefault("original_config_snapshot", json.loads(json.dumps(rec.get("hardware", {}) or {}, ensure_ascii=False)))
            self.dm.refresh_record_business_fields(rec)
            if not rec.get("renter", {}).get("name") or not rec.get("lease_info", {}).get("start_date"):
                messagebox.showerror("错误", "记录生成失败，请检查输入内容")
                return

            self._sync_inventory_for_record(rec, direction=1, trans_type="新增租赁", note="创建租赁记录占用库存")
            ok = self.dm.add_record(rec)
            if ok:
                messagebox.showinfo("成功", f"已创建记录 {rec['id']}")
                if win is not None and win.winfo_exists():
                    win.destroy()
                self._refresh()
            else:
                messagebox.showerror("错误", "保存失败，请检查记录内容后重试")
        except ValueError as e:
            messagebox.showerror("错误", f"数值错误: {e}")
        except Exception as e:
            messagebox.showerror("错误", f"保存失败: {e}")
    
    def _make_field(self, parent, label, default="", width=26):
        """创建紧凑输入行，返回 Entry 控件"""
        return self._make_field_in(parent, label, default, width=width, bg=DarkTheme.BG_PRIMARY, fill_expand=True, compact=False)

    def _make_field_in(self, parent, label, default="", width=18, bg=None, fill_expand=True, compact=False):
        """创建输入行，返回 Entry 控件。"""
        bg = bg or DarkTheme.BG_PRIMARY
        row = tk.Frame(parent, bg=bg)
        pady = 0 if compact else 2
        row.pack(fill=tk.X, pady=pady, padx=6)
        tk.Label(row, text=label, font=DarkTheme.FONT_SMALL, fg=DarkTheme.TEXT_SECONDARY,
                 bg=bg, width=5, anchor=tk.W).pack(side=tk.LEFT, padx=(0, 2))
        ent = ttk.Entry(row, width=width, font=("微软雅黑", 9))
        ent.configure(width=width)
        ent.pack(side=tk.LEFT, fill=tk.X if fill_expand else tk.NONE, expand=fill_expand, padx=(0, 2))
        if default is not None:
            ent.insert(0, str(default))
        return ent

    def _save_new_record(self, refs):
        return self._save_new_record_from_dialog(refs, None)

    def _show_edit_form(self, rec):
        """打开编辑记录弹窗。"""
        win = tk.Toplevel(self)
        win.title(f"编辑租赁记录 - {rec.get('id', '')}")
        win.geometry("1240x760")
        win.minsize(1120, 680)
        win.configure(bg=DarkTheme.BG_PRIMARY)
        self._center_on_main(win, 1240, 760)

        hdr = tk.Frame(win, bg=DarkTheme.BG_PRIMARY)
        hdr.pack(fill=tk.X, padx=18, pady=(14, 10))
        tk.Label(hdr, text=f"✏️ 编辑 {rec.get('id', '')}", font=DarkTheme.FONT_SUBTITLE,
                 fg=DarkTheme.ACCENT_YELLOW, bg=DarkTheme.BG_PRIMARY).pack(side=tk.LEFT)
        tk.Button(hdr, text="✕", font=DarkTheme.FONT_SMALL, fg=DarkTheme.TEXT_MUTED,
                  bg=DarkTheme.BG_PRIMARY, relief=tk.FLAT, cursor="hand2",
                  command=win.destroy).pack(side=tk.RIGHT)

        main_frame = tk.Frame(win, bg=DarkTheme.BG_PRIMARY)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=18, pady=(0, 14))
        left_col = tk.Frame(main_frame, bg=DarkTheme.BG_PRIMARY, width=390)
        left_col.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
        left_col.pack_propagate(False)
        right_col = tk.Frame(main_frame, bg=DarkTheme.BG_PRIMARY, width=750)
        right_col.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(10, 0))
        right_col.pack_propagate(False)

        refs = {'record': rec}
        renter = rec.get('renter', {})
        lease = rec.get('lease_info', {})
        cbg = DarkTheme.BG_CARD

        form = tk.Frame(left_col, bg=cbg, highlightbackground=DarkTheme.BORDER_COLOR, highlightthickness=1)
        form.pack(fill=tk.BOTH, expand=True)

        form_canvas = tk.Canvas(form, bg=cbg, highlightthickness=0)
        form_scroll = ttk.Scrollbar(form, orient="vertical", command=form_canvas.yview)
        form_inner = tk.Frame(form_canvas, bg=cbg)
        form_inner.bind("<Configure>", lambda e: form_canvas.configure(scrollregion=form_canvas.bbox("all")))
        form_canvas.create_window((0, 0), window=form_inner, anchor="nw", width=340)
        form_canvas.configure(yscrollcommand=form_scroll.set)
        form_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        form_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        tk.Label(form_inner, text="📋 基本信息", font=DarkTheme.FONT_SUBTITLE,
                 fg=DarkTheme.ACCENT_CYAN, bg=cbg).pack(anchor=tk.W, padx=12, pady=(8, 4))
        refs['name_e'] = self._make_field_in(form_inner, "租赁人*", renter.get("name", ""), bg=cbg, width=14, compact=True)
        refs['phone_e'] = self._make_field_in(form_inner, "联系电话*", renter.get("phone", ""), bg=cbg, width=14, compact=True)
        refs['id_e'] = self._make_field_in(form_inner, "身份证", renter.get("id_card", ""), bg=cbg, width=14, compact=True)
        refs['quantity_e'] = self._make_field_in(form_inner, "数量", str(rec.get("quantity", 1)), bg=cbg, width=14, compact=True)
        refs['start_e'] = self._make_field_in(form_inner, "起租日期", lease.get("start_date", "") or datetime.now().strftime("%Y-%m-%d"), bg=cbg, width=14, compact=True)
        refs['months_e'] = self._make_field_in(form_inner, "月数", str(lease.get("lease_months", "1")), bg=cbg, width=14, compact=True)
        refs['monthly_e'] = self._make_field_in(form_inner, "月租", lease.get("monthly_rent", "0"), bg=cbg, width=14, compact=True)
        refs['total_e'] = self._make_field_in(form_inner, "总租金", lease.get("total_rent", "0"), bg=cbg, width=14, compact=True)
        refs['end_e'] = self._make_field_in(form_inner, "到期日期", lease.get("end_date", ""), bg=cbg, width=14, compact=True)
        refs['deposit_e'] = self._make_field_in(form_inner, "押金", lease.get("deposit", "0"), bg=cbg, width=14, compact=True)
        refs['paid_e'] = self._make_field_in(form_inner, "已付金额", rec.get("paid_amount", "0"), bg=cbg, width=14, compact=True)
        refs['settle_e'] = self._make_field_in(form_inner, "结算金额", rec.get("settlement_amount", ""), bg=cbg, width=14, compact=True)

        # 让滚动区域默认在顶部，避免输入框靠下不可见
        form_canvas.yview_moveto(0)

        st_row = tk.Frame(form, bg=cbg); st_row.pack(fill=tk.X, pady=1, padx=14)
        tk.Label(st_row, text="状态", font=DarkTheme.FONT_LABEL, fg=DarkTheme.TEXT_SECONDARY, bg=cbg, width=7, anchor=tk.W).pack(side=tk.LEFT, padx=(0, 4))
        refs['status_var'] = tk.StringVar(value=rec.get("status", "在租"))
        ttk.Combobox(st_row, textvariable=refs['status_var'], state="readonly", values=["在租", "已退租", "已丢失", "已买断", "已逾期"], width=16).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 4))

        hw_card = tk.Frame(right_col, bg=cbg, highlightbackground=DarkTheme.BORDER_COLOR, highlightthickness=1)
        hw_card.pack(fill=tk.BOTH, expand=True)
        tk.Label(hw_card, text="⚙️ 租赁配置 / 硬件配置", font=DarkTheme.FONT_SUBTITLE, fg=DarkTheme.ACCENT_PURPLE, bg=cbg).pack(anchor=tk.W, padx=12, pady=(8, 6))
        refs['hardware_data'] = dict(rec.get("hardware", {}))
        refs['hardware_auto_sync'] = tk.BooleanVar(value=True)
        hw_btn_row = tk.Frame(hw_card, bg=cbg)
        hw_btn_row.pack(fill=tk.X, padx=12, pady=(0, 6))
        tk.Checkbutton(hw_btn_row, text="联动租赁字段", variable=refs['hardware_auto_sync'],
                       font=DarkTheme.FONT_SMALL, bg=cbg, fg=DarkTheme.TEXT_PRIMARY,
                       activebackground=cbg, activeforeground=DarkTheme.TEXT_PRIMARY,
                       selectcolor=DarkTheme.BG_INPUT).pack(side=tk.LEFT, padx=(0, 8))
        tk.Button(hw_btn_row, text="编辑硬件", font=DarkTheme.FONT_BUTTON, fg="white", bg=DarkTheme.ACCENT_PURPLE,
                  relief=tk.FLAT, cursor="hand2", command=lambda: self._edit_hardware_inline(refs['hardware_data'], on_updated=lambda hw: self._sync_edit_form_from_hardware(refs, hw) if refs['hardware_auto_sync'].get() else None), padx=12, pady=4).pack(side=tk.LEFT)

        btn = tk.Frame(win, bg=DarkTheme.BG_PRIMARY)
        btn.pack(fill=tk.X, padx=16, pady=(0, 12))
        tk.Button(btn, text="💾 保存", font=DarkTheme.FONT_BUTTON, fg="white", bg=DarkTheme.ACCENT_BLUE,
                  relief=tk.FLAT, cursor="hand2", command=lambda: self._save_edit_record(refs, win), padx=14, pady=8).pack(side=tk.RIGHT, padx=(8, 0))
        tk.Button(btn, text="取消", font=DarkTheme.FONT_BUTTON, fg="white", bg=DarkTheme.BG_HOVER,
                  relief=tk.FLAT, cursor="hand2", command=win.destroy, padx=14, pady=8).pack(side=tk.RIGHT)

        self._current_form_refs = refs

    # ═══ 操作记录面板 ═══


    def _show_version_detail_dialog(self, version, rec):
        """查看历史版本详情"""
        win = tk.Toplevel(self.winfo_toplevel())
        win.title(f"📋 {version.get('action', '')} - {version.get('created_at', '')}")
        win.geometry("750x550")
        win.transient(self.winfo_toplevel())
        win.grab_set()
        win.configure(bg=DarkTheme.BG_PRIMARY)

        def _safe_close(w=win):
            try:
                w.grab_release()
            except Exception:
                pass
            w.destroy()

        win.protocol("WM_DELETE_WINDOW", _safe_close)

        # 子窗口中心点对齐主窗口中心
        self._center_on_main(win, 750, 550)

        main = tk.Frame(win, bg=DarkTheme.BG_PRIMARY)
        main.pack(fill=tk.BOTH, expand=True, padx=16, pady=12)

        # 基本信息
        tk.Label(main, text=f"操作时间：{version.get('created_at', '')}",
                 font=DarkTheme.FONT_LABEL, fg=DarkTheme.TEXT_PRIMARY, bg=DarkTheme.BG_PRIMARY).pack(anchor=tk.W, pady=(0, 4))
        tk.Label(main, text=f"操作类型：{version.get('action', '')}",
                 font=DarkTheme.FONT_LABEL, fg=DarkTheme.TEXT_PRIMARY, bg=DarkTheme.BG_PRIMARY).pack(anchor=tk.W, pady=(0, 4))
        tk.Label(main, text=f"备注：{version.get('note', '')}",
                 font=DarkTheme.FONT_LABEL, fg=DarkTheme.TEXT_PRIMARY, bg=DarkTheme.BG_PRIMARY).pack(anchor=tk.W, pady=(0, 12))

        # 数据快照
        tk.Label(main, text="📊 数据快照", font=DarkTheme.FONT_SUBTITLE,
                 fg=DarkTheme.ACCENT_CYAN, bg=DarkTheme.BG_PRIMARY).pack(anchor=tk.W, pady=(0, 6))

        data = version.get("data", {})
        if data:
            # 可滚动文本
            text_frame = tk.Frame(main, bg=DarkTheme.BG_INPUT)
            text_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 8))
            text = tk.Text(text_frame, wrap=tk.WORD, font=DarkTheme.FONT_NORMAL,
                          bg=DarkTheme.BG_INPUT, fg=DarkTheme.TEXT_PRIMARY,
                          relief=tk.FLAT, padx=6, pady=4)
            scrollbar = ttk.Scrollbar(text_frame, orient="vertical", command=text.yview)
            text.configure(yscrollcommand=scrollbar.set)
            text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            # 格式化输出
            output_lines = []
            renter = data.get("renter", {})
            lease = data.get("lease_info", {})
            output_lines.append(f"记录ID: {data.get('id', 'N/A')}")
            output_lines.append(f"租赁人: {renter.get('name', 'N/A')}")
            output_lines.append(f"电话: {renter.get('phone', 'N/A')}")
            output_lines.append(f"身份证: {renter.get('id_card', 'N/A')}")
            output_lines.append(f"地址: {renter.get('address', 'N/A')}")
            output_lines.append(f"数量: {data.get('quantity', 'N/A')}")
            output_lines.append(f"起租日期: {lease.get('start_date', 'N/A')}")
            output_lines.append(f"到期日期: {lease.get('end_date', 'N/A')}")
            output_lines.append(f"月租: ¥{lease.get('monthly_rent', 0):.2f}")
            output_lines.append(f"总租金: ¥{lease.get('total_rent', 0):.2f}")
            output_lines.append(f"押金: ¥{lease.get('deposit', 0):.2f}")
            output_lines.append(f"已付金额: ¥{data.get('paid_amount', 0):.2f}")
            output_lines.append(f"状态: {data.get('status', 'N/A')}")
            hardware_summary = self.dm.summarize_hardware(data)
            output_lines.append(f"\n硬件信息:\n{hardware_summary}")

            text.insert("1.0", "\n".join(output_lines))
            text.config(state=tk.DISABLED)
        else:
            tk.Label(main, text="（无数据快照）", font=DarkTheme.FONT_LABEL,
                     fg=DarkTheme.TEXT_MUTED, bg=DarkTheme.BG_PRIMARY).pack(pady=20)

        # 按钮
        btn_frame = tk.Frame(main, bg=DarkTheme.BG_PRIMARY)
        btn_frame.pack(fill=tk.X, pady=(8, 0))

        ver_id = version.get("version_id")
        if ver_id and ver_id != rec.get("_current_version_id"):
            tk.Button(btn_frame, text="↩️ 回滚到此版本", font=DarkTheme.FONT_BUTTON, fg="white",
                      bg=DarkTheme.ACCENT_RED, relief=tk.FLAT, cursor="hand2",
                      command=lambda: self._rollback_to_version(rec, ver_id, win),
                      padx=12, pady=6).pack(side=tk.LEFT, padx=(0, 8))

        tk.Button(btn_frame, text="关闭", font=DarkTheme.FONT_BUTTON, fg="white",
                  bg=DarkTheme.BG_HOVER, relief=tk.FLAT, cursor="hand2",
                  command=_safe_close, padx=14, pady=6).pack(side=tk.LEFT)

    def _rollback_to_version(self, rec, version_id, win):
        """回滚到指定版本"""
        if not messagebox.askyesno("确认回滚", f"确定要回滚到版本 {version_id} 吗？\n此操作将覆盖当前记录。"):
            return
        rid = rec.get("id", "")
        ok = self.dm.rollback_record(rid, version_id)
        if ok:
            messagebox.showinfo("成功", f"已回滚到版本 {version_id}")
            win.destroy()
            self._refresh()
            self._show_detail_panel(rec)
        else:
            messagebox.showerror("错误", "回滚失败")


    def _sync_edit_form_from_hardware(self, refs, hardware_data):
        """根据硬件配置联动回填编辑表单中的数量与租金字段。"""
        try:
            items = hardware_data.get("items", []) if isinstance(hardware_data, dict) else []
            qty = 0.0
            rent = 0.0
            for item in items:
                q = float(item.get("quantity", 0) or 0)
                r = float(item.get("unit_rent", 0) or 0)
                qty += q
                rent += q * r
            if "quantity_e" in refs:
                refs["quantity_e"].delete(0, tk.END)
                refs["quantity_e"].insert(0, str(int(qty)) if qty == int(qty) else f"{qty:g}")
            if "monthly_e" in refs:
                refs["monthly_e"].delete(0, tk.END)
                refs["monthly_e"].insert(0, f"{rent:.2f}")
            if "total_e" in refs:
                months = 0.0
                try:
                    months = float(refs.get("months_e").get().strip() or 0) if refs.get("months_e") else 0.0
                except Exception:
                    months = 0.0
                refs["total_e"].delete(0, tk.END)
                refs["total_e"].insert(0, f"{rent * max(months, 0):.2f}")
        except Exception:
            pass

    def _save_edit_record(self, refs, win=None):
        """保存编辑记录"""
        try:
            rec = refs['record']
            name = refs['name_e'].get().strip()
            phone = refs['phone_e'].get().strip()
            if not name:
                messagebox.showwarning("提示", "租赁人不能为空")
                return
            if not phone:
                messagebox.showwarning("提示", "联系电话不能为空")
                return
            if not phone.isdigit() or len(phone) != 11:
                messagebox.showwarning("提示", "联系电话应为11位数字")
                return

            start = refs['start_e'].get().strip()
            end = refs['end_e'].get().strip()

            def parse_num(label, text):
                if text == "":
                    return 0.0
                val = float(text)
                if val < 0:
                    raise ValueError(f"{label}不能为负数")
                return val

            def check_date(label, text):
                if not text:
                    return None
                return datetime.strptime(text, "%Y-%m-%d")

            start_dt = check_date("起租日期", start)
            end_dt = check_date("到期日期", end)
            if start_dt and end_dt and start_dt > end_dt:
                messagebox.showwarning("提示", "起租日期不能晚于到期日期")
                return

            monthly = parse_num("月租", refs['monthly_e'].get().strip())
            total = parse_num("总租金", refs['total_e'].get().strip())
            deposit = parse_num("押金", refs['deposit_e'].get().strip())
            paid = parse_num("已付金额", refs['paid_e'].get().strip())
            if paid > total:
                messagebox.showwarning("提示", "已付金额不能大于总租金")
                return

            hardware_data = refs['hardware_data']
            current_snapshot = json.loads(json.dumps(hardware_data, ensure_ascii=False)) if hardware_data else {}
            updates = {
                "renter": {
                    "name": name,
                    "phone": phone,
                    "id_card": refs['id_e'].get().strip(),
                    "address": refs['addr_t'].get("1.0", tk.END).strip(),
                },
                "quantity": parse_num("数量", refs['quantity_e'].get().strip()),
                "lease_info": {
                    "start_date": start,
                    "end_date": end,
                    "monthly_rent": monthly,
                    "total_rent": total,
                    "deposit": deposit,
                },
                "hardware": hardware_data,
                "current_config_snapshot": current_snapshot,
                "paid_amount": paid,
                "status": refs['status_var'].get(),
                "config_updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            if not rec.get("original_config_snapshot"):
                updates["original_config_snapshot"] = current_snapshot

            settle_text = refs['settle_e'].get().strip()
            if settle_text:
                updates["settlement_amount"] = parse_num("结算金额", settle_text)
                updates["settlement_date"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            else:
                updates["settlement_amount"] = None
                updates["settlement_date"] = None

            old_paid = rec.get("paid_amount", 0) or 0
            if paid != old_paid and paid > old_paid:
                self.dm.append_payment_history(
                    rec, paid - old_paid, operator=self.app.username or "系统",
                    method="编辑付款", note="编辑记录时更新字段"
                )

            old_status = rec.get("status")
            old_hardware = json.loads(json.dumps(rec.get("hardware", {}) or {}, ensure_ascii=False))
            if self.dm.update_record(rec.get("id", ""), updates):
                rec.update(updates)
                self._sync_inventory_for_status_change(rec, old_status, rec.get("status"))
                self._sync_inventory_for_hardware_change(rec, old_hardware, rec.get("hardware", {}) or {})
                messagebox.showinfo("成功", "记录已更新")
                if win is not None:
                    win.destroy()
                self._refresh()
            else:
                messagebox.showerror("错误", "记录更新失败")
        except Exception as e:
            messagebox.showerror("错误", f"保存失败：{e}")

    def _show_renew_form(self, rec):
        """打开续租/买断弹窗。"""
        rid = rec.get("id", "")
        lease = rec.get("lease_info", {})
        end_str = lease.get("end_date", "")
        total = float(lease.get("total_rent", 0))
        renter = rec.get("renter", {})

        if not end_str:
            messagebox.showerror("错误", "该记录缺少到期时间")
            return
        if rec.get("status") in ("已退租", "已丢失", "已买断"):
            messagebox.showwarning("提示", f"状态为'{rec['status']}'，无法续租")
            return

        win = tk.Toplevel(self.winfo_toplevel())
        win.title(f"续租 / 买断 - {rid}")
        win.geometry("720x660")
        win.transient(self.winfo_toplevel())
        win.grab_set()
        win.configure(bg=DarkTheme.BG_PRIMARY)
        self._center_on_main(win, 720, 660)

        main = tk.Frame(win, bg=DarkTheme.BG_PRIMARY)
        main.pack(fill=tk.BOTH, expand=True, padx=12, pady=8)
        tk.Label(main, text=f"🔄 续租 / 买断 — {rid}", font=DarkTheme.FONT_SUBTITLE, fg=DarkTheme.ACCENT_BLUE, bg=DarkTheme.BG_PRIMARY).pack(anchor=tk.W)
        tk.Label(main, text=f"租赁人：{renter.get('name', '')}", font=DarkTheme.FONT_SMALL, fg=DarkTheme.TEXT_SECONDARY, bg=DarkTheme.BG_PRIMARY).pack(anchor=tk.W, pady=(1, 0))
        tk.Label(main, text=f"当前到期：{end_str}", font=DarkTheme.FONT_SMALL, fg=DarkTheme.TEXT_SECONDARY, bg=DarkTheme.BG_PRIMARY).pack(anchor=tk.W, pady=(1, 0))
        refs = {'record': rec, 'lease': lease, 'cur_total': total, 'unit_var': tk.StringVar(value='月'), 'action_var': tk.StringVar(value='renew')}
        refs['renew_end_var'] = tk.StringVar(value=end_str)
        refs['settle_mode_var'] = tk.StringVar(value='refund')
        tk.Label(main, textvariable=refs['renew_end_var'], font=DarkTheme.FONT_SMALL, fg=DarkTheme.ACCENT_BLUE, bg=DarkTheme.BG_PRIMARY).pack(anchor=tk.W, pady=(1, 0))
        tk.Label(main, text=f"当前租金总额：¥{total:.2f}", font=DarkTheme.FONT_SMALL, fg=DarkTheme.TEXT_SECONDARY, bg=DarkTheme.BG_PRIMARY).pack(anchor=tk.W, pady=(1, 3))

        form = tk.Frame(main, bg=DarkTheme.BG_PRIMARY)
        form.pack(fill=tk.BOTH, expand=True)

        mode_row = tk.Frame(form, bg=DarkTheme.BG_PRIMARY)
        mode_row.pack(fill=tk.X, pady=(0, 4))
        tk.Label(mode_row, text='处理方式：', font=DarkTheme.FONT_SMALL, fg=DarkTheme.TEXT_SECONDARY, bg=DarkTheme.BG_PRIMARY, width=9, anchor=tk.W).pack(side=tk.LEFT)
        mode_box = tk.Frame(mode_row, bg=DarkTheme.BG_PRIMARY)
        mode_box.pack(side=tk.LEFT)
        for val, txt in (('renew', '续租'), ('buyout', '买断'), ('return', '退租'), ('lost', '丢失')):
            tk.Radiobutton(mode_box, text=txt, variable=refs['action_var'], value=val, font=DarkTheme.FONT_SMALL,
                           bg=DarkTheme.BG_PRIMARY, fg=DarkTheme.TEXT_PRIMARY, selectcolor=DarkTheme.BG_PRIMARY,
                           activebackground=DarkTheme.BG_PRIMARY, indicatoron=1, padx=0, pady=0).pack(side=tk.LEFT, padx=(0, 8))

        row = tk.Frame(form, bg=DarkTheme.BG_PRIMARY)
        row.pack(fill=tk.X, pady=(0, 1))
        tk.Label(row, text='时间单位：', font=DarkTheme.FONT_SMALL, fg=DarkTheme.TEXT_SECONDARY, bg=DarkTheme.BG_PRIMARY, width=9, anchor=tk.W).pack(side=tk.LEFT)
        unit_box = tk.Frame(row, bg=DarkTheme.BG_PRIMARY)
        unit_box.pack(side=tk.LEFT)
        for val in ('天', '月'):
            tk.Radiobutton(unit_box, text=val, variable=refs['unit_var'], value=val, font=DarkTheme.FONT_SMALL,
                           bg=DarkTheme.BG_PRIMARY, fg=DarkTheme.TEXT_PRIMARY, selectcolor=DarkTheme.BG_PRIMARY,
                           activebackground=DarkTheme.BG_PRIMARY, indicatoron=1, padx=0, pady=0).pack(side=tk.LEFT, padx=(0, 4))

        def _field(parent, label, value="", readonly=False):
            r = tk.Frame(parent, bg=DarkTheme.BG_PRIMARY)
            r.pack(fill=tk.X, pady=1)
            tk.Label(r, text=label, font=DarkTheme.FONT_SMALL, fg=DarkTheme.TEXT_SECONDARY, bg=DarkTheme.BG_PRIMARY, width=11, anchor=tk.W).pack(side=tk.LEFT)
            e = ttk.Entry(r, font=("微软雅黑", 9))
            e.pack(side=tk.LEFT, fill=tk.X, expand=True)
            e.insert(0, str(value))
            if readonly:
                e.configure(state="readonly")
            return e

        refs['time_e'] = _field(form, "续租时间", "")
        refs['unit_price_e'] = _field(form, "单价(¥)", "")
        refs['amt_e'] = _field(form, "费用(¥)", "", readonly=True)
        refs['paid_e'] = _field(form, "已付金额(¥)", str(rec.get("paid_amount", 0)))
        refs['buyout_e'] = _field(form, "买断金额(¥)", f"{max(0.0, total - float(rec.get('paid_amount', 0) or 0)):.2f}")
        refs['settle_e'] = _field(form, "退租/丢失金额(¥)", "0")

        hint = tk.Label(form, text="续租与买断会更新到期时间；退租/丢失会直接结算并标记订单状态。退租金额可填写负数表示退还租金，正数表示补差。",
                        font=DarkTheme.FONT_SMALL, fg=DarkTheme.TEXT_MUTED, bg=DarkTheme.BG_PRIMARY, wraplength=660, justify=tk.LEFT)
        hint.pack(anchor=tk.W, pady=(8, 0))

        btn = tk.Frame(form, bg=DarkTheme.BG_PRIMARY)
        btn.pack(fill=tk.X, pady=(8, 0))
        tk.Button(btn, text="取消", font=DarkTheme.FONT_BUTTON, fg="white", bg=DarkTheme.BG_HOVER,
                  relief=tk.FLAT, cursor="hand2", command=win.destroy, padx=14, pady=6).pack(side=tk.RIGHT)
        tk.Button(btn, text="确认", font=DarkTheme.FONT_BUTTON, fg="white", bg=DarkTheme.ACCENT_BLUE,
                  relief=tk.FLAT, cursor="hand2", command=lambda: self._do_renew_inline(refs, win), padx=14, pady=6).pack(side=tk.RIGHT, padx=(0, 6))

        def _calc_renew_amount(*_):
            try:
                action = refs['action_var'].get()
                if action == 'buyout':
                    buyout_amount = float(refs['buyout_e'].get().strip() or 0)
                    refs['amt_e'].configure(state="normal")
                    refs['amt_e'].delete(0, tk.END)
                    refs['amt_e'].insert(0, f"{buyout_amount:.2f}")
                    refs['amt_e'].configure(state="readonly")
                    refs['renew_end_var'].set(f"买断后状态：已买断")
                    return
                if action in ('return', 'lost'):
                    settle_val = float(refs['settle_e'].get().strip() or 0)
                    refs['amt_e'].configure(state="normal")
                    refs['amt_e'].delete(0, tk.END)
                    refs['amt_e'].insert(0, f"{settle_val:.2f}")
                    refs['amt_e'].configure(state="readonly")
                    refs['renew_end_var'].set("结算后将更新为对应状态")
                    return

                t = float(refs['time_e'].get().strip() or 0)
                p = float(refs['unit_price_e'].get().strip() or 0)
                refs['amt_e'].configure(state="normal")
                refs['amt_e'].delete(0, tk.END)
                if t > 0 and p >= 0:
                    refs['amt_e'].insert(0, f"{t * p:.2f}")
                refs['amt_e'].configure(state="readonly")

                cur_end = datetime.strptime(end_str, "%Y-%m-%d")
                if refs['unit_var'].get() == '天':
                    new_end = cur_end + timedelta(days=int(t))
                else:
                    new_end = cur_end + timedelta(days=int(t * 30))
                refs['renew_end_var'].set(f"续租后到期：{new_end.strftime('%Y-%m-%d')}")
            except Exception:
                refs['renew_end_var'].set(f"续租后到期：{end_str}")

        def _sync_renew_unit(*_):
            unit = refs['unit_var'].get()
            base_month_price = cur_total / max(float(lease.get('lease_months', 1) or 1), 1)
            default_price = base_month_price if unit == '月' else round(base_month_price / 30.0, 2)
            refs['unit_price_e'].delete(0, tk.END)
            refs['unit_price_e'].insert(0, f"{default_price:.2f}")
            _calc_renew_amount()

        def _sync_mode(*_):
            action = refs['action_var'].get()
            is_settlement = action in ('buyout', 'return', 'lost')
            refs['time_e'].configure(state="disabled" if is_settlement else "normal")
            refs['unit_price_e'].configure(state="disabled" if is_settlement else "normal")
            refs['buyout_e'].configure(state="normal" if action == 'buyout' else "disabled")
            refs['paid_e'].configure(state="normal" if action == 'buyout' else "disabled")
            refs['settle_e'].configure(state="normal" if action in ('return', 'lost') else "disabled")
            if action == 'buyout':
                refs['renew_end_var'].set("买断后状态：已买断")
            elif action == 'return':
                refs['renew_end_var'].set("退租后状态：已退租")
            elif action == 'lost':
                refs['renew_end_var'].set("丢失后状态：已丢失")
            else:
                _sync_renew_unit()
            _calc_renew_amount()

        refs['time_e'].bind('<KeyRelease>', _calc_renew_amount)
        refs['unit_price_e'].bind('<KeyRelease>', _calc_renew_amount)
        refs['buyout_e'].bind('<KeyRelease>', _calc_renew_amount)
        refs['settle_e'].bind('<KeyRelease>', _calc_renew_amount)
        refs['unit_var'].trace_add('write', _sync_renew_unit)
        refs['action_var'].trace_add('write', _sync_mode)
        refs['unit_var'].set('月')
        _sync_renew_unit()
        _sync_mode()


    def _do_renew_inline(self, refs, win=None):
        """执行续租或买断操作。"""
        try:
            rec = refs['record']
            lease = refs['lease']
            cur_total = refs['cur_total']
            unit_var = refs['unit_var']
            action = refs.get('action_var').get() if refs.get('action_var') else 'renew'

            action = refs.get('action_var').get() if refs.get('action_var') else 'renew'
            if action == 'buyout':
                buyout_amount_str = refs['buyout_e'].get().strip()
                if not buyout_amount_str:
                    messagebox.showwarning("提示", "请输入买断金额")
                    return
                buyout_amount = float(buyout_amount_str)
                if buyout_amount < 0:
                    messagebox.showwarning("提示", "买断金额不能为负")
                    return

                paid_str = refs['paid_e'].get().strip()
                new_paid = float(paid_str) if paid_str else float(rec.get("paid_amount", 0) or 0)
                if new_paid < 0:
                    messagebox.showwarning("提示", "已付金额不能为负")
                    return

                old_status = rec.get("status")
                new_total = float(lease.get("total_rent", 0) or 0)
                new_end_str = datetime.now().strftime("%Y-%m-%d")
                updates = {
                    "lease_info": {
                        **lease,
                        "end_date": new_end_str,
                        "total_rent": new_total,
                    },
                    "paid_amount": new_paid,
                    "status": "已买断",
                    "settlement_amount": buyout_amount,
                    "settlement_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "renew_history": list(rec.get("renew_history", [])) + [{
                        "renew_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "renew_time": 0,
                        "renew_unit": "次",
                        "renew_amount": buyout_amount,
                        "old_end_date": lease.get("end_date", ""),
                        "new_end_date": new_end_str,
                        "operator": self.app.username or "系统",
                        "mode": "buyout",
                    }],
                }
                if new_paid != rec.get("paid_amount", 0):
                    paid_change = new_paid - (rec.get("paid_amount", 0) or 0)
                    if paid_change > 0:
                        self.dm.append_payment_history(
                            rec, paid_change, operator=self.app.username or "系统",
                            method="买断收款", note="买断结算"
                        )
                if self.dm.update_record(rec.get("id", ""), updates):
                    self._sync_inventory_for_status_change(rec, old_status, "已买断")
                    messagebox.showinfo("成功", f"买断成功！\n买断金额：¥{buyout_amount:.2f}\n记录已标记为已买断")
                    if win is not None:
                        win.destroy()
                    self._refresh()
                else:
                    messagebox.showerror("错误", "买断失败")
                return

            if action in ('return', 'lost'):
                settle_str = refs['settle_e'].get().strip() or "0"
                settle_amount = float(settle_str)
                if action == 'return' and settle_amount < 0:
                    messagebox.showwarning("提示", "退租金额不能小于 0")
                    return
                if action == 'lost' and settle_amount < 0:
                    messagebox.showwarning("提示", "亏损金额不能小于 0")
                    return

                new_status = '已退租' if action == 'return' else '已丢失'
                old_status = rec.get('status')
                updates = {
                    'status': new_status,
                    'settlement_amount': settle_amount,
                    'settlement_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'renew_history': list(rec.get('renew_history', [])) + [{
                        'renew_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'renew_time': 0,
                        'renew_unit': '次',
                        'renew_amount': settle_amount,
                        'old_end_date': lease.get('end_date', ''),
                        'new_end_date': datetime.now().strftime('%Y-%m-%d'),
                        'operator': self.app.username or '系统',
                        'mode': action,
                    }],
                }
                if self.dm.update_record(rec.get('id', ''), updates):
                    rec.update(updates)
                    self._sync_inventory_for_status_change(rec, old_status, new_status)
                    messagebox.showinfo('成功', f"{'退租' if action == 'return' else '丢失'}处理完成，订单已标记为{new_status}")
                    if win is not None:
                        win.destroy()
                    self._refresh()
                else:
                    messagebox.showerror('错误', '处理失败')
                return

            t_str = refs['time_e'].get().strip()
            if not t_str:
                messagebox.showwarning("提示", "请输入续租时间")
                return
            t_val = float(t_str)
            if t_val <= 0:
                messagebox.showwarning("提示", "时间必须 > 0")
                return

            a_str = refs['amt_e'].get().strip()
            if not a_str:
                messagebox.showwarning("提示", "请输入续租总价")
                return
            amt = float(a_str)
            if amt < 0:
                messagebox.showwarning("提示", "金额不能为负")
                return

            p_str = refs['paid_e'].get().strip()
            new_paid = float(p_str) if p_str else rec.get("paid_amount", 0)
            if new_paid < 0:
                messagebox.showwarning("提示", "已付金额不能为负")
                return
            new_total = cur_total + amt
            if new_paid > new_total:
                messagebox.showwarning("提示", "已付金额不能大于总租金")
                return

            cur_end = datetime.strptime(lease["end_date"], "%Y-%m-%d")
            if unit_var.get() == "天":
                new_end = cur_end + timedelta(days=int(t_val))
                add_months = t_val / 30.0
            else:
                new_end = cur_end + timedelta(days=int(t_val * 30))
                add_months = t_val
            new_end_str = new_end.strftime("%Y-%m-%d")

            old_start = lease.get("start_date", "")
            if old_start:
                old_start_dt = datetime.strptime(old_start, "%Y-%m-%d")
                total_months = (new_end - old_start_dt).days / 30.0
            else:
                total_months = float(lease.get("lease_months", 0) or 0) + add_months

            updates = {
                "lease_info": {
                    **lease,
                    "end_date": new_end_str,
                    "total_rent": new_total,
                    "lease_months": round(total_months, 2),
                },
                "paid_amount": new_paid,
                "renew_history": list(rec.get("renew_history", [])) + [{
                    "renew_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "renew_time": t_val, "renew_unit": unit_var.get(), "renew_amount": amt,
                    "old_end_date": cur_end.strftime("%Y-%m-%d"), "new_end_date": new_end_str,
                    "operator": self.app.username or "系统",
                    "mode": "renew",
                }],
            }
            if new_paid != rec.get("paid_amount", 0):
                paid_change = new_paid - (rec.get("paid_amount", 0) or 0)
                if paid_change > 0:
                    self.dm.append_payment_history(
                        rec, paid_change, operator=self.app.username or "系统",
                        method="续租时支付", note=f"续租{t_val}{unit_var.get()}"
                    )
            if self.dm.update_record(rec.get("id", ""), updates):
                messagebox.showinfo("成功", f"续租成功！\n时间：{t_val}{unit_var.get()}\n金额：¥{amt:.2f}\n新到期：{new_end_str}")
                if win is not None:
                    win.destroy()
                self._refresh()
            else:
                messagebox.showerror("错误", "续租失败")
        except Exception as e:
            messagebox.showerror("错误", f"续租失败：{e}")

    def _update_change_preview(self, refs):
        preview = refs.get('change_preview')
        if not preview:
            return
        try:
            selected = refs.get('selected_model')
            old_name = refs.get('old_name').get().strip() if refs.get('old_name') else ""
            new_name = refs.get('new_name').get().strip() if refs.get('new_name') else ""
            change_type = refs.get('change_type').get().strip() if refs.get('change_type') else ""
            fee = refs.get('fee').get().strip() if refs.get('fee') else ""
            mode = refs.get('change_mode').get().strip() if refs.get('change_mode') else "换件"
            model_text = "未选择型号"
            if selected:
                model_text = f"{selected.get('brand', '')} {selected.get('model_name', '')}｜参考租金 ¥{float(selected.get('reference_rent') or 0):.2f}｜参考成本 ¥{float(selected.get('reference_cost') or 0):.2f}"
            preview.configure(text=f"模式：{mode}\n旧件：{old_name or '未填写'}\n新件：{new_name or '未填写'}\n类型：{change_type or '未填写'}\n费用：{fee or '0'}\n型号：{model_text}")
        except Exception as e:
            preview.configure(text=f"预览更新失败：{e}")

    def _sync_change_mode(self, refs):
        mode = refs.get('change_mode').get().strip() if refs.get('change_mode') else "换件"
        if mode == "重写":
            if refs.get('change_type'):
                refs['change_type'].delete(0, tk.END)
                refs['change_type'].insert(0, "重写配置")
            if refs.get('old_name') and not refs['old_name'].get().strip():
                refs['old_name'].insert(0, self._safe_guess_current_part_name(refs.get('current_snapshot')))
            if refs.get('reason') and not refs['reason'].get().strip():
                refs['reason'].insert(0, "重新整理/重写设备配置")
        else:
            if refs.get('change_type') and refs['change_type'].get().strip() == "重写配置":
                refs['change_type'].delete(0, tk.END)
                refs['change_type'].insert(0, "更换硬盘")
        self._update_change_preview(refs)

    def _save_change_record(self, refs, win):
        try:
            if not refs.get('change_time') or not refs.get('change_type') or not refs.get('operator'):
                messagebox.showerror("错误", "换件登记表单未初始化完整，请重新打开页面")
                return
            change_time = refs['change_time'].get().strip() or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            change_mode = refs.get('change_mode').get().strip() if refs.get('change_mode') else "换件"
            change_type = refs['change_type'].get().strip() or ("重写配置" if change_mode == "重写" else "硬件变更")
            old_name = refs['old_name'].get().strip()
            new_name = refs['new_name'].get().strip()
            reason = refs['reason'].get().strip() or ("重新整理/重写设备配置" if change_mode == "重写" else "")
            operator = refs['operator'].get().strip() or getattr(self.app, "username", "系统")
            fee = float(refs['fee'].get().strip() or 0)
            remark = refs['remark'].get().strip() or reason

            if not new_name and not refs.get('selected_model'):
                messagebox.showwarning("提示", "请填写新件名称或先选择一个型号")
                return

            old_hw = json.loads(json.dumps(self.rec.get("current_config_snapshot") or self.rec.get("hardware") or {}, ensure_ascii=False))
            selected_model = refs.get('selected_model')
            new_hw = self._apply_change_to_snapshot(old_hw, change_type, old_name, new_name)

            if selected_model:
                display_name = f"{selected_model.get('brand', '')} {selected_model.get('model_name', '')}".strip()
                new_name = new_name or display_name
                if not fee:
                    fee = float(selected_model.get('reference_rent') or selected_model.get('reference_cost') or 0)
                self.dm.update_model(selected_model.get('id'), is_active=True)

            self.dm.append_config_change(
                self.rec.get("id", ""),
                change_type=change_type,
                old_part_name=old_name or self._guess_current_part_name(old_hw),
                new_part_name=new_name or "新件",
                operator=operator,
                reason=reason,
                fee=fee,
                remark=remark,
            )
            if selected_model and selected_model.get('id'):
                self.dm.add_inventory_transaction(
                    int(selected_model['id']),
                    trans_type="出库",
                    quantity=1,
                    related_rental_id=self.rec.get("id", ""),
                    operator=operator,
                    note=remark or reason or change_type,
                )
                self.dm.update_model(int(selected_model['id']), is_active=True)
            self.dm.append_hardware_history(
                self.rec,
                old_hw,
                new_hw,
                operator=operator,
                note=remark or reason or change_type,
            )
            self.dm.update_record(self.rec.get("id", ""), {
                "hardware": new_hw,
                "current_config_snapshot": new_hw,
                "original_config_snapshot": self.rec.get("original_config_snapshot") or old_hw,
                "config_updated_at": change_time,
            })
            self.rec = self.dm.get_record_by_id(self.rec.get("id", "")) or self.rec
            self.dm.refresh_record_business_fields(self.rec)
            self.dm.save()
            messagebox.showinfo("成功", "换件登记已保存")
            win.destroy()
            self._refresh()
        except Exception as e:
            messagebox.showerror("错误", f"保存失败：{e}")

    def _guess_current_part_name(self, snapshot):
        if not isinstance(snapshot, dict):
            return ""
        items = snapshot.get("items")
        if isinstance(items, list) and items:
            first = items[0] or {}
            return first.get("name") or first.get("device_type") or first.get("model") or first.get("cpu") or first.get("disk") or first.get("ram") or ""
        for key in ("disk", "ram", "cpu", "gpu", "monitor", "brand_model", "model", "name"):
            if snapshot.get(key):
                return str(snapshot.get(key))
        return ""

    def _safe_guess_current_part_name(self, snapshot):
        try:
            return self._guess_current_part_name(snapshot)
        except Exception:
            return ""

    def _search_model_picker(self, refs):
        try:
            query = refs.get('model_query')
            if not query:
                return
            q = query.get().strip()
            if not q:
                messagebox.showwarning("提示", "请输入要搜索的型号关键字")
                return
            results = []
            for cat in ("cpu", "mb", "ram", "disk", "gpu", "psu", "case", "cooler", "monitor"):
                try:
                    results.extend(self.dm.search_models(q, category=cat, limit=8))
                except Exception:
                    continue
            if not results:
                messagebox.showinfo("提示", "未找到匹配型号")
                return
            self._show_model_pick_dialog(results, refs)
        except Exception as e:
            messagebox.showerror("错误", f"搜索型号失败：{e}")

    def _show_model_pick_dialog(self, models, refs):
        if not models:
            messagebox.showinfo("提示", "没有可选型号")
            return
        owner = self.win.winfo_toplevel() if hasattr(self, "win") and self.win else self.winfo_toplevel()
        win = tk.Toplevel(owner)
        win.title("选择型号")
        win.geometry("940x500")
        win.transient(owner)
        win.grab_set()
        win.configure(bg=DarkTheme.BG_PRIMARY)
        self._center_on_main(win, 940, 500)

        main = tk.Frame(win, bg=DarkTheme.BG_PRIMARY)
        main.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)
        tk.Label(main, text="选择与当前换件匹配的型号", font=DarkTheme.FONT_SUBTITLE, fg=DarkTheme.ACCENT_CYAN, bg=DarkTheme.BG_PRIMARY).pack(anchor=tk.W, pady=(0, 8))
        cols = ("分类", "品牌", "型号", "参考租金", "参考成本")
        tree = ttk.Treeview(main, columns=cols, show="headings", height=14)
        for c, w in [("分类", 120), ("品牌", 170), ("型号", 300), ("参考租金", 120), ("参考成本", 120)]:
            tree.heading(c, text=c)
            tree.column(c, width=w, anchor="center")
        tree.pack(fill=tk.BOTH, expand=True)
        for m in models:
            tree.insert("", tk.END, iid=str(m.get('id')), values=(m.get('category', ''), m.get('brand', ''), m.get('model_name', ''), f"¥{float(m.get('reference_rent') or 0):.2f}", f"¥{float(m.get('reference_cost') or 0):.2f}"))

        hint = tk.Label(main, text="双击即可选中，选中后会自动回填新件名称、默认费用和换件类型。", font=DarkTheme.FONT_SMALL, fg=DarkTheme.TEXT_MUTED, bg=DarkTheme.BG_PRIMARY)
        hint.pack(anchor=tk.W, pady=(8, 0))

        def _choose():
            sel = tree.selection()
            if not sel:
                return
            mid = int(sel[0])
            model = next((m for m in models if int(m.get('id')) == mid), None)
            if not model:
                return
            display = f"{model.get('brand', '')} {model.get('model_name', '')}".strip()
            refs['selected_model'] = model
            refs['new_name'].delete(0, tk.END)
            refs['new_name'].insert(0, display)
            fee = model.get('reference_rent') or model.get('reference_cost') or 0
            refs['fee'].delete(0, tk.END)
            refs['fee'].insert(0, f"{float(fee):.2f}" if fee else "0")
            refs['change_type'].delete(0, tk.END)
            refs['change_type'].insert(0, self._change_type_from_category(model.get('category', '')))
            self._update_change_preview(refs)
            win.destroy()

        btn = tk.Frame(main, bg=DarkTheme.BG_PRIMARY)
        btn.pack(fill=tk.X, pady=(8, 0))
        tk.Button(btn, text="选中", font=DarkTheme.FONT_BUTTON, fg="white", bg=DarkTheme.ACCENT_GREEN, relief=tk.FLAT, cursor="hand2", command=_choose, padx=12, pady=6).pack(side=tk.LEFT)
        tk.Button(btn, text="关闭", font=DarkTheme.FONT_BUTTON, fg="white", bg=DarkTheme.BG_HOVER, relief=tk.FLAT, cursor="hand2", command=win.destroy, padx=12, pady=6).pack(side=tk.LEFT, padx=(8, 0))
        tree.bind("<Double-1>", lambda e: _choose())

    def _sync_model_stock(self, refs):
        try:
            model = refs.get('selected_model')
            if not model:
                messagebox.showwarning("提示", "请先搜索并选择一个型号")
                return
            mid = model.get('id')
            current = self.dm.get_model_by_id(mid)
            if not current:
                messagebox.showwarning("提示", "未找到该型号")
                return
            stock_total = int(current.get('stock_total') or 0)
            stock_rented = int(current.get('stock_rented') or 0)
            self.dm.update_model(mid, stock_total=stock_total + 1, stock_rented=max(0, stock_rented - 1))
            self.dm.add_inventory_transaction(
                model_id=mid,
                trans_type="同步",
                quantity=1,
                operator=getattr(self.app, "username", "系统"),
                note="租赁记录内同步库存",
            )
            messagebox.showinfo("成功", f"库存已同步：{current.get('brand', '')} {current.get('model_name', '')}")
        except Exception as e:
            messagebox.showerror("错误", f"同步库存失败：{e}")

    def _change_type_from_category(self, category):
        mapping = {"cpu": "更换CPU", "mb": "更换主板", "ram": "更换内存", "disk": "更换硬盘", "gpu": "更换显卡", "psu": "更换电源", "case": "更换机箱", "cooler": "更换散热", "monitor": "更换显示器"}
        return mapping.get(category, "更换配件")

    def _apply_change_to_snapshot(self, snapshot, change_type, old_name, new_name):
        snap = json.loads(json.dumps(snapshot or {}, ensure_ascii=False))
        target = new_name or old_name
        if isinstance(snap, dict) and isinstance(snap.get("items"), list) and snap["items"]:
            first = snap["items"][0]
            lower_type = (change_type or "").lower()
            if any(k in lower_type for k in ["硬盘", "disk"]):
                first["disk"] = target
            elif any(k in lower_type for k in ["内存", "ram"]):
                first["ram"] = target
            elif any(k in lower_type for k in ["显卡", "gpu"]):
                first["gpu"] = target
            elif any(k in lower_type for k in ["显示器", "monitor"]):
                first["monitor_model"] = target
            elif any(k in lower_type for k in ["主板", "motherboard", "mb"]):
                first["motherboard"] = target
            elif any(k in lower_type for k in ["cpu"]):
                first["cpu"] = target
            else:
                first["name"] = target or first.get("name", "")
                first["device_type"] = first.get("device_type", "设备")
            return snap
        if isinstance(snap, dict):
            key = "disk" if "硬盘" in (change_type or "") else "ram" if "内存" in (change_type or "") else "cpu"
            snap[key] = target or snap.get(key, "")
        return snap

    def _show_payment_form(self, rec):
        """收款记录弹窗"""
        rid = rec.get("id", "")
        renter = rec.get("renter", {})
        name = renter.get("name", "")
        paid = float(rec.get("paid_amount", 0) or 0)
        total = float(rec.get("lease_info", {}).get("total_rent", 0) or 0)
        unpaid = total - paid

        owner = self.winfo_toplevel()
        win = tk.Toplevel(owner)
        win.title(f"💰 收款 - {rid}")
        win.geometry("460x390")
        win.transient(owner)
        win.grab_set()
        win.configure(bg=DarkTheme.BG_PRIMARY)
        self._center_on_main(win, 460, 390)

        def _safe_close(w=win):
            try:
                w.grab_release()
            except Exception:
                pass
            w.destroy()

        win.protocol("WM_DELETE_WINDOW", _safe_close)

        main = tk.Frame(win, bg=DarkTheme.BG_PRIMARY)
        main.pack(fill=tk.BOTH, expand=True, padx=16, pady=14)

        tk.Label(main, text="💰 收款登记", font=DarkTheme.FONT_SUBTITLE, fg=DarkTheme.ACCENT_GREEN, bg=DarkTheme.BG_PRIMARY).pack(anchor=tk.W, pady=(0, 8))
        tk.Label(main, text=f"租赁人：{name}", font=DarkTheme.FONT_LABEL, fg=DarkTheme.TEXT_PRIMARY, bg=DarkTheme.BG_PRIMARY).pack(anchor=tk.W, pady=(0, 4))
        tk.Label(main, text=f"总租金：¥{total:.2f}", font=DarkTheme.FONT_LABEL, fg=DarkTheme.ACCENT_BLUE, bg=DarkTheme.BG_PRIMARY).pack(anchor=tk.W, pady=(0, 4))
        tk.Label(main, text=f"已付：¥{paid:.2f}", font=DarkTheme.FONT_LABEL, fg=DarkTheme.ACCENT_GREEN, bg=DarkTheme.BG_PRIMARY).pack(anchor=tk.W, pady=(0, 4))
        tk.Label(main, text=f"未付：¥{unpaid:.2f}", font=DarkTheme.FONT_LABEL, fg=DarkTheme.ACCENT_RED if unpaid > 0 else DarkTheme.ACCENT_GREEN, bg=DarkTheme.BG_PRIMARY).pack(anchor=tk.W, pady=(0, 12))

        refs = {}
        refs['amount_e'] = self._make_field(main, "收款金额", str(unpaid) if unpaid > 0 else "")
        refs['method_e'] = self._make_field(main, "收款方式", "现金")
        refs['note_e'] = self._make_field(main, "备注", "")

        btn = tk.Frame(main, bg=DarkTheme.BG_PRIMARY)
        btn.pack(fill=tk.X, pady=(12, 0))
        tk.Button(btn, text="✅ 确认收款", font=DarkTheme.FONT_BUTTON, fg="white", bg=DarkTheme.ACCENT_GREEN, relief=tk.FLAT, cursor="hand2", command=lambda: self._process_payment(rec, refs, win), padx=14, pady=8).pack(side=tk.LEFT, padx=(0, 8))
        tk.Button(btn, text="取消", font=DarkTheme.FONT_BUTTON, fg="white", bg=DarkTheme.BG_HOVER, relief=tk.FLAT, cursor="hand2", command=_safe_close, padx=14, pady=8).pack(side=tk.LEFT)

    def _process_payment(self, rec, refs, win):
        """处理收款记录"""
        try:
            amount_str = refs['amount_e'].get().strip()
            if not amount_str:
                messagebox.showwarning("提示", "请输入收款金额")
                return
            amount = float(amount_str)
            if amount <= 0:
                messagebox.showwarning("提示", "金额必须大于 0")
                return

            method = refs['method_e'].get().strip() or "现金"
            note = refs['note_e'].get().strip()

            old_paid = float(rec.get("paid_amount", 0) or 0)
            new_paid = old_paid + amount

            new_payment_history = list(rec.get("payment_history", []))
            new_payment_history.append({
                "payment_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "amount": amount,
                "method": method,
                "note": note,
                "operator": self.app.username or "系统",
            })

            updates = {
                "paid_amount": new_paid,
                "payment_history": new_payment_history,
            }
            old_status = rec.get("status")
            if self.dm.update_record(rec.get("id", ""), updates):
                rec["paid_amount"] = new_paid
                rec["payment_history"] = new_payment_history
                rec["status"] = rec.get("status", old_status)
                self.dm.refresh_record_business_fields(rec)
                self.dm.save()
                new_unpaid = float(rec.get("lease_info", {}).get("total_rent", 0) or 0) - rec["paid_amount"]
                messagebox.showinfo("成功", f"收款 ¥{amount:.2f} 已记录\n剩余未付: ¥{max(0, new_unpaid):.2f}")
                win.destroy()
                self._refresh()
            else:
                messagebox.showerror("错误", "收款失败")
        except ValueError:
            messagebox.showerror("错误", "金额格式不正确")
        except Exception as e:
            messagebox.showerror("错误", f"收款失败: {e}")

    def _navigate_to_record(self, rid):
        """从AI助手导航到指定记录"""
        rec = self._find_record(rid)
        if rec:
            # 清除搜索过滤以显示完整列表
            self.search_var.set("")
            self.status_var.set("全部")
            self._shown = list(self._all)
            target_idx = next((i for i, r in enumerate(self._shown) if r.get("id") == rid), -1)
            if target_idx >= 0:
                self._current_page = target_idx // self._page_size
            self._render_tree()
            # 选中对应记录并显示详情
            for item in self.tree.get_children():
                if self.tree.item(item)["values"][0] == rid:
                    self.tree.selection_set(item)
                    self.tree.focus(item)
                    self.tree.see(item)
                    break
            self._show_detail_panel(rec)
        else:
            messagebox.showinfo("提示", f"未找到记录 {rid}")

    def _export_contract(self, rec):
        """导出租赁合同为文本文件"""
        from tkinter import filedialog
        rid = rec.get("id", "")
        renter = rec.get("renter", {})
        lease = rec.get("lease_info", {})
        name = renter.get("name", "")
        phone = renter.get("phone", "")
        id_card = renter.get("id_card", "")
        address = renter.get("address", "")
        start_date = lease.get("start_date", "")
        end_date = lease.get("end_date", "")
        monthly_rent = float(lease.get("monthly_rent", 0) or 0)
        total_rent = float(lease.get("total_rent", 0) or 0)
        deposit = float(lease.get("deposit", 0) or 0)
        quantity = int(rec.get("quantity", 1) or 1)
        status = rec.get("status", "")
        paid_amount = float(rec.get("paid_amount", 0) or 0)
        hardware = rec.get("hardware", {})
        hardware_summary = self.dm.summarize_hardware(rec)

        contract_text = f"""========================================
        速维电脑租赁合同
========================================

合同编号: {rid}
签订日期: {start_date}

【出租方】
名称: 速维电脑租赁
联系电话: 13800138000
地址: XX市XX区XX路XX号

【承租方】
姓名: {name}
联系电话: {phone}
身份证号: {id_card}
联系地址: {address}

【租赁物品清单】
数量: {quantity} 套
{hardware_summary}

【租赁期限】
起租日期: {start_date}
到期日期: {end_date}

【租金及押金】
月租金: ¥{monthly_rent:.2f}
总租金: ¥{total_rent:.2f}
押金: ¥{deposit:.2f}
已付金额: ¥{paid_amount:.2f}
未付金额: ¥{max(0, total_rent - paid_amount):.2f}

【合同状态】
当前状态: {status}

【备注】
{rec.get("notes", "无")}

========================================
出租方签字: ______________    日期: ____________
承租方签字: ______________    日期: ____________
========================================
"""

        fp = filedialog.asksaveasfilename(
            title="导出租赁合同",
            defaultextension=".txt",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")],
            initialfile=f"租赁合同_{rid}_{name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        )
        if not fp:
            return

        try:
            with open(fp, "w", encoding="utf-8") as f:
                f.write(contract_text)
            messagebox.showinfo("成功", f"合同已导出\n{fp}")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {e}")

    def _navigate_to_record(self, rid):
        """从AI助手导航到指定记录"""
        rec = self._find_record(rid)
        if rec:
            # 清除搜索过滤以显示完整列表
            self.search_var.set("")
            self.status_var.set("全部")
            self._shown = list(self._all)
            target_idx = next((i for i, r in enumerate(self._shown) if r.get("id") == rid), -1)
            if target_idx >= 0:
                self._current_page = target_idx // self._page_size
            self._render_tree()
            # 选中对应记录并显示详情
            for item in self.tree.get_children():
                if self.tree.item(item)["values"][0] == rid:
                    self.tree.selection_set(item)
                    self.tree.focus(item)
                    self.tree.see(item)
                    break
            self._show_detail_panel(rec)
        else:
            messagebox.showinfo("提示", f"未找到记录 {rid}")

    def _export_contract(self, rec):
        """导出租赁合同为文本文件"""
        from tkinter import filedialog
        rid = rec.get("id", "")
        renter = rec.get("renter", {})
        lease = rec.get("lease_info", {})
        name = renter.get("name", "")
        phone = renter.get("phone", "")
        id_card = renter.get("id_card", "")
        address = renter.get("address", "")
        start_date = lease.get("start_date", "")
        end_date = lease.get("end_date", "")
        monthly_rent = float(lease.get("monthly_rent", 0) or 0)
        total_rent = float(lease.get("total_rent", 0) or 0)
        deposit = float(lease.get("deposit", 0) or 0)
        quantity = int(rec.get("quantity", 1) or 1)
        status = rec.get("status", "")
        paid_amount = float(rec.get("paid_amount", 0) or 0)
        hardware = rec.get("hardware", {})
        hardware_summary = self.dm.summarize_hardware(rec)

        contract_text = f"""========================================
        速维电脑租赁合同
========================================

合同编号: {rid}
签订日期: {start_date}

【出租方】
名称: 速维电脑租赁
联系电话: 13800138000
地址: XX市XX区XX路XX号

【承租方】
姓名: {name}
联系电话: {phone}
身份证号: {id_card}
联系地址: {address}

【租赁物品清单】
数量: {quantity} 套
{hardware_summary}

【租赁期限】
起租日期: {start_date}
到期日期: {end_date}

【租金及押金】
月租金: ¥{monthly_rent:.2f}
总租金: ¥{total_rent:.2f}
押金: ¥{deposit:.2f}
已付金额: ¥{paid_amount:.2f}
未付金额: ¥{max(0, total_rent - paid_amount):.2f}

【合同状态】
当前状态: {status}

【备注】
{rec.get("notes", "无")}

========================================
出租方签字: ______________    日期: ____________
承租方签字: ______________    日期: ____________
========================================
"""

        fp = filedialog.asksaveasfilename(
            title="导出租赁合同",
            defaultextension=".txt",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")],
            initialfile=f"租赁合同_{rid}_{name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        )
        if not fp:
            return

        try:
            with open(fp, "w", encoding="utf-8") as f:
                f.write(contract_text)
            messagebox.showinfo("成功", f"合同已导出\n{fp}")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {e}")

